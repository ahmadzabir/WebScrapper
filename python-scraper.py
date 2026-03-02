import asyncio
import aiohttp
import pandas as pd
import re
import os
import sys
import time
import csv
import zipfile
from html import unescape, escape
import streamlit as st
from datetime import datetime
from io import BytesIO, StringIO
import json
import random
import threading
from collections import deque
from dataclasses import dataclass, field
from typing import Optional, Dict, List
import hashlib
import sqlite3
import tempfile

# Windows: stdout/stderr UTF-8 so print() never raises 'charmap' codec errors
if sys.platform == "win32":
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# Phase 3: Extraction pipeline imports
try:
    from bs4 import BeautifulSoup
    BEAUTIFULSOUP_AVAILABLE = True
except ImportError:
    BEAUTIFULSOUP_AVAILABLE = False

try:
    import trafilatura
    TRAFILATURA_AVAILABLE = True
except ImportError:
    TRAFILATURA_AVAILABLE = False

# AI imports (optional - only if API keys provided)
try:
    from openai import AsyncOpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

# -------------------------
# Utilities
# -------------------------

def _get_system_ram_gb() -> float:
    """Return total system RAM in GB, or 0 if unknown."""
    try:
        import psutil
        return psutil.virtual_memory().total / (1024 ** 3)
    except Exception:
        return 0.0


def get_system_tier() -> str:
    """
    Detect system capability for auto-tuning. Used so 300k runs don't crash on low-RAM
    machines and run at full force on capable hardware. Target: 300k in <6h on i5 + 8GB.
    Returns: 'low' (<10GB RAM or ≤4 CPUs), 'medium' (10–20GB or 5–8 CPUs), 'high' (20GB+ or 8+ CPUs).
    """
    try:
        import psutil
        ram_gb = psutil.virtual_memory().total / (1024 ** 3)
        cpus = psutil.cpu_count(logical=True) or os.cpu_count() or 2
    except Exception:
        ram_gb = 0.0
        cpus = os.cpu_count() or 2
    if ram_gb > 0:
        # 32GB+ or 8+ logical CPUs (e.g. 10th gen 6-core) = high
        if ram_gb >= 20 or cpus >= 8:
            return "high"
        if ram_gb >= 16 or cpus >= 6:
            return "high"  # 16GB+ or 6+ cores: treat as high for better speed
        if ram_gb >= 10 or cpus >= 5:
            return "medium"
        return "low"  # <10GB RAM or ≤4 CPUs
    # No RAM info: use CPU only
    if cpus >= 6:
        return "high"
    if cpus >= 5:
        return "medium"
    return "low"


def _is_low_resource_default() -> bool:
    """Auto-detect: use low-resource mode when system tier is 'low' (limited RAM/CPU)."""
    return get_system_tier() == "low"


def _get_concurrency_max() -> int:
    """Max parallel workers: liberal formula for I/O-bound scraping. 8 cores -> 96, 16 cores -> 192."""
    try:
        n = os.cpu_count() or 2
        return min(max(20, n * 12), 250)  # 8 cores -> 96, 16 cores -> 192
    except Exception:
        return 80


def _should_use_fast_mode(total_urls: int) -> bool:
    """Auto-enable fast mode for 500+ URLs. In cloud, only use for small/medium runs (<5k) so big runs stay stable."""
    if total_urls < 500:
        return False
    if is_cloud_mode() and total_urls >= 5000:
        return False  # Cloud large run: keep gentle
    return True


def _cloud_concurrency_max() -> int:
    """Max parallel workers in cloud. Small runs (<5k) can use up to 20; large runs are capped lower in run_scraper."""
    return 20


def _cloud_concurrency_default() -> int:
    """Default workers in cloud: balanced for typical runs (e.g. hundreds to a few thousand URLs)."""
    return 12


def get_auto_run_settings(total_urls: int) -> dict:
    """
    All speed/reliability settings from URL count and environment.
    Local: auto-tunes by system tier (RAM/CPU) so 300k runs full force on fast machines
    and stays stable and under ~6h on modest hardware (e.g. i5 + 8GB). Cloud = gentle.
    """
    cloud = is_cloud_mode()
    tier = get_system_tier() if not cloud else "low"
    if cloud:
        if total_urls < 5000:
            concurrency = 20
        elif total_urls < 20000:
            concurrency = 12
        elif total_urls < 100000:
            concurrency = 8
        else:
            concurrency = 4  # 100k–300k: very low for stability
        timeout = 45
        rows_per_file = 1000 if total_urls >= 5000 else 2000
        if total_urls >= 200000:
            rows_per_file = 1000  # smaller chunks for 200k+ to reduce memory
    else:
        # Local: tier-based so fast machines run full force (32GB + 10th gen = high)
        max_workers = _get_concurrency_max()
        if total_urls < 5000:
            # Small runs: use hardware fully (96–120 workers on 8+ cores)
            concurrency = min(120, max_workers) if tier == "high" else min(80, max_workers)
        elif total_urls < 20000:
            concurrency = min(96, max_workers) if tier == "high" else min(64, max_workers)
        elif total_urls < 100000:
            concurrency = min(64, max_workers) if tier == "high" else min(48, max_workers)
        elif total_urls < 200000:
            concurrency = min(24, max_workers) if tier == "high" else min(16, max_workers)
        else:
            # 200k–300k: high = 12, medium = 8, low = 6 (target <6h on 8GB)
            if tier == "high":
                concurrency = min(12, max_workers)
            elif tier == "medium":
                concurrency = min(8, max_workers)
            else:
                concurrency = min(6, max_workers)
        # High-tier: longer timeout so slow sites don't time out (60s for 32GB/10th gen)
        timeout = 60 if tier == "high" else (45 if tier == "medium" else 30)
        rows_per_file = 2000
        if total_urls >= 200000:
            rows_per_file = 1000  # smaller chunks for 200k+ to reduce memory and flush often
    return {
        "concurrency": concurrency,
        "timeout": timeout,
        "retries": 3,
        "depth": 3,
        "rows_per_file": rows_per_file,
        "max_chars": 10000,
    }


# -------------------------
# Phase 0: Measurement Infrastructure
# -------------------------

@dataclass
class ScrapeResult:
    """Structured result object for tracking scrape outcomes."""
    url: str
    result_status: str = "unknown"  # success, failed, partial
    http_status: Optional[int] = None
    exception_type: Optional[str] = None
    stage_failed: Optional[str] = None  # connect, read, parse, decode
    html_bytes: int = 0
    cleaned_chars: int = 0
    detected_js_shell: bool = False
    detected_block_page: bool = False
    retries_used: int = 0
    total_seconds: float = 0.0
    content_type: Optional[str] = None
    encoding_used: Optional[str] = None
    error_message: Optional[str] = None
    sample_html: Optional[str] = None  # First 2000 chars for debugging
    extracted_method: str = "unknown"  # direct, cache, playwright, jsonld, head_only
    
    def to_dict(self) -> dict:
        return {
            'url': self.url,
            'result_status': self.result_status,
            'http_status': self.http_status,
            'exception_type': self.exception_type,
            'stage_failed': self.stage_failed,
            'html_bytes': self.html_bytes,
            'cleaned_chars': self.cleaned_chars,
            'detected_js_shell': self.detected_js_shell,
            'detected_block_page': self.detected_block_page,
            'retries_used': self.retries_used,
            'total_seconds': round(self.total_seconds, 2),
            'content_type': self.content_type,
            'encoding_used': self.encoding_used,
            'error_message': self.error_message,
            'extracted_method': self.extracted_method,
        }


class FailureTracker:
    """Track and categorize failures for analysis."""
    
    def __init__(self):
        self.results: List[ScrapeResult] = []
        self.failure_buckets: Dict[str, List[ScrapeResult]] = {}
        self.html_samples: Dict[str, str] = {}  # Store sample HTML per bucket
        
    def add_result(self, result: ScrapeResult):
        self.results.append(result)
        
        if result.result_status != "success":
            bucket_key = self._get_bucket_key(result)
            if bucket_key not in self.failure_buckets:
                self.failure_buckets[bucket_key] = []
            self.failure_buckets[bucket_key].append(result)
            
            # Store one sample HTML per bucket (first 2000 chars)
            if bucket_key not in self.html_samples and result.sample_html:
                self.html_samples[bucket_key] = result.sample_html[:2000]
    
    def _get_bucket_key(self, result: ScrapeResult) -> str:
        """Generate failure bucket key from result."""
        if result.exception_type:
            return f"{result.stage_failed}:{result.exception_type}"
        elif result.http_status:
            return f"http_{result.http_status}"
        elif result.detected_block_page:
            return "detected:block_page"
        elif result.cleaned_chars < 200:
            return "content:insufficient"
        else:
            return "unknown"
    
    def get_top_failures(self, n: int = 10) -> List[tuple]:
        """Get top N failure buckets with counts."""
        sorted_buckets = sorted(
            self.failure_buckets.items(),
            key=lambda x: len(x[1]),
            reverse=True
        )
        return [(key, len(results)) for key, results in sorted_buckets[:n]]
    
    def get_failure_csv(self) -> str:
        """Generate CSV of all failures."""
        import csv
        import io
        output = io.StringIO()
        if self.results:
            fieldnames = ['url', 'result_status', 'http_status', 'exception_type', 
                         'stage_failed', 'html_bytes', 'cleaned_chars', 'error_message']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for result in self.results:
                if result.result_status != "success":
                    writer.writerow(result.to_dict())
        return output.getvalue()
    
    def get_summary_stats(self) -> dict:
        """Get summary statistics."""
        total = len(self.results)
        successes = sum(1 for r in self.results if r.result_status == "success")
        partials = sum(1 for r in self.results if r.result_status == "partial")
        failures = total - successes - partials
        
        return {
            'total': total,
            'successes': successes,
            'partials': partials,
            'failures': failures,
            'success_rate': round(successes / total * 100, 1) if total > 0 else 0,
            'top_failures': self.get_top_failures(5),
            'html_samples': self.html_samples
        }


# Global failure tracker (per session)
_failure_tracker = None

def get_failure_tracker() -> FailureTracker:
    global _failure_tracker
    if _failure_tracker is None:
        _failure_tracker = FailureTracker()
    return _failure_tracker

def reset_failure_tracker():
    global _failure_tracker
    _failure_tracker = FailureTracker()


# -------------------------
# Large Dataset Support: SQLite Persistence for 150k+ leads
# -------------------------

class ScrapeDatabase:
    """SQLite database for persisting scrape results - handles 150k+ leads."""
    
    def __init__(self, db_path: Optional[str] = None):
        if db_path is None:
            # Use temp directory for cloud compatibility
            db_path = os.path.join(tempfile.gettempdir(), f"scrape_{int(time.time())}.db")
        self.db_path = db_path
        self._init_db()
    
    def _init_db(self):
        """Initialize database tables."""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS scrape_results (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    url TEXT NOT NULL,
                    row_index INTEGER,
                    scraped_text TEXT,
                    company_summary TEXT,
                    email_copy TEXT,
                    result_status TEXT DEFAULT 'pending',
                    http_status INTEGER,
                    error_message TEXT,
                    extracted_method TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            conn.execute("""
                CREATE TABLE IF NOT EXISTS scrape_jobs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_id TEXT UNIQUE NOT NULL,
                    total_urls INTEGER,
                    processed_urls INTEGER DEFAULT 0,
                    failed_urls INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'running',
                    config_json TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_results_url ON scrape_results(url)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_results_status ON scrape_results(result_status)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_results_row ON scrape_results(row_index)
            """)
    
    def insert_result(self, url: str, row_index: int, scraped_text: str = "", 
                     company_summary: str = "", email_copy: str = "",
                     result_status: str = "pending", http_status: Optional[int] = None,
                     error_message: str = "", extracted_method: str = ""):
        """Insert a scrape result."""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                INSERT OR REPLACE INTO scrape_results 
                (url, row_index, scraped_text, company_summary, email_copy, 
                 result_status, http_status, error_message, extracted_method)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (url, row_index, scraped_text, company_summary, email_copy,
                 result_status, http_status, error_message, extracted_method))
    
    def get_results(self, status: Optional[str] = None, limit: int = 1000, offset: int = 0) -> pd.DataFrame:
        """Get results as DataFrame."""
        with sqlite3.connect(self.db_path) as conn:
            if status:
                query = """
                    SELECT url, row_index, scraped_text, company_summary, email_copy,
                           result_status, http_status, error_message, extracted_method
                    FROM scrape_results 
                    WHERE result_status = ?
                    ORDER BY row_index
                    LIMIT ? OFFSET ?
                """
                return pd.read_sql_query(query, conn, params=(status, limit, offset))
            else:
                query = """
                    SELECT url, row_index, scraped_text, company_summary, email_copy,
                           result_status, http_status, error_message, extracted_method
                    FROM scrape_results 
                    ORDER BY row_index
                    LIMIT ? OFFSET ?
                """
                return pd.read_sql_query(query, conn, params=(limit, offset))
    
    def get_count(self, status: Optional[str] = None) -> int:
        """Get count of results."""
        with sqlite3.connect(self.db_path) as conn:
            if status:
                cursor = conn.execute("SELECT COUNT(*) FROM scrape_results WHERE result_status = ?", (status,))
            else:
                cursor = conn.execute("SELECT COUNT(*) FROM scrape_results")
            return cursor.fetchone()[0]
    
    def get_stats(self) -> dict:
        """Get scrape statistics."""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN result_status = 'success' THEN 1 ELSE 0 END) as successes,
                    SUM(CASE WHEN result_status = 'partial' THEN 1 ELSE 0 END) as partials,
                    SUM(CASE WHEN result_status = 'failed' THEN 1 ELSE 0 END) as failures
                FROM scrape_results
            """)
            row = cursor.fetchone()
            return {
                'total': row[0],
                'successes': row[1] or 0,
                'partials': row[2] or 0,
                'failures': row[3] or 0
            }
    
    def export_to_csv(self, output, chunk_size: int = 5000):
        """Export all results to CSV in chunks (memory efficient).
        
        Args:
            output: Either a file path (str) or a file-like object (StringIO/BytesIO)
            chunk_size: Number of rows per chunk
        """
        offset = 0
        first_chunk = True
        csv_buffer = None
        
        # Check if output is a file path or file-like object
        is_file_path = isinstance(output, str)
        
        while True:
            df = self.get_results(limit=chunk_size, offset=offset)
            if df.empty:
                break
            
            if first_chunk:
                if is_file_path:
                    df.to_csv(output, index=False, encoding='utf-8-sig')
                else:
                    # For StringIO, accumulate in buffer
                    csv_buffer = df.to_csv(index=False, encoding='utf-8-sig')
                first_chunk = False
            else:
                if is_file_path:
                    df.to_csv(output, index=False, encoding='utf-8-sig', mode='a', header=False)
                else:
                    # For StringIO, append without header
                    csv_buffer += df.to_csv(index=False, encoding='utf-8-sig', header=False)
            
            offset += chunk_size
        
        # If using StringIO, write accumulated buffer
        if not is_file_path and csv_buffer:
            output.write(csv_buffer)
    
    def close(self):
        """Clean up database file."""
        try:
            if os.path.exists(self.db_path):
                os.remove(self.db_path)
        except:
            pass


# Global database instance (per session)
_scrape_db = None

def get_scrape_db() -> Optional[ScrapeDatabase]:
    global _scrape_db
    return _scrape_db

def init_scrape_db() -> ScrapeDatabase:
    global _scrape_db
    _scrape_db = ScrapeDatabase()
    return _scrape_db

def reset_scrape_db():
    global _scrape_db
    if _scrape_db:
        _scrape_db.close()
    _scrape_db = None


# -------------------------
# Phase 1: Cloud Mode & Compression Settings
# -------------------------

def is_cloud_mode() -> bool:
    """Detect if running in cloud/limited environment (Streamlit Cloud, Kubernetes, etc.). Local runs return False so they use full CPU/network."""
    cloud_indicators = [
        'STREAMLIT_SHARING_MODE',
        'STREAMLIT_CLOUD',
        'KUBERNETES_SERVICE_HOST',
    ]
    for indicator in cloud_indicators:
        if os.environ.get(indicator):
            return True
    # No cloud env vars = assume local; use more workers and resources
    return False


def get_safe_headers() -> dict:
    """Get headers that work reliably in cloud mode."""
    base_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        # Phase 1: Disable Brotli to avoid edge cases
        "Accept-Encoding": "gzip, deflate",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    }
    return base_headers


def is_html_content(content_type: str) -> bool:
    """Check if content type indicates HTML."""
    if not content_type:
        return False
    content_type = content_type.lower()
    html_indicators = ['text/html', 'application/xhtml', 'application/xhtml+xml']
    return any(ind in content_type for ind in html_indicators)


def is_likely_error_page(html: str) -> tuple[bool, str]:
    """Detect if HTML is likely an error page."""
    if not html or len(html) < 100:
        return True, "too_short"
    
    html_lower = html.lower()
    
    # Check for block pages
    block_indicators = [
        "access denied", "forbidden", "blocked", "captcha", 
        "challenge", "one moment", "checking your browser",
        "please wait", "rate limit", "too many requests"
    ]
    for indicator in block_indicators:
        if indicator in html_lower:
            return True, f"block_page:{indicator}"
    
    # Check for error HTTP status codes in content
    if "404 not found" in html_lower or "error 404" in html_lower:
        return True, "error:404"
    
    if "500 internal server" in html_lower or "error 500" in html_lower:
        return True, "error:500"
    
    return False, ""


# -------------------------
# Phase 3: 3-Layer Extraction Pipeline
# -------------------------

class ExtractionResult:
    """Result from extraction pipeline."""
    def __init__(self, text: str, method: str, confidence: float, metadata: dict = None):
        self.text = text
        self.method = method
        self.confidence = confidence
        self.metadata = metadata or {}
        self.length = len(text) if text else 0


def extract_head_content(html: str) -> dict:
    """Extract content from head section (Layer 3)."""
    result = {}
    if not html:
        return result
    
    try:
        # Parse with BeautifulSoup if available
        if BEAUTIFULSOUP_AVAILABLE:
            soup = BeautifulSoup(html, 'html.parser')
        else:
            # Fallback to regex
            result['title'] = re.search(r'<title[^>]*>(.*?)</title>', html, re.IGNORECASE | re.DOTALL)
            if result['title']:
                result['title'] = result['title'].group(1).strip()
            return result
        
        # Extract title
        title = soup.find('title')
        if title:
            result['title'] = title.get_text().strip()
        
        # Extract meta description
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        if meta_desc:
            result['description'] = meta_desc.get('content', '').strip()
        
        # Extract OpenGraph tags
        og_tags = {}
        for tag in soup.find_all('meta', property=re.compile(r'^og:')):
            prop = tag.get('property', '').replace('og:', '')
            content = tag.get('content', '').strip()
            if prop and content:
                og_tags[prop] = content
        result['opengraph'] = og_tags
        
        # Extract JSON-LD
        jsonld_scripts = []
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                data = json.loads(script.string)
                jsonld_scripts.append(data)
            except:
                pass
        result['jsonld'] = jsonld_scripts
        
        # Check for Next.js data
        next_data = soup.find('script', id='__NEXT_DATA__')
        if next_data:
            try:
                result['nextjs'] = json.loads(next_data.string)
            except:
                pass
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


def layer1_dom_extraction(html: str) -> ExtractionResult:
    """Layer 1: Basic DOM text extraction with BeautifulSoup."""
    if not BEAUTIFULSOUP_AVAILABLE or not html:
        # Fallback to regex-based cleanup
        cleaned = cleanup_html(html) if html else ""
        return ExtractionResult(
            text=cleaned,
            method="regex_fallback",
            confidence=0.3 if cleaned else 0.0
        )
    
    try:
        soup = BeautifulSoup(html, 'html.parser')
        
        # Remove unwanted elements
        for tag in soup(['script', 'style', 'nav', 'footer', 'header', 'aside', 
                         'noscript', 'iframe', 'svg', 'canvas', 'form', 'button']):
            tag.decompose()
        
        # Remove elements with noise classes/ids
        noise_patterns = re.compile(r'(cookie|ad|popup|modal|banner|promo|newsletter|subscribe)')
        for tag in soup.find_all(class_=noise_patterns):
            tag.decompose()
        for tag in soup.find_all(id=noise_patterns):
            tag.decompose()
        
        # Extract text from meaningful elements
        text_parts = []
        for tag in soup.find_all(['h1', 'h2', 'h3', 'h4', 'p', 'li', 'article', 'section']):
            text = tag.get_text().strip()
            if text and len(text) > 10:
                text_parts.append(text)
        
        # Also get body text
        body = soup.find('body')
        if body:
            full_text = body.get_text(separator='\n', strip=True)
            # Deduplicate
            seen = set()
            unique_parts = []
            for part in text_parts:
                if part not in seen:
                    seen.add(part)
                    unique_parts.append(part)
            
            result_text = '\n\n'.join(unique_parts) if unique_parts else full_text
        else:
            result_text = '\n\n'.join(text_parts)
        
        # Clean up excessive whitespace
        result_text = re.sub(r'\n{3,}', '\n\n', result_text)
        result_text = re.sub(r' {2,}', ' ', result_text)
        
        confidence = min(0.6, 0.3 + len(result_text) / 5000)
        
        return ExtractionResult(
            text=result_text,
            method="dom_extraction",
            confidence=confidence
        )
        
    except Exception as e:
        return ExtractionResult(
            text="",
            method="dom_extraction_failed",
            confidence=0.0,
            metadata={'error': str(e)}
        )


def layer2_readability_extraction(html: str, url: str = "") -> ExtractionResult:
    """Layer 2: Readability-style extraction."""
    if TRAFILATURA_AVAILABLE:
        try:
            # Use trafilatura for content extraction
            text = trafilatura.extract(
                html,
                url=url,
                include_comments=False,
                include_tables=False,
                no_fallback=True
            )
            
            if text and len(text) > 100:
                return ExtractionResult(
                    text=text,
                    method="readability_trafilatura",
                    confidence=0.85
                )
        except Exception as e:
            pass
    
    # Fallback to layer 1
    return layer1_dom_extraction(html)


def layer3_head_signals(html: str) -> ExtractionResult:
    """Layer 3: Extract from head section when body is thin."""
    head_data = extract_head_content(html)
    
    parts = []
    
    if 'title' in head_data:
        parts.append(f"Title: {head_data['title']}")
    
    if 'description' in head_data:
        parts.append(f"Description: {head_data['description']}")
    
    if 'opengraph' in head_data:
        og = head_data['opengraph']
        if 'title' in og:
            parts.append(f"OG Title: {og['title']}")
        if 'description' in og:
            parts.append(f"OG Description: {og['description']}")
    
    # Extract from JSON-LD
    if 'jsonld' in head_data:
        for item in head_data['jsonld']:
            if isinstance(item, dict):
                if 'description' in item:
                    parts.append(f"JSON-LD: {item['description']}")
                if 'name' in item:
                    parts.append(f"Name: {item['name']}")
    
    result_text = '\n\n'.join(parts)
    
    confidence = min(0.5, 0.2 + len(result_text) / 1000)
    
    return ExtractionResult(
        text=result_text,
        method="head_signals",
        confidence=confidence,
        metadata=head_data
    )


def phase4_js_shell_detection(html: str) -> tuple[bool, dict]:
    """Phase 4: Detect JavaScript shell patterns and extract embedded data."""
    if not html:
        return False, {}
    
    html_lower = html.lower()
    indicators = {
        'script_ratio': 0,
        'visible_text_ratio': 0,
        'has_framework_markers': False,
        'has_next_data': False,
        'has_jsonld': False
    }
    
    # Check for framework markers
    framework_markers = ['react', 'vue', 'angular', 'next.js', 'nuxt', 'gatsby']
    indicators['has_framework_markers'] = any(marker in html_lower for marker in framework_markers)
    
    # Check for Next.js data
    indicators['has_next_data'] = '__NEXT_DATA__' in html
    
    # Check for JSON-LD
    indicators['has_jsonld'] = 'application/ld+json' in html
    
    # Calculate script ratio
    script_tags = len(re.findall(r'<script', html_lower))
    total_len = len(html)
    indicators['script_ratio'] = script_tags / max(total_len / 1000, 1)
    
    # Check visible text
    visible_text = re.sub(r'<[^>]+>', '', html)
    visible_text = re.sub(r'\s+', ' ', visible_text).strip()
    indicators['visible_text_ratio'] = len(visible_text) / max(total_len, 1)
    
    # JS shell: high script ratio, low visible text
    is_js_shell = (
        indicators['script_ratio'] > 0.1 and 
        indicators['visible_text_ratio'] < 0.1
    ) or indicators['has_framework_markers']
    
    return is_js_shell, indicators


def extract_from_js_payloads(html: str) -> ExtractionResult:
    """Phase 4: Extract content from JavaScript payloads."""
    results = []
    
    # Try Next.js data
    next_match = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if next_match:
        try:
            data = json.loads(next_match.group(1))
            # Extract from props
            if 'props' in data and 'pageProps' in data['props']:
                page_props = data['props']['pageProps']
                if isinstance(page_props, dict):
                    # Look for content fields
                    for key in ['content', 'description', 'body', 'text', 'data']:
                        if key in page_props and isinstance(page_props[key], str):
                            results.append(page_props[key])
        except:
            pass
    
    # Try to extract from JSON-LD
    jsonld_pattern = re.compile(r'<script type="application/ld\+json"[^>]*>(.*?)</script>', re.DOTALL)
    for match in jsonld_pattern.finditer(html):
        try:
            data = json.loads(match.group(1))
            if isinstance(data, dict):
                for key in ['description', 'articleBody', 'text']:
                    if key in data and isinstance(data[key], str):
                        results.append(data[key])
        except:
            pass
    
    if results:
        combined = '\n\n'.join(results)
        return ExtractionResult(
            text=combined,
            method="js_payload_extraction",
            confidence=0.6
        )
    
    return ExtractionResult(
        text="",
        method="js_payload_extraction_failed",
        confidence=0.0
    )


def run_extraction_pipeline(html: str, url: str = "", result_tracker: Optional[ScrapeResult] = None) -> tuple[str, str, float]:
    """
    Run the full 3-layer extraction pipeline.
    Returns: (extracted_text, method_used, confidence_score)
    """
    if not html:
        if result_tracker:
            result_tracker.stage_failed = "parse"
            result_tracker.exception_type = "empty_html"
        return "", "no_content", 0.0
    
    # Phase 4: Check for JS shell
    is_js_shell, js_indicators = phase4_js_shell_detection(html)
    
    if is_js_shell:
        # Try to extract from JS payloads first
        js_result = extract_from_js_payloads(html)
        if js_result.length > 200:
            if result_tracker:
                result_tracker.extracted_method = "js_payload"
                result_tracker.detected_js_shell = True
            return js_result.text, js_result.method, js_result.confidence
    
    # Layer 2: Readability extraction
    readability_result = layer2_readability_extraction(html, url)
    if readability_result.length >= 200:
        if result_tracker:
            result_tracker.extracted_method = readability_result.method
        return readability_result.text, readability_result.method, readability_result.confidence
    
    # Layer 1: DOM extraction (fallback)
    dom_result = layer1_dom_extraction(html)
    if dom_result.length >= 200:
        if result_tracker:
            result_tracker.extracted_method = dom_result.method
        return dom_result.text, dom_result.method, dom_result.confidence
    
    # Layer 3: Head signals (when body is thin)
    head_result = layer3_head_signals(html)
    if head_result.length > 50:
        if result_tracker:
            result_tracker.extracted_method = head_result.method
        # Mark as partial success
        combined = head_result.text + "\n\n" + dom_result.text if dom_result.text else head_result.text
        return combined, "head_plus_partial", 0.4
    
    # If all layers failed, return whatever we got from DOM
    if result_tracker:
        result_tracker.extracted_method = "dom_fallback"
        result_tracker.stage_failed = "content_insufficient"
    
    return dom_result.text, "extraction_failed", min(0.2, dom_result.confidence)


# -------------------------
# Phase 5: Circuit Breaker & Smart Retries
# -------------------------

class DomainCircuitBreaker:
    """Circuit breaker for per-domain rate limiting and backoff."""
    
    def __init__(self):
        self.domain_states: Dict[str, dict] = {}
        self.domain_identities: Dict[str, dict] = {}  # Phase 2: stable identity
    
    def get_domain_state(self, domain: str) -> dict:
        if domain not in self.domain_states:
            self.domain_states[domain] = {
                'consecutive_errors': 0,
                'last_error_time': 0,
                'cooldown_until': 0,
                'health_score': 100,
                'requests_this_minute': 0,
                'last_request_time': 0
            }
        return self.domain_states[domain]
    
    def get_identity(self, domain: str) -> dict:
        """Phase 2: Get stable identity for domain."""
        if domain not in self.domain_identities:
            # Create consistent identity
            import random
            user_agents = [
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0"
            ]
            accept_langs = ["en-US,en;q=0.9", "en-GB,en;q=0.9", "en-US,en;q=0.9,fr;q=0.8"]
            viewports = [(1920, 1080), (1366, 768), (1440, 900)]
            
            identity = {
                'user_agent': random.choice(user_agents),
                'accept_language': random.choice(accept_langs),
                'viewport': random.choice(viewports),
                'accept': "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                'created_at': time.time()
            }
            self.domain_identities[domain] = identity
        return self.domain_identities[domain]
    
    def can_request(self, domain: str) -> tuple[bool, float]:
        """Check if request is allowed and return wait time if not."""
        state = self.get_domain_state(domain)
        now = time.time()
        
        # Check cooldown
        if now < state['cooldown_until']:
            wait = state['cooldown_until'] - now
            return False, wait
        
        # Check per-minute rate
        if now - state['last_request_time'] > 60:
            state['requests_this_minute'] = 0
        
        if state['requests_this_minute'] >= 10:  # Max 10 per minute per domain
            return False, 6.0  # Wait 6 seconds
        
        # Health-based rate limiting
        if state['health_score'] < 50:
            return False, 30.0  # Cool down for 30s
        
        return True, 0.0
    
    def record_result(self, domain: str, success: bool, http_status: int = None):
        """Record request result and update health."""
        state = self.get_domain_state(domain)
        now = time.time()
        
        state['requests_this_minute'] += 1
        state['last_request_time'] = now
        
        if success:
            state['consecutive_errors'] = 0
            state['health_score'] = min(100, state['health_score'] + 10)
        else:
            state['consecutive_errors'] += 1
            state['health_score'] = max(0, state['health_score'] - 20)
            
            # Activate cooldown on repeated errors
            if state['consecutive_errors'] >= 3:
                if http_status in [429, 403]:
                    state['cooldown_until'] = now + 60  # 1 minute cooldown
                else:
                    state['cooldown_until'] = now + 30  # 30s cooldown
    
    def get_headers_for_domain(self, domain: str, cloud_mode: bool = True) -> dict:
        """Get stable headers for domain."""
        identity = self.get_identity(domain)
        
        if cloud_mode:
            # Safe headers for cloud
            return {
                "User-Agent": identity['user_agent'],
                "Accept": identity['accept'],
                "Accept-Language": identity['accept_language'],
                "Accept-Encoding": "gzip, deflate",  # No Brotli
                "DNT": "1",
                "Connection": "keep-alive",
            }
        else:
            # Full headers for local
            return {
                "User-Agent": identity['user_agent'],
                "Accept": identity['accept'],
                "Accept-Language": identity['accept_language'],
                "Accept-Encoding": "gzip, deflate, br",
                "DNT": "1",
                "Connection": "keep-alive",
            }


# Global circuit breaker instance
_circuit_breaker = DomainCircuitBreaker()

def get_circuit_breaker() -> DomainCircuitBreaker:
    return _circuit_breaker


# -------------------------
# Phase 6: Intelligent Link Discovery
# -------------------------

LINK_PRIORITY_KEYWORDS = [
    ('about', 10),
    ('services', 9),
    ('service', 9),
    ('products', 8),
    ('product', 8),
    ('solutions', 8),
    ('solution', 8),
    ('pricing', 7),
    ('features', 7),
    ('feature', 7),
    ('case-studies', 6),
    ('casestudies', 6),
    ('case-study', 6),
    ('contact', 5),
    ('team', 4),
    ('company', 4),
    ('industries', 3),
    ('industry', 3),
]

def score_link(url: str, link_text: str = "", base_domain: str = "") -> int:
    """Score a link based on priority keywords and path characteristics."""
    score = 0
    url_lower = url.lower()
    text_lower = link_text.lower()
    
    # Check for priority keywords
    for keyword, weight in LINK_PRIORITY_KEYWORDS:
        if keyword in url_lower or keyword in text_lower:
            score += weight
    
    # Prefer shorter paths
    path_depth = url.count('/')
    if path_depth <= 1:
        score += 3
    elif path_depth <= 2:
        score += 1
    
    # Prefer same domain
    if base_domain and base_domain in url:
        score += 2
    
    # Penalize assets
    asset_extensions = ['.pdf', '.jpg', '.png', '.gif', '.css', '.js', '.zip']
    if any(url_lower.endswith(ext) for ext in asset_extensions):
        score -= 10
    
    return max(0, score)


def discover_links_intelligent(html: str, base_url: str, max_links: int = 5) -> List[str]:
    """Phase 6: Score and return top N links."""
    if not html or not BEAUTIFULSOUP_AVAILABLE:
        # Fallback to simple regex
        return extract_links_simple(html, base_url)[:max_links]
    
    try:
        from urllib.parse import urljoin, urlparse
        
        soup = BeautifulSoup(html, 'html.parser')
        base_domain = urlparse(base_url).netloc
        
        scored_links = []
        seen = set()
        
        for a_tag in soup.find_all('a', href=True):
            href = a_tag['href']
            link_text = a_tag.get_text().strip()
            
            # Resolve relative URLs
            full_url = urljoin(base_url, href)
            parsed = urlparse(full_url)
            
            # Skip non-HTTP, anchors, and external domains
            if not parsed.scheme.startswith('http'):
                continue
            if href.startswith('#') or href.startswith('mailto:'):
                continue
            if parsed.netloc != base_domain and not parsed.netloc.endswith('.' + base_domain):
                continue
            
            # Normalize and dedupe
            normalized = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
            if normalized in seen:
                continue
            seen.add(normalized)
            
            # Score
            link_score = score_link(normalized, link_text, base_domain)
            scored_links.append((normalized, link_score))
        
        # Sort by score descending
        scored_links.sort(key=lambda x: x[1], reverse=True)
        
        # Return top N
        return [url for url, score in scored_links[:max_links]]
        
    except Exception as e:
        # Fallback
        return extract_links_simple(html, base_url)[:max_links]


def extract_links_simple(html: str, base_url: str) -> List[str]:
    """Simple regex-based link extraction as fallback."""
    from urllib.parse import urljoin, urlparse
    
    if not html:
        return []
    
    base_domain = urlparse(base_url).netloc
    links = []
    seen = set()
    
    # Simple href extraction
    matches = re.findall(r'href=["\']([^"\']+)["\']', html, re.IGNORECASE)
    
    for href in matches:
        try:
            full_url = urljoin(base_url, href)
            parsed = urlparse(full_url)
            
            if parsed.netloc == base_domain or parsed.netloc.endswith('.' + base_domain):
                normalized = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
                if normalized not in seen:
                    seen.add(normalized)
                    links.append(normalized)
        except:
            continue
    
    return links


# -------------------------
# Phase 8: Ethics & Robots.txt
# -------------------------

def check_robots_txt(domain: str, user_agent: str = "*") -> tuple[bool, float]:
    """
    Phase 8: Optional robots.txt check.
    Returns: (is_allowed, crawl_delay)
    """
    try:
        robots_url = f"https://{domain}/robots.txt"
        import urllib.request
        
        req = urllib.request.Request(
            robots_url,
            headers={'User-Agent': 'WebScraperBot/1.0'},
            timeout=5
        )
        
        with urllib.request.urlopen(req) as response:
            robots_content = response.read().decode('utf-8', errors='ignore')
        
        # Simple parsing
        lines = robots_content.split('\n')
        current_ua = None
        crawl_delay = 0
        
        for line in lines:
            line = line.strip().lower()
            if line.startswith('user-agent:'):
                current_ua = line.split(':', 1)[1].strip()
            elif line.startswith('crawl-delay:') and (current_ua == '*' or current_ua in user_agent.lower()):
                try:
                    crawl_delay = float(line.split(':', 1)[1].strip())
                except:
                    pass
            elif line.startswith('disallow:') and (current_ua == '*' or current_ua in user_agent.lower()):
                # Simplified check - actual path matching would be more complex
                pass
        
        return True, crawl_delay
        
    except Exception as e:
        # If robots.txt fails, assume allowed
        return True, 0


def normalize_url(url: str) -> str:
    """
    Normalize URL by adding https:// if protocol is missing.
    More lenient - handles various URL formats and edge cases.
    """
    if not url:
        return url
    
    url = str(url).strip()
    
    # Remove any leading/trailing whitespace, quotes, or brackets
    url = url.strip(' "\'[]()')
    
    # Skip empty URLs
    if not url:
        return url
    
    # If URL doesn't start with http:// or https://, add https://
    if not url.startswith(("http://", "https://")):
        # Remove any protocol-like prefixes that might be malformed
        url = url.lstrip("://")
        
        # If it starts with www., add https://
        if url.startswith("www."):
            url = "https://" + url
        else:
            # Otherwise, add https://
            url = "https://" + url
    
    # Clean up any double slashes (except after http:// or https://)
    url = url.replace(":///", "://")
    url = url.replace(":////", "://")
    
    # Basic validation - ensure it looks like a URL
    # But be lenient - let the actual HTTP request determine validity
    if "://" not in url:
        # Still try to fix it
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
    
    return url


def _truncate_at_word(text: str, max_len: int) -> str:
    """Truncate text at word boundary to avoid cutting mid-word."""
    if not text or len(text) <= max_len:
        return text
    cut = text[:max_len]
    last_space = cut.rfind(' ')
    if last_space > max_len * 0.7:  # Only break at word if we keep most of the content
        return cut[:last_space].rstrip()
    return cut.rstrip()


def process_keywords(keyword_string: str) -> list:
    """
    Process keywords string into a clean list.
    Handles: commas with/without spaces, case-insensitive, hyphens, special chars.
    
    Examples:
    - "about,service,product" -> ["about", "service", "product"]
    - "about, service, product" -> ["about", "service", "product"]
    - "About, Service-Product" -> ["about", "service-product"]
    """
    if not keyword_string or not keyword_string.strip():
        return []
    
    # Split by comma and clean each keyword
    keywords = []
    for kw in keyword_string.split(","):
        kw = kw.strip()  # Remove leading/trailing spaces
        if kw:  # Only add non-empty keywords
            keywords.append(kw.lower())  # Normalize to lowercase for matching
    
    return keywords


def match_keyword_in_url(keyword: str, url: str) -> bool:
    """
    Check if keyword matches in URL (case-insensitive, handles hyphens/underscores).
    
    Features:
    - Case-insensitive matching
    - Handles hyphens, underscores, and spaces in URLs
    - Partial word matching (e.g., "about" matches "/about-us")
    - Handles URL encoding
    
    Examples:
    - keyword="about" matches "/about", "/about-us", "/About-Us", "/ABOUT"
    - keyword="service" matches "/services", "/our-services", "/SERVICE"
    """
    if not keyword or not url:
        return False
    
    # Normalize both to lowercase for comparison
    keyword_lower = keyword.lower().strip()
    url_lower = url.lower()
    
    # Replace common separators with the keyword separator for better matching
    # This helps match "about-us" with keyword "about"
    url_normalized = url_lower.replace("-", "").replace("_", "").replace("/", "/")
    
    # Check if keyword appears in URL
    # Also check with common separators
    keyword_variations = [
        keyword_lower,
        keyword_lower.replace("-", ""),
        keyword_lower.replace("_", ""),
    ]
    
    for kw_var in keyword_variations:
        if kw_var in url_lower or f"/{kw_var}" in url_lower or f"{kw_var}/" in url_lower:
            return True
    
    return False


# -------------------------
# Checkpoint for crash recovery and resume
# -------------------------

CHECKPOINT_FILENAME = "checkpoint.json"


def save_checkpoint(checkpoint_data: dict, checkpoint_path: str) -> None:
    """Save checkpoint to disk. Atomic write + retries. Never raise — data must never be lost."""
    if not checkpoint_path or not checkpoint_data:
        return
    data = checkpoint_data.copy()
    try:
        data["completed_urls"] = list(data.get("completed_urls", []))
    except Exception:
        data["completed_urls"] = []
    data["last_updated"] = datetime.now().isoformat()
    tmp_path = checkpoint_path + ".tmp"
    for attempt in range(3):
        try:
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
                f.flush()
                if hasattr(f, 'fileno'):
                    try:
                        os.fsync(f.fileno())
                    except Exception:
                        pass
            os.replace(tmp_path, checkpoint_path)
            return
        except Exception as e:
            if attempt < 2:
                time.sleep(0.2 * (attempt + 1))
            else:
                try:
                    print(f"⚠️ Checkpoint save failed after 3 attempts: {e}")
                except Exception:
                    pass
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass


def load_checkpoint(checkpoint_path: str) -> dict | None:
    """Load checkpoint from disk. Returns None if file doesn't exist or is invalid."""
    if not checkpoint_path or not os.path.exists(checkpoint_path):
        return None
    try:
        with open(checkpoint_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return None
        completed = data.get("completed_urls", [])
        if isinstance(completed, list):
            data["completed_urls"] = set(completed)
        elif isinstance(completed, set):
            data["completed_urls"] = completed
        else:
            data["completed_urls"] = set()
        return data
    except Exception as e:
        print(f"⚠️ Checkpoint load failed: {e}")
        return None


def get_checkpoint_path(output_dir: str) -> str:
    """Return the checkpoint file path for a run."""
    return os.path.join(output_dir, CHECKPOINT_FILENAME)


def get_actual_completed_from_files(run_path: str) -> tuple[int, set]:
    """Count actual rows and extract URLs from output_part_*.csv files. Returns (row_count, set of urls)."""
    total_rows = 0
    urls = set()
    if not run_path or not os.path.isdir(run_path):
        return 0, urls
    try:
        for f in sorted(os.listdir(run_path)):
            if f.startswith("output_part_") and f.endswith(".csv") and "combined" not in f.lower():
                path = os.path.join(run_path, f)
                try:
                    df = pd.read_csv(path, encoding="utf-8-sig", usecols=[0], nrows=None)  # Website col
                    if len(df.columns) > 0:
                        col = df.columns[0]
                        for v in df[col].dropna().astype(str).str.strip():
                            if v and v.strip():
                                urls.add(v)
                        total_rows += len(df)
                except Exception:
                    pass
    except Exception:
        pass
    return total_rows, urls


def get_latest_run_with_results():
    """
    Find the most recent run folder in outputs/ that has result files (ZIP or output_part_*.csv).
    Returns dict with output_dir, run_folder, zip_path, zip_name, csv_files, excel_files, total
    or None if no such run exists. Keeps download section visible after refresh when session state is cleared.
    """
    outputs_dir = "outputs"
    if not os.path.isdir(outputs_dir):
        return None
    for run_folder in sorted(os.listdir(outputs_dir), reverse=True):
        run_path = os.path.join(outputs_dir, run_folder)
        if not os.path.isdir(run_path):
            continue
        zip_name = f"{run_folder}.zip"
        zip_path = os.path.join(run_path, zip_name)
        has_zip = os.path.isfile(zip_path) and os.path.getsize(zip_path) > 0
        try:
            files = os.listdir(run_path)
        except Exception:
            continue
        csv_files = [f for f in files if f.endswith(".csv")]
        excel_files = [f for f in files if f.endswith(".xlsx")]
        has_parts = any(f.startswith("output_part_") and f.endswith(".csv") for f in files)
        if not (has_zip or has_parts or csv_files or excel_files):
            continue
        ck = load_checkpoint(get_checkpoint_path(run_path))
        total = len(ck.get("urls", [])) if ck else 0
        if total == 0:
            actual_rows, _ = get_actual_completed_from_files(run_path)
            total = actual_rows if actual_rows > 0 else 1
        return {
            "output_dir": run_path,
            "run_folder": run_folder,
            "zip_path": zip_path if has_zip else None,
            "zip_name": zip_name,
            "csv_files": csv_files,
            "excel_files": excel_files,
            "total": total,
        }
    return None


def reconcile_checkpoint_with_files(output_dir: str) -> int:
    """If checkpoint has more completed_urls than actual rows in files, fix checkpoint. Returns actual count."""
    ck_path = get_checkpoint_path(output_dir)
    ck = load_checkpoint(ck_path)
    if not ck:
        return 0
    actual_rows, actual_urls = get_actual_completed_from_files(output_dir)
    completed = ck.get("completed_urls", set())
    if len(completed) > actual_rows and actual_rows >= 0:
        ck["completed_urls"] = actual_urls
        part_files = [f for f in os.listdir(output_dir) if f.startswith("output_part_") and (f.endswith(".csv") or f.endswith(".xlsx"))]
        part_nums = []
        for f in part_files:
            try:
                # output_part_1.csv -> 1
                n = int(f.replace("output_part_", "").replace(".csv", "").replace(".xlsx", ""))
                part_nums.append(n)
            except ValueError:
                pass
        ck["last_part"] = max(part_nums) if part_nums else 0
        save_checkpoint(ck, ck_path)
        return actual_rows
    return len(completed)


# -------------------------
# AI Summary Generation
# -------------------------

def _col_to_placeholder(col_name: str) -> str:
    """Convert column name to placeholder key: 'Company Name' -> 'company_name'."""
    if not col_name:
        return ""
    key = re.sub(r'[^\w\s]', '', str(col_name).strip())
    key = re.sub(r'\s+', '_', key).lower()
    return key or "col"


def _format_var_value(v) -> str:
    """Format any value for safe insertion into prompts. Handles None, NaN, control chars. Preserves newlines/formatting."""
    if v is None:
        return ""
    try:
        import math
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "n/a"):
        return ""
    s = s.replace("\x00", "")
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", s)
    return s.strip()


def _replace_prompt_variables(template: str, data: dict) -> str:
    """
    Replace {key} and {{key}} placeholders with actual values. Both brace styles get the full value.
    Longest keys first so e.g. {company_name} is not broken by {company}. Formatting preserved.
    """
    if template is None or not template:
        return "" if template is None else template
    keys_sorted = sorted(data.keys(), key=len, reverse=True)
    out = template
    for k in keys_sorted:
        if k == "scraped_content":
            sval = str(data.get(k, ""))[:15000].replace("\x00", "")
        else:
            sval = _format_var_value(data.get(k, "")) or ""
        out = out.replace("{{" + k + "}}", sval)
        out = out.replace("{" + k + "}", sval)
    return out


def _get_variable_definitions() -> list[tuple[str, str, str]]:
    """
    Return list of (display_label, placeholder, key) for prompt variables.
    Driven 100% by the uploaded sheet: URL column + all other columns + company_name + scraped_content.
    """
    result = []
    csv_cfg = st.session_state.get("csv_config") or {}
    url_col = csv_cfg.get("url_column")
    df = csv_cfg.get("df_preview")
    has_headers = csv_cfg.get("has_headers", True)
    col_list = list(df.columns) if df is not None and has_headers else ([f"Column {i+1}" for i in range(len(df.columns))] if df is not None else [])

    result.append(("URL", "{url}", "url"))
    result.append(("Company name", "{company_name}", "company_name"))
    for col in col_list:
        if col == url_col:
            continue
        key = _col_to_placeholder(col)
        if not key:
            continue
        result.append((str(col), "{" + key + "}", key))
    result.append(("Scraped content", "{scraped_content}", "scraped_content"))
    return result


def _get_available_placeholders() -> list[str]:
    """Return list of placeholder keys available for prompts."""
    return [r[2] for r in _get_variable_definitions()]


@st.dialog("Sample prompt — filled with one lead's data")
def _sample_prompt_dialog(prompt_key: str, sample_row_index: int = 0):
    """Show the prompt with variables replaced by one lead's data. No persistent state — only opens when button is clicked."""
    prompt_text = st.session_state.get(prompt_key, "") or ""
    sample = _get_lead_sample_from_row(sample_row_index)
    sample["scraped_content"] = EXAMPLE_SCRAPED_CONTENT
    if prompt_key == "master_prompt":
        filled = build_company_summary_prompt(prompt_text, sample, EXAMPLE_SCRAPED_CONTENT)
        filled = filled + COMPANY_SUMMARY_FINAL_REMINDER
    else:
        filled = build_email_copy_prompt(prompt_text, sample, EXAMPLE_SCRAPED_CONTENT)
    filled = filled.replace("{scraped_content}", EXAMPLE_SCRAPED_CONTENT).replace("{{scraped_content}}", EXAMPLE_SCRAPED_CONTENT)
    st.caption("This is the exact prompt (including formatting) that would be sent to the AI for this lead.")
    st.text_area("Filled prompt", value=filled, height=400, disabled=True, key="sample_dialog_ta", label_visibility="collapsed")
    if st.button("Close", key="sample_dialog_close"):
        st.rerun()


# Example scraped content shown in previews so users see how {scraped_content} will look
EXAMPLE_SCRAPED_CONTENT = """About Us

Example Company helps businesses grow through digital transformation. Founded in 2020, we provide consulting and software solutions across industries.

Our Services
- Strategy and planning: We work with leadership to define roadmaps and priorities.
- Implementation and integration: Our team delivers projects on time and on budget.
- Training and support: We ensure your people get the most from new systems.

Why choose us? We combine deep industry experience with a practical, no-nonsense approach. Our clients include mid-market and enterprise organizations in technology, healthcare, and financial services.

Contact us at info@example.com or visit our website for more information. We typically respond within one business day."""


def _get_sample_lead_data_for_preview() -> dict:
    """Get sample values from first CSV row for prompt preview. Returns dict of key -> value."""
    return _get_lead_sample_from_row(0)


def _get_lead_sample_from_row(row_index: int | None = None) -> dict:
    """Get sample values from one CSV row. If row_index is None, picks a random row. scraped_content uses EXAMPLE_SCRAPED_CONTENT so preview shows realistic example."""
    out = {"url": "https://example.com", "company_name": "Example Company", "scraped_content": EXAMPLE_SCRAPED_CONTENT}
    csv_cfg = st.session_state.get("csv_config") or {}
    df = csv_cfg.get("df_preview")
    if df is None or df.empty:
        return out
    has_headers = csv_cfg.get("has_headers", True)
    n = len(df)
    if n == 0:
        return out
    idx = row_index if row_index is not None and 0 <= row_index < n else random.randint(0, n - 1)
    row = df.iloc[idx]
    url_col = csv_cfg.get("url_column")
    if url_col:
        try:
            if has_headers and url_col in df.columns:
                v = row[url_col]
            else:
                ci = int(str(url_col).replace("Column ", "")) - 1
                v = row.iloc[ci]
            u = "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v).strip()
            if u:
                out["url"] = u
                out["company_name"] = u.replace("https://", "").replace("http://", "").split("/")[0]
        except Exception:
            pass
    lead_cols = csv_cfg.get("lead_data_columns") or {}
    for key, col_name in lead_cols.items():
        try:
            if has_headers and col_name in df.columns:
                val = row[col_name]
            else:
                ci = int(str(col_name).replace("Column ", "")) - 1
                val = row.iloc[ci]
            v = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val).strip()
            out[key] = v
        except Exception:
            out[key] = ""
    return out


# Standard output column names (3 or 4 when email copy enabled)
EXCEL_COLS_3 = ["Website", "ScrapedText", "CompanySummary"]
EXCEL_COLS_4 = ["Website", "ScrapedText", "CompanySummary", "EmailCopy"]


def _clean_excel_value(val) -> str:
    """Single source of truth: clean any value for Excel (prevents corruption, formula injection)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    val = str(val).strip()
    if not val:
        return ""
    val = val.replace('\x00', '')
    val = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', val)
    val = ''.join(c for c in val if (ord(c) == 0x9 or ord(c) == 0xA or ord(c) == 0xD or
                                      0x20 <= ord(c) <= 0xD7FF or 0xE000 <= ord(c) <= 0xFFFD))
    val = re.sub(r'\s+', ' ', val.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')).strip()
    val = val[:32767]
    if val and val[0] in ('=', '+', '-', '@'):
        val = "'" + val
    return val


def _write_excel_sheet(ws, cols: list, df: "pd.DataFrame", sheet_title: str = "Scraped Data") -> None:
    """Write DataFrame to Excel sheet with proper formatting, column widths, frozen header."""
    ws.title = sheet_title
    ws.append(cols)
    
    # Write data rows
    for idx in range(len(df)):
        row = df.iloc[idx]
        row_data = [_clean_excel_value(row[c] if c in df.columns else "") for c in cols]
        row_num = idx + 2
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = str(val) if val is not None else ""
            # Enable text wrapping for long content
            cell.alignment = cell.alignment.copy(wrapText=True)
    
    # Column widths - smart defaults based on content type
    from openpyxl.utils import get_column_letter
    
    for i, col_name in enumerate(cols, start=1):
        col_letter = get_column_letter(i)
        
        # Determine width based on column name
        col_lower = col_name.lower()
        if 'url' in col_lower or 'website' in col_lower:
            width = 50  # URLs need moderate width
        elif 'scraped' in col_lower or 'content' in col_lower or 'summary' in col_lower or 'email' in col_lower:
            width = 100  # Long text columns need more space
        elif 'name' in col_lower or 'title' in col_lower or 'company' in col_lower:
            width = 40  # Names
        elif 'id' in col_lower or 'phone' in col_lower:
            width = 20  # IDs, phone numbers
        elif any(date_word in col_lower for date_word in ['date', 'time', 'created', 'updated']):
            width = 25  # Dates
        else:
            width = 35  # Default for other columns
        
        ws.column_dimensions[col_letter].width = min(width, 120)
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _normalize_scrape_error_for_display(msg: str) -> str:
    """Convert emoji-prefixed scrape errors so narrow columns show clear messages.
    Uses ': ' (colon+space) for split to avoid breaking on URLs like http://."""
    if not msg or not isinstance(msg, str):
        return ""
    s = str(msg).strip()
    if not s.startswith("❌"):
        return s
    rest = s[1:].lstrip()
    if "Timeout" in rest or "timeout" in rest:
        m = re.search(r'\((\d+)s?\)', rest)
        secs = m.group(1) if m else "?"
        return f"Scrape timeout ({secs}s exceeded)"
    # Split only on ': ' (colon+space) to avoid breaking URLs (http://, https://)
    if ": " in rest:
        _, detail = rest.split(": ", 1)
        return f"Scrape error: {detail.strip()}"
    # If rest looks like just a URL (no descriptive message), use generic message
    if rest.startswith(("http://", "https://", "www.")) and len(rest) < 120:
        return "Scrape error: Site unreachable or blocked"
    return f"Scrape error: {rest}"


def _normalize_ai_error_for_display(msg: str, prefix: str = "Summary") -> str:
    """Convert emoji-prefixed errors to clearer format for CSV/Excel.
    Uses 'Summary failed:' prefix so narrow columns never truncate to 'AI Sum'."""
    if not msg or not isinstance(msg, str):
        return ""
    s = str(msg).strip()
    if not s.startswith("❌"):
        return s
    rest = s[1:].lstrip()
    if ":" in rest:
        lead, detail = rest.split(":", 1)
        detail = detail.strip()
        if "skipped" in lead.lower():
            return f"{prefix} skipped: {detail}" if detail else f"{prefix} skipped"
        return f"{prefix} failed: {detail}" if detail else f"{prefix} failed"
    return f"{prefix} failed: {rest}" if rest else f"{prefix} failed"


# Appended to company summary prompt when sending to AI (so sample preview matches exactly)
COMPANY_SUMMARY_FINAL_REMINDER = """

===========================================================
CRITICAL OUTPUT FORMAT REQUIREMENTS:
===========================================================
1. Use EXACT section headers: ===SUMMARY===, ===FACTS===, ===HYPOTHESES===
2. Do NOT use markdown formatting (**, __, #). Use plain text only.
3. Evidence quotes must be COMPLETE sentences, not truncated with "..."
4. Do NOT truncate words mid-word. Always end at sentence boundaries.
5. Keep sections separate. Do NOT mix facts into hypotheses or vice versa.
6. Facts must be grounded (from webcopy). Hypotheses must be inferred (marked with (obs)).
7. Use consistent format:
   - Fact: [statement]
   - Evidence Quote: "[complete quote from webcopy]"
   - Source: [page name or Webcopy]
   - Hypothesis: [inference] (obs)
   - Signal: "[specific wording from webcopy]"
   - Commercial implication: [why it matters]
   - Confidence: High/Medium/Low
8. Do NOT include inline labels like "Confidence: Medium Signal: ..." - use structured format above.
===========================================================
"""


def build_company_summary_prompt(base_prompt: str, lead_data: dict | None, scraped_content: str | None) -> str:
    """
    Build the complete prompt for AI company summary generation.
    Scraped content is inserted only where {scraped_content} or {{scraped_content}} appears;
    it is never injected elsewhere, so no double-adding.
    
    Args:
        base_prompt: User-provided base prompt (with placeholders)
        lead_data: Dictionary with lead information (URL, company name, etc.)
        scraped_content: Scraped website content (replaces {scraped_content} in prompt)
    
    Returns:
        Complete prompt string ready for AI
    """
    # Default prompt structure - Commercial analysis focused
    default_base = """You are Hypothesis Bot… an advanced commercial analysis agent.

INPUT
You will receive ONLY one input: raw website copy scraped from a company's website (may include multiple pages).

CORE JOB
Turn messy webcopy into:
1) a sharp company intelligence SUMMARY
2) a clean set of FACTS grounded in the text
3) commercially relevant HYPOTHESES inferred from signals, gaps, tone, and structure

You are not writing outreach. You are building the intelligence layer that enables personalization later.

NON NEGOTIABLE RULES
1) Do not invent facts. If it is not explicitly supported by the webcopy, it is not a fact.
2) Facts must be short and must include Evidence Quote… an exact snippet from the webcopy.
3) Hypotheses must be explicitly labeled as hypotheses and must include:
   • Signal… the specific wording or structural clue that triggered the inference
   • Commercial implication… why it matters in a sales or growth context
   • Confidence: High, Medium, or Low
4) Never mention any external tools or data sources (Apollo, LinkedIn, Crunchbase, funding, headcount, etc.). You only have webcopy.
5) If the company name is unclear, write "Company: Not explicitly stated" and proceed.
6) Avoid cringe adjectives like great, amazing, innovative. Be surgical.
7) Avoid criticism. Frame gaps neutrally as "signals" or "absence suggests".
8) Keep it concise. No fluff. No extra sections.

ANALYSIS GUIDELINES
When extracting facts, prioritize:
• What they do (offer categories, deliverables)
• Who they serve (industries, segments, buyer language)
• How they sell (engagement models, pricing mentions, process)
• Proof (case studies, client names, testimonials, quantified claims)
• Differentiators (positioning phrases, guarantees, compliance, security)
• Operational signals (hiring, support hours, global language, locations)
• Technology signals (stacks, platforms, integrations) only if stated

When generating hypotheses, prioritize:
• Likely buying triggers (growth, hiring, new initiatives, modernization)
• Likely pain points (capacity, speed, differentiation, trust, compliance, delivery)
• Likely maturity level (specialist vs generalist, product vs services)
• Likely stakeholder priorities (risk reduction, outcomes, speed, cost certainty)
• Contradictions or gaps between claims and proof
Each hypothesis must connect to a commercial implication.

LEAD INFORMATION:
- Website URL: {url}
- Company Name: {company_name}

WEBSITE CONTENT (ONLY SOURCE OF INFORMATION):
{scraped_content}

OUTPUT FORMAT (STRICT - NO MARKDOWN, NO TRUNCATION)
Return exactly these 3 sections with EXACT headers:

===SUMMARY===
Write 3 to 5 sentences in one paragraph describing:
• what the company appears to do
• who it appears to serve
• how it positions itself
• one notable proof element (only if present)
Mark inferences with (obs).
Do NOT use markdown (**, __, #). Use plain text only.

===FACTS===
Provide 10 to 18 facts.
Format each EXACTLY as:
Fact: [statement]
Evidence Quote: "[COMPLETE quote from webcopy - do NOT truncate with ...]"
Source: [Page name if present, otherwise "Webcopy"]

CRITICAL: Evidence quotes must be COMPLETE sentences from the webcopy. Do NOT truncate mid-word or mid-sentence.

===HYPOTHESES===
Provide 6 to 12 hypotheses.
Format each EXACTLY as:
Hypothesis: [inference statement] (obs)
Signal: "[specific wording or structural clue from webcopy]"
Commercial implication: [why it matters in sales/growth context]
Confidence: High or Medium or Low

CRITICAL: 
- Do NOT use markdown formatting (**, __, #)
- Do NOT truncate words mid-word
- Do NOT mix facts into hypotheses section
- Keep facts grounded (from webcopy), hypotheses inferred (from signals)"""
    
    # Use user prompt if provided, otherwise use default
    prompt_template = base_prompt if base_prompt.strip() else default_base
    
    lead_data = dict(lead_data or {})
    # Ensure company_name (default from URL if missing)
    company_name = _format_var_value(lead_data.get('company_name')) or ''
    if not company_name or company_name == 'Unknown':
        url_str = lead_data.get('url', '')
        company_name = url_str.replace('https://', '').replace('http://', '').split('/')[0] if url_str else 'Unknown'
    lead_data['company_name'] = company_name
    lead_data['url'] = _format_var_value(lead_data.get('url')) or 'N/A'
    lead_data['scraped_content'] = (scraped_content or "")[:15000]

    def _humanize_key(k):
        return k.replace('_', ' ').title()

    # LEAD INFORMATION block: only URL and company name by default (no auto-dump of all CSV columns).
    # Other variables (employees, city, etc.) only appear if the user puts {employees}, {city}, etc. in their prompt.
    lead_info_parts = []
    for k in ("url", "company_name"):
        v = _format_var_value(lead_data.get(k))
        if v:
            lead_info_parts.append(f"- {_humanize_key(k)}: {v}")
    lead_info_section = "LEAD INFORMATION:\n" + "\n".join(lead_info_parts) if lead_info_parts else "LEAD INFORMATION:\n(no additional lead data)"

    # Replace all placeholders including {scraped_content}; scraped content is never injected elsewhere
    prompt = _replace_prompt_variables(prompt_template, lead_data)

    if 'LEAD INFORMATION:' in prompt:
        prompt = re.sub(
            r'LEAD INFORMATION:.*?(?=WEBSITE CONTENT:|$)',
            lead_info_section + '\n\n',
            prompt,
            flags=re.DOTALL
        )
    else:
        if 'WEBSITE CONTENT:' in prompt:
            prompt = prompt.replace('WEBSITE CONTENT:', lead_info_section + '\n\nWEBSITE CONTENT:')
        else:
            prompt = lead_info_section + '\n\n' + prompt

    return prompt


async def fetch_openai_models(api_key: str) -> list:
    """Fetch ALL available OpenAI models from API in real-time."""
    if not OPENAI_AVAILABLE:
        return ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"]  # Fallback
    
    try:
        client = AsyncOpenAI(api_key=api_key)
        models_response = await client.models.list()
        
        # Get ALL GPT models (chat completion capable)
        all_models = [model.id for model in models_response.data]
        
        # Filter for GPT models - Get ALL GPT models that can be used for chat completions
        # Only exclude: embeddings, deprecated models, and base/completion models (ada, babbage, curie, davinci)
        # Include ALL GPT models including variants like gpt-4o, gpt-4-turbo, gpt-3.5-turbo, gpt-4o-mini, etc.
        # IMPORTANT: Include ALL GPT models, don't filter by 'chat' or 'instruct' as those are valid chat models
        gpt_models = []
        skip_patterns = ['embedding', 'ada', 'babbage', 'curie', 'davinci', 'deprecated']
        
        for model_id in all_models:
            model_lower = model_id.lower()
            # Include if it's a GPT model
            if 'gpt' in model_lower:
                # Skip only if it's clearly an embedding or base model (not chat models)
                # Include ALL GPT models including instruct, turbo, o, etc.
                if not any(skip in model_lower for skip in skip_patterns):
                    gpt_models.append(model_id)
        
        # Debug: Log how many models we found (remove in production)
        # print(f"Found {len(gpt_models)} GPT models: {gpt_models[:10]}...")
        
        # Remove duplicates and sort
        unique_models = sorted(list(set(gpt_models)), reverse=True)
        
        # Prioritize common/recommended models at the top
        priority_models = ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-4", "gpt-3.5-turbo", "gpt-3.5-turbo-16k"]
        ordered = []
        
        # Add priority models first (if they exist)
        for pm in priority_models:
            if pm in unique_models:
                ordered.append(pm)
        
        # Add ALL remaining models (not just filtered ones)
        for model in unique_models:
            if model not in ordered:
                ordered.append(model)
        
        return ordered if ordered else priority_models
    except Exception as e:
        # Return fallback models on error
        return ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"]


async def fetch_gemini_models(api_key: str) -> list:
    """Fetch available Gemini models from API."""
    if not GEMINI_AVAILABLE:
        return ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-pro"]  # Fallback
    
    try:
        genai.configure(api_key=api_key)
        # Gemini models are typically fixed, but we can verify
        # Common Gemini models
        common_models = ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-pro", "gemini-1.5-flash-latest", "gemini-1.5-pro-latest"]
        
        # Try to get list of models (if API supports it)
        try:
            # Note: Gemini API doesn't have a direct models.list() endpoint
            # So we'll use common models and verify they work
            return common_models
        except:
            return common_models
    except Exception as e:
        return ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-pro"]


async def fetch_openrouter_models(api_key: str = "") -> list:
    """Fetch available models from OpenRouter API. No API key needed for listing."""
    fallback = [
        "openai/gpt-4o-mini",
        "openai/gpt-4o",
        "google/gemini-2.0-flash-exp:free",
        "anthropic/claude-3.5-sonnet",
        "google/gemini-1.5-flash",
    ]
    try:
        async with aiohttp.ClientSession() as session:
            headers = {}
            if api_key and api_key.strip():
                headers["Authorization"] = f"Bearer {api_key.strip()}"
            async with session.get("https://openrouter.ai/api/v1/models", headers=headers) as resp:
                if resp.status != 200:
                    return fallback
                data = await resp.json()
        models_data = data.get("data", [])
        if not models_data:
            return fallback
        model_ids = [m.get("id") for m in models_data if m.get("id")]
        if not model_ids:
            return fallback
        priority = [
            "openai/gpt-4o-mini",
            "openai/gpt-4o",
            "google/gemini-2.0-flash-exp:free",
            "google/gemini-2.0-flash",
            "anthropic/claude-3.5-sonnet",
            "google/gemini-1.5-flash",
        ]
        ordered = [m for m in priority if m in model_ids]
        for m in model_ids:
            if m not in ordered:
                ordered.append(m)
        return ordered
    except Exception:
        return fallback


# System message for email copy so the model doesn't use summary-format or refuse B2B outreach
EMAIL_COPY_SYSTEM = "You are a professional B2B copywriter. The user will provide company information and ask for a short personalized business email. Output only the email body (no subject line, no explanations). Write in a professional, specific tone. Do not refuse; write the requested email."


async def generate_openrouter_summary(api_key: str, model: str, prompt: str, max_retries: int = 5, status_callback=None, system_message: str | None = None) -> str:
    """Generate company summary using OpenRouter API (OpenAI-compatible). Optional system_message for email copy."""
    if not OPENAI_AVAILABLE:
        return "❌ OpenAI library not installed. Install with: pip install openai"
    
    system_content = system_message if system_message else "You are Hypothesis Bot, an advanced commercial analysis agent. Your job is to turn messy webcopy into structured intelligence: ===SUMMARY===, ===FACTS=== (with complete evidence quotes), and ===HYPOTHESES=== (with signals, commercial implications, and confidence levels). CRITICAL: Use EXACT section headers (===SUMMARY===, ===FACTS===, ===HYPOTHESES===). Do NOT use markdown formatting (**, __, #). Evidence quotes must be COMPLETE sentences, not truncated. Do NOT truncate words mid-word. Never invent facts. Only use what's explicitly in the webcopy. Be surgical and concise. Follow the output format STRICTLY."
    
    client = AsyncOpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key.strip(),
    )
    
    for attempt in range(max_retries):
        try:
            response = await client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_content},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.2,
                max_tokens=4000,
                top_p=0.9,
            )
            content = (response.choices[0].message.content or "").strip()
            if not content:
                return "❌ OpenRouter returned empty response; try again or another model."
            return content
        except Exception as e:
            error_str = str(e).lower()
            error_msg = str(e)
            
            is_rate_limit = (
                "rate limit" in error_str or
                "429" in error_str or
                "quota" in error_str or
                "too many requests" in error_str or
                "credit" in error_str
            )
            
            if is_rate_limit:
                wait_time = min(2 ** attempt, 60)
                if status_callback:
                    status_callback(f"⏳ Rate limit hit. Waiting {wait_time}s before retry {attempt + 1}/{max_retries}...")
                await asyncio.sleep(wait_time)
                continue
            
            is_retryable = (
                "timeout" in error_str or
                "connection" in error_str or
                "503" in error_str or
                "502" in error_str or
                "500" in error_str
            )
            
            if is_retryable and attempt < max_retries - 1:
                wait_time = min(2 ** attempt, 30)
                if status_callback:
                    status_callback(f"⚠️ API error. Retrying in {wait_time}s ({attempt + 1}/{max_retries})...")
                await asyncio.sleep(wait_time)
                continue
            
            if attempt == max_retries - 1:
                if is_rate_limit:
                    return f"❌ OpenRouter Rate Limit: Exceeded after {max_retries} retries."
                return f"❌ OpenRouter API Error: {error_msg}"
            
            await asyncio.sleep(1 + attempt)
    
    return "❌ OpenRouter API failed after retries"


async def generate_openai_summary(api_key: str, model: str, prompt: str, max_retries: int = 5, status_callback=None, system_message: str | None = None) -> str:
    """Generate company summary using OpenAI API with automatic rate limit handling. Optional system_message for email copy."""
    if not OPENAI_AVAILABLE:
        return "❌ OpenAI library not installed. Install with: pip install openai"
    
    system_content = system_message if system_message else "You are Hypothesis Bot, an advanced commercial analysis agent. Your job is to turn messy webcopy into structured intelligence: ===SUMMARY===, ===FACTS=== (with complete evidence quotes), and ===HYPOTHESES=== (with signals, commercial implications, and confidence levels). CRITICAL: Use EXACT section headers (===SUMMARY===, ===FACTS===, ===HYPOTHESES===). Do NOT use markdown formatting (**, __, #). Evidence quotes must be COMPLETE sentences, not truncated. Do NOT truncate words mid-word. Never invent facts. Only use what's explicitly in the webcopy. Be surgical and concise. Follow the output format STRICTLY."
    
    client = AsyncOpenAI(api_key=api_key, timeout=120.0)
    
    for attempt in range(max_retries):
        try:
            response = await client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_content},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.2,  # Lower temperature for more consistent formatting
                max_tokens=4000,  # Increased for complete quotes
                top_p=0.9  # More focused responses
            )
            content = (response.choices[0].message.content or "").strip()
            if not content:
                return "❌ OpenAI returned empty response; try again or another model."
            return content
        except Exception as e:
            error_str = str(e).lower()
            error_msg = str(e)
            
            # Check for rate limit errors
            is_rate_limit = (
                "rate limit" in error_str or
                "429" in error_str or
                "quota" in error_str or
                "too many requests" in error_str or
                "requests per minute" in error_str or
                "rate_limit_exceeded" in error_str
            )
            
            if is_rate_limit:
                # Calculate exponential backoff: 2^attempt seconds, max 60 seconds
                wait_time = min(2 ** attempt, 60)
                
                if status_callback:
                    status_callback(f"⏳ Rate limit hit. Waiting {wait_time}s before retry {attempt + 1}/{max_retries}...")
                
                await asyncio.sleep(wait_time)
                continue  # Retry after waiting
            
            # Check for other retryable errors
            is_retryable = (
                "timeout" in error_str or
                "connection" in error_str or
                "503" in error_str or
                "502" in error_str or
                "500" in error_str
            )
            
            if is_retryable and attempt < max_retries - 1:
                wait_time = min(2 ** attempt, 30)
                
                if status_callback:
                    status_callback(f"⚠️ API error. Retrying in {wait_time}s ({attempt + 1}/{max_retries})...")
                
                await asyncio.sleep(wait_time)
                continue
            
            # Final attempt or non-retryable error
            if attempt == max_retries - 1:
                if is_rate_limit:
                    return f"❌ OpenAI Rate Limit: Exceeded after {max_retries} retries. Please wait a few minutes and try again, or upgrade your API plan."
                return f"❌ OpenAI API Error: {error_msg}"
            
            # Default wait for other errors
            await asyncio.sleep(1 + attempt)


async def generate_gemini_summary(api_key: str, model: str, prompt: str, max_retries: int = 5, status_callback=None, system_message: str | None = None) -> str:
    """Generate company summary using Google Gemini API with automatic rate limit handling. Optional system_message for email copy."""
    if not GEMINI_AVAILABLE:
        return "❌ Gemini library not installed. Install with: pip install google-generativeai"
    
    try:
        genai.configure(api_key=api_key)
    except Exception as e:
        return f"❌ Gemini API Configuration Error: {str(e)}"
    
    # When system_message is provided (e.g. email copy), prepend so model follows role
    content_prompt = (system_message + "\n\n---\n\n" + prompt) if system_message else prompt
    
    for attempt in range(max_retries):
        try:
            # Configure generation parameters for analytical but factual output
            generation_config = {
                "temperature": 0.2,  # Lower temperature for more consistent formatting
                "max_output_tokens": 4000,  # Increased for complete quotes
                "top_p": 0.9,
                "top_k": 40,  # More focused
            }
            
            gemini_model = genai.GenerativeModel(
                model,
                generation_config=generation_config
            )
            # Run in thread pool to avoid blocking
            loop = asyncio.get_event_loop()
            response = await loop.run_in_executor(
                None,
                lambda p=content_prompt: gemini_model.generate_content(p)
            )
            if hasattr(response, 'text'):
                return response.text.strip()
            else:
                return f"❌ Gemini API returned unexpected response format"
        except Exception as e:
            error_str = str(e).lower()
            error_msg = str(e)
            
            # Check for rate limit errors
            is_rate_limit = (
                "rate limit" in error_str or
                "429" in error_str or
                "quota" in error_str or
                "resource_exhausted" in error_str or
                "too many requests" in error_str or
                "requests per minute" in error_str or
                "per_minute" in error_str
            )
            
            if is_rate_limit:
                # Calculate exponential backoff: 2^attempt seconds, max 60 seconds
                wait_time = min(2 ** attempt, 60)
                
                if status_callback:
                    status_callback(f"⏳ Rate limit hit. Waiting {wait_time}s before retry {attempt + 1}/{max_retries}...")
                
                await asyncio.sleep(wait_time)
                continue  # Retry after waiting
            
            # Check for other retryable errors
            is_retryable = (
                "timeout" in error_str or
                "connection" in error_str or
                "503" in error_str or
                "502" in error_str or
                "500" in error_str or
                "unavailable" in error_str
            )
            
            if is_retryable and attempt < max_retries - 1:
                wait_time = min(2 ** attempt, 30)
                
                if status_callback:
                    status_callback(f"⚠️ API error. Retrying in {wait_time}s ({attempt + 1}/{max_retries})...")
                
                await asyncio.sleep(wait_time)
                continue
            
            # Final attempt or non-retryable error
            if attempt == max_retries - 1:
                if is_rate_limit:
                    return f"❌ Gemini Rate Limit: Exceeded after {max_retries} retries. Please wait a few minutes and try again, or check your API quota."
                return f"❌ Gemini API Error: {error_msg}"
            
            # Default wait for other errors
            await asyncio.sleep(1 + attempt)
    
    return "❌ Gemini API failed after retries"


def clean_and_structure_ai_output(raw_output: str) -> str:
    """
    Post-process AI output to fix all formatting issues:
    - Remove markdown formatting
    - Enforce strict section boundaries
    - Fix quote truncation
    - Ensure sentence-aware truncation
    - Standardize section headers
    - Clean up semantic noise
    """
    if not raw_output or raw_output.startswith("❌"):
        return raw_output
    
    import re
    
    # Step 1: Remove all markdown formatting (**, __, #, etc.)
    cleaned = raw_output
    # Remove bold/italic markdown
    cleaned = re.sub(r'\*\*([^*]+)\*\*', r'\1', cleaned)  # Remove **text**
    cleaned = re.sub(r'__([^_]+)__', r'\1', cleaned)  # Remove __text__
    cleaned = re.sub(r'\*([^*]+)\*', r'\1', cleaned)  # Remove *text*
    cleaned = re.sub(r'_([^_]+)_', r'\1', cleaned)  # Remove _text_
    # Remove headers
    cleaned = re.sub(r'^#{1,6}\s+', '', cleaned, flags=re.MULTILINE)
    
    # Step 2: Normalize section headers to consistent format
    # Find and standardize section headers
    section_patterns = [
        (r'^#*\s*\*?\*?SUMMARY\*?\*?\s*#*:?\s*$', '===SUMMARY==='),
        (r'^#*\s*\*?\*?FACTS\*?\*?\s*#*:?\s*$', '===FACTS==='),
        (r'^#*\s*\*?\*?HYPOTHESES\*?\*?\s*#*:?\s*$', '===HYPOTHESES==='),
        (r'^SUMMARY\s*:?\s*$', '===SUMMARY==='),
        (r'^FACTS\s*:?\s*$', '===FACTS==='),
        (r'^HYPOTHESES\s*:?\s*$', '===HYPOTHESES==='),
    ]
    
    for pattern, replacement in section_patterns:
        cleaned = re.sub(pattern, replacement, cleaned, flags=re.MULTILINE | re.IGNORECASE)
    
    # Step 3: Split into sections
    sections = {}
    current_section = None
    current_content = []
    
    lines = cleaned.split('\n')
    for line in lines:
        line = line.strip()
        if line.startswith('===') and line.endswith('==='):
            # Save previous section
            if current_section:
                sections[current_section] = '\n'.join(current_content).strip()
            # Start new section
            current_section = line.replace('===', '').upper()
            current_content = []
        elif current_section:
            current_content.append(line)
        elif not current_section and line:
            # Content before first section - assume it's SUMMARY
            if 'SUMMARY' not in sections:
                current_section = 'SUMMARY'
                current_content = [line]
    
    # Save last section
    if current_section:
        sections[current_section] = '\n'.join(current_content).strip()
    
    # Step 4: Clean each section
    cleaned_sections = {}
    
    for section_name, content in sections.items():
        if not content:
            continue
        
        # Clean SUMMARY section
        if section_name == 'SUMMARY':
            # Remove any remaining markdown, normalize whitespace
            cleaned_content = re.sub(r'\s+', ' ', content).strip()
            # Ensure it's a paragraph (no line breaks unless intentional)
            cleaned_content = cleaned_content.replace('\n', ' ')
            cleaned_content = re.sub(r'\s+', ' ', cleaned_content)
            # Limit length to reasonable size (sentence-aware)
            if len(cleaned_content) > 1000:
                # Truncate at sentence boundary
                sentences = re.split(r'([.!?]\s+)', cleaned_content[:1100])
                if len(sentences) > 1:
                    cleaned_content = ''.join(sentences[:-1]).strip()
                else:
                    cleaned_content = cleaned_content[:1000] + '...'
            cleaned_sections[section_name] = cleaned_content
        
        # Clean FACTS section
        elif section_name == 'FACTS':
            facts = []
            # Split by "Fact:" pattern
            fact_blocks = re.split(r'(?=Fact\s*:)', content, flags=re.IGNORECASE)
            
            for block in fact_blocks:
                if not block.strip() or 'Fact:' not in block[:20]:
                    continue
                
                # Extract fact
                fact_match = re.search(r'Fact\s*:\s*(.+?)(?=Evidence Quote:|Source:|$)', block, re.DOTALL | re.IGNORECASE)
                fact_text = fact_match.group(1).strip() if fact_match else ''
                
                # Extract evidence quote (must be complete, not truncated)
                quote_match = re.search(r'Evidence Quote\s*:\s*"([^"]+)"', block, re.DOTALL | re.IGNORECASE)
                if not quote_match:
                    # Try without quotes
                    quote_match = re.search(r'Evidence Quote\s*:\s*(.+?)(?=Source:|$)', block, re.DOTALL | re.IGNORECASE)
                
                quote_text = ''
                if quote_match:
                    quote_text = quote_match.group(1).strip().strip('"\'')
                    # Remove ellipses that indicate truncation (we want complete quotes)
                    quote_text = re.sub(r'\s*\.\.\.\s*$', '', quote_text)
                    # Ensure quote is not mid-word truncated
                    if quote_text and not quote_text[-1].isalnum() and quote_text[-1] not in '.!?"\'':
                        # Might be truncated, try to find sentence end
                        pass
                
                # Extract source
                source_match = re.search(r'Source\s*:\s*(.+?)(?=Fact:|$)', block, re.DOTALL | re.IGNORECASE)
                source_text = source_match.group(1).strip() if source_match else 'Webcopy'
                
                if fact_text:
                    fact_entry = f"Fact: {fact_text}"
                    if quote_text:
                        fact_entry += f'\nEvidence Quote: "{quote_text}"'
                    fact_entry += f'\nSource: {source_text}'
                    facts.append(fact_entry)
            
            cleaned_sections[section_name] = '\n\n'.join(facts[:18])  # Limit to 18 facts
        
        # Clean HYPOTHESES section
        elif section_name == 'HYPOTHESES':
            hypotheses = []
            # Split by "Hypothesis:" pattern
            hyp_blocks = re.split(r'(?=Hypothesis\s*:)', content, flags=re.IGNORECASE)
            
            for block in hyp_blocks:
                if not block.strip() or 'Hypothesis:' not in block[:30]:
                    continue
                
                # Extract hypothesis
                hyp_match = re.search(r'Hypothesis\s*:\s*(.+?)(?=Signal:|Commercial implication:|Confidence:|$)', block, re.DOTALL | re.IGNORECASE)
                hyp_text = hyp_match.group(1).strip() if hyp_match else ''
                hyp_text = re.sub(r'\s*\(obs\)\s*', ' (obs)', hyp_text)  # Normalize (obs) marker
                
                # Extract signal
                signal_match = re.search(r'Signal\s*:\s*"([^"]+)"', block, re.DOTALL | re.IGNORECASE)
                if not signal_match:
                    signal_match = re.search(r'Signal\s*:\s*(.+?)(?=Commercial implication:|Confidence:|$)', block, re.DOTALL | re.IGNORECASE)
                signal_text = signal_match.group(1).strip().strip('"\'') if signal_match else ''
                
                # Extract commercial implication
                impl_match = re.search(r'Commercial implication\s*:\s*(.+?)(?=Confidence:|Signal:|$)', block, re.DOTALL | re.IGNORECASE)
                impl_text = impl_match.group(1).strip() if impl_match else ''
                
                # Extract confidence
                conf_match = re.search(r'Confidence\s*:\s*(High|Medium|Low)', block, re.IGNORECASE)
                conf_text = conf_match.group(1) if conf_match else 'Medium'
                
                if hyp_text:
                    hyp_entry = f"Hypothesis: {hyp_text}"
                    if signal_text:
                        hyp_entry += f'\nSignal: "{signal_text}"'
                    if impl_text:
                        hyp_entry += f'\nCommercial implication: {impl_text}'
                    hyp_entry += f'\nConfidence: {conf_text}'
                    hypotheses.append(hyp_entry)
            
            cleaned_sections[section_name] = '\n\n'.join(hypotheses[:12])  # Limit to 12 hypotheses
    
    # Step 5: Reconstruct output with strict separators
    output_parts = []
    
    if 'SUMMARY' in cleaned_sections:
        output_parts.append('===SUMMARY===')
        output_parts.append(cleaned_sections['SUMMARY'])
        output_parts.append('')
    
    if 'FACTS' in cleaned_sections:
        output_parts.append('===FACTS===')
        output_parts.append(cleaned_sections['FACTS'])
        output_parts.append('')
    
    if 'HYPOTHESES' in cleaned_sections:
        output_parts.append('===HYPOTHESES===')
        output_parts.append(cleaned_sections['HYPOTHESES'])
    
    result = '\n'.join(output_parts).strip()
    
    # Step 6: Final cleanup - remove any remaining markdown artifacts
    result = re.sub(r'\*\*', '', result)
    result = re.sub(r'__', '', result)
    result = re.sub(r'#{1,6}\s+', '', result)
    
    # Ensure no mid-word truncation (check for patterns like "word...word")
    result = re.sub(r'(\w+)\.\.\.(\w+)', r'\1 \2', result)
    
    return result if result else raw_output  # Return original if cleaning failed


async def generate_ai_summary(
    api_key: str,
    provider: str,
    model: str,
    prompt: str,
    lead_data: dict,
    scraped_content: str,
    status_callback=None
) -> str:
    """
    Generate AI summary for a company.
    
    Args:
        api_key: API key for the AI provider
        provider: 'openai' or 'gemini'
        model: Model name (e.g., 'gpt-4', 'gemini-pro')
        prompt: Base prompt template
        lead_data: Lead information dictionary
        scraped_content: Scraped website content
    
    Returns:
        Generated summary string (cleaned and structured)
    """
    if not api_key or not api_key.strip():
        return "❌ No API key provided"
    
    if not scraped_content or scraped_content.startswith("❌"):
        return "❌ No valid scraped content available"
    
    # Validate scraped content has meaningful content before sending to AI
    if len(scraped_content.strip()) < 50:
        return "❌ Insufficient content scraped. Content too short to generate meaningful summary."
    
    # Build complete prompt with STRICT formatting requirements
    full_prompt = build_company_summary_prompt(prompt, lead_data, scraped_content)
    
    full_prompt = full_prompt + COMPANY_SUMMARY_FINAL_REMINDER
    
    # Generate summary based on provider
    raw_output = ""
    if provider.lower() == 'openai':
        raw_output = await generate_openai_summary(api_key, model, full_prompt, status_callback=status_callback)
    elif provider.lower() == 'gemini':
        raw_output = await generate_gemini_summary(api_key, model, full_prompt, status_callback=status_callback)
    elif provider.lower() == 'openrouter':
        raw_output = await generate_openrouter_summary(api_key, model, full_prompt, status_callback=status_callback)
    else:
        return f"❌ Unknown provider: {provider}"
    
    # Post-process to fix all formatting issues
    cleaned_output = clean_and_structure_ai_output(raw_output)
    
    # Never return blank when we had a real response (so CSV always has an explicit reason if no summary)
    if not (cleaned_output or "").strip() and raw_output and not raw_output.startswith("❌"):
        return "❌ Summary empty: AI returned no usable content after formatting."
    return cleaned_output


def build_email_copy_prompt(prompt_template: str, lead_data: dict | None, scraped_content: str | None) -> str:
    """Build the complete prompt for email copy. Replaces {key} and {{key}} with formatted values."""
    template = (prompt_template or "").strip()
    lead_data = dict(lead_data or {})
    url = _format_var_value(lead_data.get("url")) or ""
    company_name = _format_var_value(lead_data.get("company_name")) or ""
    if not company_name and url:
        company_name = url.replace("https://", "").replace("http://", "").split("/")[0]
    lead_data["url"] = url
    lead_data["company_name"] = company_name
    lead_data["scraped_content"] = str(scraped_content or "")
    result = _replace_prompt_variables(template, lead_data)
    return result if result is not None else ""


def _is_ai_refusal(text: str) -> bool:
    """Detect common model refusal phrases so we can retry or show a helpful error."""
    if not text or len(text.strip()) < 10:
        return False
    t = text.strip().lower()
    refusal_phrases = (
        "i cannot assist",
        "i can't assist",
        "i'm sorry, but i cannot",
        "i am sorry, but i cannot",
        "i'm unable to assist",
        "i am unable to assist",
        "i can't help with that",
        "i cannot help with that",
        "i'm not able to",
        "i am not able to",
        "cannot fulfill this request",
        "cannot complete your request",
        "against my guidelines",
        "against my programming",
        "not able to provide",
    )
    return any(p in t for p in refusal_phrases)


async def generate_email_copy(
    api_key: str, provider: str, model: str, prompt_template: str,
    lead_data: dict, scraped_content: str, status_callback=None
) -> str:
    """Generate email copy for a lead. Uses email-specific system message so the model doesn't refuse or output summary format."""
    if not api_key or not api_key.strip():
        return "❌ No API key provided"
    if not scraped_content or scraped_content.startswith("❌"):
        return "❌ No valid scraped content available"
    if len(scraped_content.strip()) < 50:
        return "❌ Insufficient content scraped. Content too short to generate email copy."
    full_prompt = build_email_copy_prompt(prompt_template, lead_data, scraped_content)
    raw_output = ""
    # Use email-specific system message so the model writes the email instead of refusing or using summary format
    if provider.lower() == "openai":
        raw_output = await generate_openai_summary(api_key, model, full_prompt, status_callback=status_callback, system_message=EMAIL_COPY_SYSTEM)
    elif provider.lower() == "gemini":
        raw_output = await generate_gemini_summary(api_key, model, full_prompt, status_callback=status_callback, system_message=EMAIL_COPY_SYSTEM)
    elif provider.lower() == "openrouter":
        raw_output = await generate_openrouter_summary(api_key, model, full_prompt, status_callback=status_callback, system_message=EMAIL_COPY_SYSTEM)
    else:
        return f"❌ Unknown provider: {provider}"
    # Detect refusal: model said "I cannot assist" etc. — return helpful error so user can rephrase or try another model
    if raw_output and _is_ai_refusal(raw_output):
        return "❌ Email copy: The AI declined to write the email (content policy). Try rephrasing your prompt to focus on 'a short professional follow-up email based on this company's website' or use a different model (e.g. GPT-4o, Claude)."
    # Light cleanup for email copy (preserve line breaks)
    if raw_output and not raw_output.startswith("❌"):
        raw_output = raw_output.strip()
    return raw_output


def cleanup_html(html: str) -> str:
    """Clean and organize HTML content into well-structured text."""
    # Remove unwanted elements first
    html = re.sub(
        r"<(style|script|nav|footer|header|aside|noscript|iframe|svg|canvas|form|button)[\s\S]*?</\1>", 
        "", html, flags=re.IGNORECASE)
    
    # Remove common noise patterns (cookies, ads, etc.)
    html = re.sub(
        r'<(div|section|span)[^>]*(?:class|id)=["\'](?:cookie|ad|advertisement|popup|modal|overlay|banner|promo|marketing|tracking|analytics|social-share|share-buttons|newsletter|subscribe|signup)[^"\']*["\'][^>]*>[\s\S]*?</\1>',
        "", html, flags=re.IGNORECASE)
    
    # Convert headings to markdown-style headers
    html = re.sub(r'<h1[^>]*>(.*?)</h1>', r'\n\n# \1\n', html, flags=re.IGNORECASE | re.DOTALL)
    html = re.sub(r'<h2[^>]*>(.*?)</h2>', r'\n\n## \1\n', html, flags=re.IGNORECASE | re.DOTALL)
    html = re.sub(r'<h3[^>]*>(.*?)</h3>', r'\n\n### \1\n', html, flags=re.IGNORECASE | re.DOTALL)
    html = re.sub(r'<h4[^>]*>(.*?)</h4>', r'\n\n#### \1\n', html, flags=re.IGNORECASE | re.DOTALL)
    html = re.sub(r'<h5[^>]*>(.*?)</h5>', r'\n\n##### \1\n', html, flags=re.IGNORECASE | re.DOTALL)
    html = re.sub(r'<h6[^>]*>(.*?)</h6>', r'\n\n###### \1\n', html, flags=re.IGNORECASE | re.DOTALL)
    
    # Convert lists to markdown-style
    html = re.sub(r'<li[^>]*>(.*?)</li>', r'\n- \1', html, flags=re.IGNORECASE | re.DOTALL)
    html = re.sub(r'<(ul|ol)[^>]*>', '\n', html, flags=re.IGNORECASE)
    html = re.sub(r'</(ul|ol)>', '\n', html, flags=re.IGNORECASE)
    
    # Convert paragraphs to double newlines
    html = re.sub(r'<p[^>]*>(.*?)</p>', r'\n\n\1\n\n', html, flags=re.IGNORECASE | re.DOTALL)
    
    # Convert line breaks
    html = re.sub(r'<br\s*/?>', '\n', html, flags=re.IGNORECASE)
    
    # Convert strong/bold to markdown
    html = re.sub(r'<(strong|b)[^>]*>(.*?)</\1>', r'**\2**', html, flags=re.IGNORECASE | re.DOTALL)
    
    # Convert emphasis/italic to markdown
    html = re.sub(r'<(em|i)[^>]*>(.*?)</\1>', r'*\2*', html, flags=re.IGNORECASE | re.DOTALL)
    
    # Remove all remaining HTML tags
    html = re.sub(r'<[^>]+>', '', html)
    
    # Decode HTML entities
    html = unescape(html)
    
    # Clean up whitespace
    html = re.sub(r'&nbsp;', ' ', html)
    html = re.sub(r'\s+', ' ', html)  # Multiple spaces to single
    html = re.sub(r' \n', '\n', html)  # Space before newline
    html = re.sub(r'\n{3,}', '\n\n', html)  # Multiple newlines to double
    html = re.sub(r'^\s+', '', html, flags=re.MULTILINE)  # Leading whitespace
    html = re.sub(r'\s+$', '', html, flags=re.MULTILINE)  # Trailing whitespace
    
    # Remove common noise phrases (more conservative - only remove standalone phrases)
    # This prevents removing legitimate content that might contain these words
    noise_patterns = [
        r'^\s*cookie\s+policy\s*$',
        r'^\s*privacy\s+policy\s*$',
        r'^\s*terms\s+of\s+service\s*$',
        r'^\s*accept\s+cookies\s*$',
        r'^\s*we\s+use\s+cookies\s*$',
        r'^\s*skip\s+to\s+content\s*$',
        r'^\s*jump\s+to\s+content\s*$',
        r'^\s*menu\s*$',
        r'^\s*close\s*$',
        r'^\s*share\s+on\s*$',
        r'^\s*follow\s+us\s*$',
        r'^\s*subscribe\s+to\s*$',
        r'^\s*learn\s+more\s*$',
    ]
    for pattern in noise_patterns:
        html = re.sub(pattern, '', html, flags=re.IGNORECASE | re.MULTILINE)
    
    # Final cleanup
    html = html.strip()
    
    # Split into lines and clean each line (more conservative filtering)
    lines = html.split('\n')
    cleaned_lines = []
    for line in lines:
        line = line.strip()
        # Skip empty lines
        if not line:
            continue
        # Skip very short lines that are likely noise (but keep single characters if they're meaningful)
        if len(line) < 2:
            continue
        # Skip lines that are ONLY punctuation or symbols (but allow lines with some punctuation)
        if re.match(r'^[^\w\s]+$', line) and len(line) < 5:
            continue
        # Skip lines that are just numbers (likely page numbers or IDs)
        if re.match(r'^\d+$', line) and len(line) < 10:
            continue
        cleaned_lines.append(line)
    
    # Join back with proper spacing
    result = '\n'.join(cleaned_lines)
    
    # Final pass: ensure proper spacing around headers
    result = re.sub(r'\n(#{1,6}\s+[^\n]+)\n+', r'\n\1\n\n', result)
    
    # Fix double/malformed headers: "## ### Title" -> "### Title", "1 ### Title" -> "1. Title"
    result = re.sub(r'#{1,6}\s+(#{1,6}\s+)', r'\1', result)
    result = re.sub(r'^(\d+)\s+(#{1,6})\s+', r'\1. ', result, flags=re.MULTILINE)
    
    # Collapse excessive newlines (max 2)
    result = re.sub(r'\n{4,}', '\n\n\n', result)
    
    return result.strip()


def extract_links(html: str, base_url: str):
    """Extract links from HTML with improved accuracy and URL resolution."""
    from urllib.parse import urljoin, urlparse
    
    # More robust regex pattern for href extraction
    matches = re.findall(
        r'<a\s+(?:[^>]*?\s+)?href\s*=\s*(["\']?)([^"\'>\s]+)\1', html, flags=re.IGNORECASE)
    
    base_domain = "/".join(base_url.split("/")[:3])
    parsed_base = urlparse(base_url)
    base_scheme = parsed_base.scheme or "https"
    base_netloc = parsed_base.netloc or parsed_base.path.split("/")[0] if "/" in parsed_base.path else ""
    
    urls = []
    seen = set()
    
    for quote_char, href in matches:
        href = href.strip()
        if not href:
            continue
            
        # Skip anchors, mailto, javascript, and other non-HTTP links
        if href.startswith(("#", "mailto:", "javascript:", "tel:", "sms:", "data:")):
            continue
        
        # Resolve relative URLs properly
        try:
            if href.startswith("//"):
                # Protocol-relative URL
                resolved_url = base_scheme + ":" + href
            elif href.startswith("/"):
                # Absolute path
                resolved_url = base_scheme + "://" + base_netloc + href
            elif href.startswith("http://") or href.startswith("https://"):
                # Absolute URL
                resolved_url = href
            else:
                # Relative URL - use urljoin for proper resolution
                resolved_url = urljoin(base_url, href)
            
            # Parse to validate and normalize
            parsed = urlparse(resolved_url)
            if not parsed.netloc:
                continue
            
            # Only include URLs from the same domain
            if parsed.netloc == base_netloc or parsed.netloc.endswith("." + base_netloc):
                # Normalize URL (remove fragment, normalize path)
                normalized = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
                if parsed.query:
                    normalized += "?" + parsed.query
                
                # Avoid duplicates
                if normalized not in seen:
                    seen.add(normalized)
                    urls.append(normalized)
        except Exception:
            # Skip invalid URLs silently
            continue
    
    return urls

# -------------------------
# Network fetch + scraping
# -------------------------


def get_random_user_agent():
    """Generate a random realistic user agent."""
    import random
    user_agents = [
        # Chrome on Windows (most common)
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
        # Chrome on Mac
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
        # Firefox on Windows
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:132.0) Gecko/20100101 Firefox/132.0",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:131.0) Gecko/20100101 Firefox/131.0",
        # Firefox on Mac
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:132.0) Gecko/20100101 Firefox/132.0",
        # Safari on Mac
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.1 Safari/605.1.15",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.6 Safari/605.1.15",
        # Edge on Windows
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0",
    ]
    return random.choice(user_agents)

def get_realistic_headers(user_agent=None, target_url=None):
    """Generate realistic browser headers based on user agent - EPIC version."""
    import random
    if user_agent is None:
        user_agent = get_random_user_agent()
    
    # Detect browser type from user agent
    is_chrome = "Chrome" in user_agent and "Edg" not in user_agent
    is_firefox = "Firefox" in user_agent
    is_safari = "Safari" in user_agent and "Chrome" not in user_agent
    is_edge = "Edg" in user_agent
    
    # EPIC referer strategy - sometimes no referer (direct navigation), sometimes search engines
    # 30% chance of no referer (direct navigation), 70% chance of search engine referer
    referers = [
        None,  # Direct navigation - no referer
        "https://www.google.com/",
        "https://www.google.com/search?q=example",
        "https://www.google.com/search?q=site",
        "https://www.bing.com/",
        "https://duckduckgo.com/",
        "https://www.yahoo.com/",
        "https://www.google.com/search?client=firefox-b-d&q=example",
    ]
    
    # If target_url provided, sometimes use it as referer (internal navigation)
    if target_url and random.random() < 0.1:
        referer = target_url.rsplit('/', 1)[0] + '/'  # Parent directory
    else:
        referer = random.choice(referers)
    
    # Random viewport sizes (common resolutions)
    viewports = [
        (1920, 1080),
        (1366, 768),
        (1536, 864),
        (1440, 900),
        (1280, 720),
    ]
    viewport_width, viewport_height = random.choice(viewports)
    
    headers = {
        "User-Agent": user_agent,
        "Accept-Language": random.choice([
            "en-US,en;q=0.9",
            "en-US,en;q=0.9,es;q=0.8",
            "en-US,en;q=0.9,fr;q=0.8",
            "en-US,en;q=0.9,de;q=0.8",
            "en-GB,en;q=0.9",
        ]),
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "DNT": random.choice(["1", "0"]),  # Sometimes DNT, sometimes not
    }
    
    # Only add referer if we have one (30% chance of None for direct navigation)
    if referer:
        headers["Referer"] = referer
    
    if is_chrome or is_edge:
        headers.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": random.choice(["none", "cross-site", "same-origin"]),
            "Sec-Fetch-User": "?1",
            "Sec-CH-UA": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            "Sec-CH-UA-Mobile": "?0",
            "Sec-CH-UA-Platform": '"Windows"',
            "Cache-Control": random.choice(["max-age=0", "no-cache"]),
            "Viewport-Width": str(viewport_width),
        })
    elif is_firefox:
        headers.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        })
    elif is_safari:
        headers.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        })
    
    return headers

# Global domain-based rate limiting - EPIC anti-bot feature
_domain_request_times = {}  # Track last request time per domain
_domain_lock = asyncio.Lock()  # Lock for thread-safe access to domain timing

# Limit concurrent Playwright instances. Local: 4 for better throughput; cloud: 2 for stability
_playwright_semaphore = None
def _get_playwright_semaphore():
    global _playwright_semaphore
    if _playwright_semaphore is None:
        n = 4 if not is_cloud_mode() else 2
        _playwright_semaphore = asyncio.Semaphore(n)
    return _playwright_semaphore


def _is_google_error_or_captcha_page(html: str) -> bool:
    """True if HTML is a Google error/captcha/trouble page, not real cached content."""
    if not html or len(html.strip()) < 100:
        return False
    h = html.lower()
    # Google "having trouble accessing" / captcha / error pages
    if "having trouble accessing" in h and "google" in h:
        return True
    if "if you're having trouble" in h and ("google search" in h or "send feedback" in h):
        return True
    if "google search" in h and ("please click here" in h or "send feedback" in h):
        return True
    if "google" in h and "please click here" in h:
        return True
    if "google" in h and "send feedback" in h and ("trouble" in h or "accessing" in h):
        return True
    if "sorry," in h and "we're having trouble" in h and "google" in h:
        return True
    if "unusual traffic" in h and "google" in h:
        return True
    if "captcha" in h and "google" in h:
        return True
    # Reject pages that are mostly Google UI with no real cached body
    if "google" in h and "webcache.googleusercontent" in h:
        if h.count("google") >= 2 and ("trouble" in h or "feedback" in h or "click here" in h):
            return True
    return False


def _is_challenge_or_verification_page(html: str) -> bool:
    """True if HTML is a Cloudflare/bot challenge or 'please wait' verification page, not real content."""
    if not html or len(html.strip()) < 20:
        return False
    h = html.lower()
    # Cloudflare / "One moment, please" / verification interstitials (match split text too)
    if "one moment" in h and "please" in h:
        return True
    if "your request is being verified" in h:
        return True
    if "request is being verified" in h:
        return True
    if "please wait" in h and ("verif" in h or "request" in h):
        return True
    if "please wait while" in h:
        return True
    if "checking your browser" in h:
        return True
    if "just a moment" in h and ("enable javascript" in h or "cloudflare" in h):
        return True
    if "cloudflare" in h and ("ray id" in h or "performance" in h or "checking" in h):
        return True
    if "ddos protection" in h and "cloudflare" in h:
        return True
    if "attention required" in h and "cloudflare" in h:
        return True
    # Very short body that is only verification text
    if len(h) < 800 and ("verif" in h or "moment" in h) and "please" in h:
        return True
    # "Please wait while your request is being verified" - catch variations
    if "being verified" in h and "wait" in h:
        return True
    return False


def _cleaned_text_is_challenge(cleaned: str) -> bool:
    """True if cleaned (visible) text is clearly a challenge/verification page, not real content."""
    if not cleaned:
        return False
    # Normalize: collapse whitespace so "one   moment" / "one\nmoment" still match
    t = re.sub(r'\s+', ' ', cleaned.strip()).lower()
    if len(t) < 40:
        return False
    if "one moment" in t and "please" in t:
        return True
    if "just a moment" in t or "just one moment" in t:
        return True
    if "request is being verified" in t or "your request is being verified" in t:
        return True
    if "please wait while" in t:
        return True
    if "please wait" in t and ("verif" in t or "moment" in t):
        return True
    if "being verified" in t and "wait" in t:
        return True
    if "checking your browser" in t and ("moment" in t or "please" in t):
        return True
    if "enable javascript" in t and ("verify" in t or "moment" in t):
        return True
    return False


# Minimum cleaned text length to accept as valid result (under this = error/blocked/insufficient)
_MIN_CONTENT_LEN = 200

# Minimum cleaned text length to accept a cache/fallback result
_MIN_FALLBACK_TEXT_LEN = 200

# Below this length, content from a different domain is treated as mismatch (wrong page)
_SHORT_CONTENT_FOR_MISMATCH = 500


def _scraped_result_is_bad(text: str) -> bool:
    """True if this scraped result should be treated as error (empty, short, challenge, or Google error)."""
    if text is None:
        return True
    s = (text if isinstance(text, str) else "").strip()
    if s.startswith("❌"):
        return False  # already an error message, don't replace
    if len(s) < _MIN_CONTENT_LEN:
        return True
    # Check full text for challenge (catches "One moment, please..." in body)
    if _cleaned_text_is_challenge(s):
        return True
    # Check each page body (content after "PAGE: url" blocks) in case of multi-page output
    page_blocks = re.split(r'\n--- PAGE:\s*[^\n]+\s+---\s*\n+|\n={20,}\s*\nPAGE:\s*[^\n]+\n={20,}\s*\n+', s)
    for block in page_blocks:
        if block.strip() and _cleaned_text_is_challenge(block):
            return True
    t = s.lower()
    if "content was a google error" in t or ("google" in t and "captcha" in t) or ("having trouble" in t and "google" in t):
        return True
    return False


def _cache_result_acceptable(url: str, html: str, max_chars: int = 50000) -> bool:
    """Return True only if this HTML yields enough real content (not error/placeholder)."""
    if not html or len(html.strip()) < 100:
        return False
    if _is_google_error_or_captcha_page(html):
        return False
    if _is_challenge_or_verification_page(html):
        return False
    text = _process_cached_html_to_text(url, html, max_chars)
    if not text or len(text.strip()) < _MIN_FALLBACK_TEXT_LEN:
        return False
    t = text.lower()
    if "having trouble accessing" in t and "google" in t:
        return False
    if "please click here" in t and "google" in t:
        return False
    if "send feedback" in t and "google" in t and ("trouble" in t or "accessing" in t):
        return False
    if "one moment" in t and "please" in t:
        return False
    if "request is being verified" in t or "your request is being verified" in t:
        return False
    if "please wait while" in t:
        return False
    # Reject challenge pages regardless of length (bug fix: was len(t) < 800, allowing long challenge pages through)
    if "please wait" in t and ("verif" in t or "moment" in t):
        return False
    if "being verified" in t and "wait" in t:
        return False
    return True


async def _fetch_from_cache_fallbacks(
    session: aiohttp.ClientSession, url: str, timeout: int,
    use_playwright_fallback: bool = False, use_common_crawl_fallback: bool = False,
    skip_google_cache: bool = False, quick_mode: bool = False,
    try_playwright_first: bool = False,
    playwright_full_time: bool = False,
) -> tuple | str:
    """
    When direct fetch fails: try archive.org, Playwright, Google cache, Common Crawl.
    try_playwright_first=True: try Playwright before archive (for Cloudflare/challenge pages).
    quick_mode=True uses shorter timeouts. For timeout recovery, Playwright uses full time (not quick).
    Returns (page_url, html) on success, or error string on failure.
    """
    from urllib.parse import quote
    cache_total = 12 if quick_mode else 20
    cache_timeout = aiohttp.ClientTimeout(total=cache_total, connect=8 if quick_mode else 10, sock_read=10 if quick_mode else 15)
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    max_chars = 50000  # for validation only

    # Try Playwright first when we know it's Cloudflare/challenge (real browser bypasses it)
    if try_playwright_first and use_playwright_fallback:
        result = await _fetch_via_playwright(url, timeout, quick_mode=False)
        if isinstance(result, tuple):
            page_url, html = result
            if _cache_result_acceptable(page_url, html, max_chars):
                return result

    # 1. Try archive.org — most reliable when site is down or blocked
    # quick_mode: 2 snapshots only; normal: 4 snapshots
    wayback_paths = (
        (f"https://web.archive.org/web/{url}", f"https://web.archive.org/web/20240601000000/{url}")
        if quick_mode
        else (
            f"https://web.archive.org/web/{url}",
            f"https://web.archive.org/web/20240601000000/{url}",
            f"https://web.archive.org/web/20240101000000/{url}",
            f"https://web.archive.org/web/20230601000000/{url}",
        )
    )
    for wayback_path in wayback_paths:
        try:
            async with session.get(wayback_path, timeout=cache_timeout, headers=headers, allow_redirects=True) as resp:
                if resp.status < 400:
                    html = await resp.text(errors="replace")
                    if html and len(html.strip()) > 200:
                        if _cache_result_acceptable(url, html, max_chars):
                            return (url, html)
        except Exception:
            pass

    # 2. Try Playwright — gets real page for Cloudflare/bot-blocked sites (playwright_full_time=use full 45s for timeout recovery)
    if use_playwright_fallback:
        result = await _fetch_via_playwright(url, timeout, quick_mode=quick_mode and not playwright_full_time)
        if isinstance(result, tuple):
            page_url, html = result
            if _cache_result_acceptable(page_url, html, max_chars):
                return result

    # 3. Try Google cache (skip when we already know content was Google error; skip in quick_mode - often slow)
    if not skip_google_cache and not quick_mode:
        try:
            cache_url = f"https://webcache.googleusercontent.com/search?q=cache:{quote(url, safe='')}"
            async with session.get(cache_url, timeout=cache_timeout, headers=headers) as resp:
                if resp.status < 400:
                    html = await resp.text(errors="replace")
                    if html and len(html.strip()) > 200 and ("<html" in html.lower() or "<!doctype" in html.lower() or "<meta" in html.lower()):
                        if _cache_result_acceptable(url, html, max_chars):
                            return (url, html)
        except Exception:
            pass

    # 4. Common Crawl — free external crawl data
    if use_common_crawl_fallback:
        result = await _fetch_from_common_crawl(session, url, timeout, quick_mode=quick_mode)
        if isinstance(result, tuple):
            page_url, html = result
            if _cache_result_acceptable(page_url, html, max_chars):
                return result
    return ""


async def _fetch_via_playwright(url: str, timeout: int, quick_mode: bool = False) -> tuple | str:
    """
    Headless browser fallback — bypasses Cloudflare and bot protection. Uses Playwright/Chromium.
    Cloudflare needs 5–10s to pass; we use 45s goto + 6s wait for reliable success.
    Returns (url, html) or "".
    """
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return ""
    sem = _get_playwright_semaphore()
    # Cloudflare needs time: 45s load + 6s wait for challenge to pass (was 25s+4s)
    goto_timeout = 30000 if quick_mode else 45000
    sleep_after = 4 if quick_mode else 6
    urls_to_try = [url]
    if url.startswith("https://"):
        urls_to_try.append(url.replace("https://", "http://", 1))
    elif url.startswith("http://"):
        urls_to_try.append(url.replace("http://", "https://", 1))
    for try_url in urls_to_try:
        try:
            async with sem:
                async with async_playwright() as p:
                    browser = await p.chromium.launch(headless=True)
                    try:
                        page = await browser.new_page()
                        await page.goto(try_url, wait_until="domcontentloaded", timeout=goto_timeout)
                        await asyncio.sleep(sleep_after)
                        html = await page.content()
                        if html and _is_challenge_or_verification_page(html):
                            await asyncio.sleep(5 if quick_mode else 8)  # Cloudflare often needs 5–8s
                            html = await page.content()
                        if html and len(html.strip()) > 200 and not _is_challenge_or_verification_page(html):
                            return (try_url, html)
                    finally:
                        await browser.close()
        except Exception:
            pass
    return ""


async def _fetch_from_common_crawl(session: aiohttp.ClientSession, url: str, timeout: int, quick_mode: bool = False) -> tuple | str:
    """
    Common Crawl fallback — free open crawl data. Try multiple indexes for 90%+ success rate.
    Returns (url, html) or "".
    """
    from urllib.parse import quote_plus
    cc_timeout = 10 if quick_mode else 15
    timeout_obj = aiohttp.ClientTimeout(total=cc_timeout)
    # Try recent indexes (update periodically). Format: CC-MAIN-YYYY-WW. quick_mode: 2 indexes only
    indexes = ("CC-MAIN-2025-51", "CC-MAIN-2025-47") if quick_mode else ("CC-MAIN-2025-51", "CC-MAIN-2025-47", "CC-MAIN-2025-43", "CC-MAIN-2024-51", "CC-MAIN-2024-41", "CC-MAIN-2024-31")
    for CC_INDEX in indexes:
        try:
            index_url = f"https://index.commoncrawl.org/{CC_INDEX}-index?url={quote_plus(url)}&output=json&limit=1"
            async with session.get(index_url, timeout=timeout_obj) as resp:
                if resp.status != 200:
                    continue
                text = await resp.text()
                lines = [ln.strip() for ln in text.strip().split("\n") if ln.strip()]
                if not lines:
                    continue
                record = json.loads(lines[0])
                warc_filename = record.get("filename")
                warc_offset = int(record.get("offset", 0) or 0)
                warc_length = int(record.get("length", 0) or 0)
                if not warc_filename or warc_length <= 0:
                    continue
            data_url = f"https://data.commoncrawl.org/{warc_filename}"
            async with session.get(data_url, timeout=timeout_obj, headers={"Range": f"bytes={warc_offset}-{warc_offset + warc_length - 1}"}) as r:
                if r.status not in (200, 206):
                    continue
                body = await r.read()
            from warcio.archiveiterator import ArchiveIterator
            from io import BytesIO
            for rec in ArchiveIterator(BytesIO(body)):
                if rec.rec_type == "response":
                    html = rec.content_stream().read().decode("utf-8", errors="replace")
                    if html and len(html.strip()) > 200:
                        return (url, html)
                    break
        except Exception:
            continue
    return ""


def _process_cached_html_to_text(url: str, html: str, max_chars: int = 50000) -> str:
    """Convert cached HTML to scraped-text format (same as scrape_site output)."""
    if not html or len(html.strip()) < 50:
        return ""
    cleaned = cleanup_html(html)
    if not cleaned or len(cleaned.strip()) < _MIN_FALLBACK_TEXT_LEN:
        return ""
    page_header = f"\n--- PAGE: {url} ---\n\n"
    content = page_header + cleaned[:max_chars]
    return content


async def fetch(session: aiohttp.ClientSession, url: str, timeout: int, retries: int, fast_mode: bool = False):
    """Fetch URL with EPIC anti-bot detection avoidance. fast_mode reduces delays for 4k+ runs."""
    from urllib.parse import urlparse
    import random
    
    parsed = urlparse(url)
    domain = parsed.netloc.replace('www.', '').lower()
    
    # Domain rate limit - much shorter in fast mode
    if fast_mode:
        min_delay = random.uniform(0.15, 0.4)
        pre_request_delay = random.uniform(0.02, 0.1)
    else:
        min_delay = random.uniform(2.0, 5.0)
        pre_request_delay = random.uniform(1.0, 3.0)
    
    async with _domain_lock:
        if domain in _domain_request_times:
            last_request_time = _domain_request_times[domain]
            time_since_last = time.time() - last_request_time
            if time_since_last < min_delay:
                wait_time = min_delay - time_since_last
                await asyncio.sleep(wait_time)
        _domain_request_times[domain] = time.time()
    
    await asyncio.sleep(pre_request_delay)
    
    # Try different URL variations and header combinations for better success rate
    url_variations = [url]
    parsed = urlparse(url)
    
    # Generate URL variations to try
    if parsed.netloc:
        domain = parsed.netloc
        path = parsed.path or '/'
        # Try with/without www
        if domain.startswith('www.'):
            url_variations.append(url.replace('www.', ''))
        else:
            url_variations.append(url.replace(domain, 'www.' + domain))
        # Try http if https
        if url.startswith('https://'):
            url_variations.append(url.replace('https://', 'http://'))
    
    # Header variations - fewer in fast_mode to fail faster and leave time for cache fallback
    header_variations = []
    n_headers = 5 if fast_mode else 15
    for _ in range(n_headers):
        headers = get_realistic_headers(target_url=url)
        header_variations.append(headers)
    if not fast_mode:
        for _ in range(5):
            headers = get_realistic_headers(target_url=url)
            headers["Accept-Encoding"] = "gzip, deflate"
            header_variations.append(headers)
    random.shuffle(header_variations)
    
    last_error = None
    
    # Try URL variations and header combinations
    for url_to_try in url_variations:
        for headers in header_variations:
            # Override with session's user agent if provided, but regenerate headers to maintain consistency
            if session.headers.get("User-Agent"):
                # Regenerate headers with the session's user agent to maintain consistency
                headers = get_realistic_headers(session.headers.get("User-Agent"), target_url=url_to_try)
            
            for attempt in range(retries + 1):
                # Increase timeout for retries (some sites are slow) - more generous on later attempts for 90%+ success
                retry_timeout = timeout + (attempt * 15)
                sock_read_cap = min(retry_timeout, 120)  # Allow up to 120s for slow sites to send data
                req_timeout = aiohttp.ClientTimeout(total=retry_timeout, connect=20, sock_read=sock_read_cap)
                
                try:
                    if attempt > 0:
                        delay = (random.uniform(0.3, 0.8) + attempt * 0.2) if fast_mode else (random.uniform(3.0, 6.0) + attempt * 1.5)
                        await asyncio.sleep(delay)
                    
                    # Parse URL
                    parsed_original = urlparse(url_to_try)
                    if not parsed_original.netloc:
                        continue
                    
                    original_domain = parsed_original.netloc.replace('www.', '').lower()
                    
                    await asyncio.sleep(random.uniform(0.02, 0.15) if fast_mode else random.uniform(0.5, 2.0))
                    
                    async with session.get(url_to_try, timeout=req_timeout, allow_redirects=True, headers=headers) as resp:
                        # Handle redirects properly - use the final URL after redirects
                        final_url = str(resp.url)
                        parsed_final = urlparse(final_url)
                        final_domain = parsed_final.netloc.replace('www.', '').lower()
                        
                        # More lenient domain validation - allow common redirect patterns
                        # Extract base domains for comparison
                        original_base = '.'.join(original_domain.split('.')[-2:]) if '.' in original_domain else original_domain
                        final_base = '.'.join(final_domain.split('.')[-2:]) if '.' in final_domain else final_domain
                        
                        # Only reject if base domains are completely different
                        if final_domain != original_domain:
                            # Check if it's a subdomain variation
                            if not (final_domain.endswith('.' + original_domain) or original_domain.endswith('.' + final_domain)):
                                # Check if base domains match (e.g., www.example.com -> example.com)
                                if original_base != final_base:
                                    # Completely different domain - but be lenient, allow it
                                    # Many sites redirect to different domains (e.g., country-specific domains)
                                    # Only log for debugging, don't block
                                    pass
                        
                        # Check status code - handle 403 and 429 specially
                        if resp.status == 403:
                            last_error = f"HTTP 403 at {final_url}"
                            backoff_delay = (random.uniform(1.0, 3.0) * (1.5 ** attempt)) if fast_mode else (random.uniform(5.0, 10.0) * (2 ** attempt))
                            backoff_delay = min(backoff_delay, 15.0 if fast_mode else 60.0)
                            await asyncio.sleep(backoff_delay)
                            async with _domain_lock:
                                _domain_request_times[domain] = time.time() + (random.uniform(2.0, 5.0) if fast_mode else random.uniform(10.0, 20.0))
                            
                            break  # Try next variation
                        elif resp.status == 429:
                            retry_after = int(resp.headers.get('Retry-After', 3 + attempt if fast_mode else 5 + attempt * 2))
                            last_error = f"HTTP 429 at {final_url} (rate limited)"
                            await asyncio.sleep(min(retry_after, 15 if fast_mode else 30))
                            continue  # Retry same URL/headers
                        elif resp.status >= 400:
                            # Other 4xx/5xx errors
                            last_error = f"HTTP {resp.status} at {final_url}"
                            if resp.status == 404:
                                # 404 is final, don't retry
                                return last_error
                            break  # Try next variation
                        
                        # Success! Read content
                        try:
                            html = await resp.text(errors="replace")
                        except Exception as decode_error:
                            error_str = str(decode_error).lower()
                            # Handle brotli encoding error
                            if 'brotli' in error_str or 'br' in error_str or 'content-encoding' in error_str:
                                # Retry without brotli
                                headers_no_br = headers.copy()
                                headers_no_br["Accept-Encoding"] = "gzip, deflate"
                                # Try once more with different encoding
                                async with session.get(url_to_try, timeout=aiohttp.ClientTimeout(total=timeout), allow_redirects=True, headers=headers_no_br) as resp2:
                                    if resp2.status < 400:
                                        html = await resp2.text(errors="replace")
                                    else:
                                        last_error = f"HTTP {resp2.status} at {str(resp2.url)}"
                                        break
                            else:
                                last_error = f"Error decoding content: {str(decode_error)}"
                                break
                        
                        # Validate it's actually HTML
                        html_lower = html.lower()
                        if '<html' not in html_lower and '<!doctype' not in html_lower:
                            # Check if it's a redirect page or error
                            if 'redirect' in html_lower[:500] or 'location.replace' in html_lower[:500] or 'window.location' in html_lower[:500]:
                                last_error = f"JavaScript redirect detected at {final_url}"
                                break
                            # Might be valid content without HTML tags (rare), but check length
                            if len(html.strip()) < 100:
                                last_error = f"Non-HTML or empty content at {final_url}"
                                break
                        
                        # Reject Cloudflare/challenge pages so we can try cache fallbacks instead
                        if _is_challenge_or_verification_page(html):
                            last_error = f"Challenge/verification page at {final_url} (e.g. Cloudflare)"
                            break
                        
                        # Success! Return the content
                        return (final_url, html)
                        
                except asyncio.TimeoutError:
                    last_error = (
                        f"Timeout: request to {url_to_try[:80]}{'…' if len(url_to_try) > 80 else ''} exceeded {retry_timeout}s "
                        f"(connection + response). Tip: increase Timeout in Step 2 Settings for slow sites, or site may be blocking."
                    )
                    if attempt < retries:
                        await asyncio.sleep(0.5 + attempt * 0.3 if fast_mode else 2 + attempt * 1)
                    continue
                except (aiohttp.ClientError, ConnectionResetError, OSError) as e:
                    error_str = str(e).lower()
                    # Handle brotli encoding error specifically
                    if 'brotli' in error_str or 'br' in error_str or 'content-encoding' in error_str:
                        last_error = f"Brotli encoding error: {str(e)}"
                        if attempt < retries:
                            await asyncio.sleep(0.3 + attempt * 0.2 if fast_mode else 1 + attempt * 0.5)
                        continue
                    if 'connection' in error_str or 'disconnected' in error_str or 'getaddrinfo' in error_str:
                        last_error = f"Error fetching {url_to_try}: {str(e)}"
                        if attempt < retries:
                            await asyncio.sleep(0.5 + attempt * 0.3 if fast_mode else 2 + attempt * 1)
                        continue
                    last_error = f"Error fetching {url_to_try}: {str(e)}"
                    if attempt < retries:
                        await asyncio.sleep(0.3 + attempt * 0.2 if fast_mode else 1 + attempt * 0.5)
                    continue
                except UnicodeDecodeError as e:
                    last_error = f"Encoding error: {str(e)}"
                    if attempt < retries:
                        await asyncio.sleep(0.3 + attempt * 0.2 if fast_mode else 1 + attempt * 0.5)
                    continue
                except Exception as e:
                    last_error = f"Unexpected error fetching {url_to_try}: {str(e)}"
                    if attempt < retries:
                        await asyncio.sleep(0.3 + attempt * 0.2 if fast_mode else 1 + attempt * 0.5)
                    continue
    
    # If we get here, all variations failed
    return last_error or f"Failed to fetch {url} after trying all variations"


async def scrape_site(session, url: str, depth: int, keywords, max_chars: int, retries: int, timeout: int, fast_mode: bool = False,
                     use_playwright_fallback: bool = False, use_common_crawl_fallback: bool = False):
    depth = max(1, int(depth) if depth is not None else 1)
    max_chars = max(100, min(int(max_chars) if max_chars is not None else 10000, 50000))
    visited, results, errors = set(), [], []
    total_chars = 0
    separator = "\n\n" + "-" * 80 + "\n\n"  # Better visual separator between pages
    separator_len = len(separator)

    # Normalize URL before fetching
    normalized_url = normalize_url(url)

    # In fast_mode, limit direct fetch to 45s so cache fallback gets 75s within the 120s budget
    fetch_budget = 45 if fast_mode else None
    fetch_retries = min(retries, 2) if fast_mode else retries
    try:
        if fetch_budget:
            homepage = await asyncio.wait_for(fetch(session, normalized_url, timeout, fetch_retries, fast_mode), fetch_budget)
        else:
            homepage = await fetch(session, normalized_url, timeout, fetch_retries, fast_mode)
    except asyncio.TimeoutError:
        homepage = (
            f"Timeout: initial connection/response exceeded {fetch_budget}s. "
            f"Tip: increase Timeout in Step 2 Settings, or site may be blocking."
        )
    if isinstance(homepage, str):
        # Direct fetch failed: try Playwright first when error suggests Cloudflare/blocking
        try_pw_first = "challenge" in homepage.lower() or "cloudflare" in homepage.lower() or "timeout" in homepage.lower()
        cache_result = await _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback, try_playwright_first=try_pw_first)
        if isinstance(cache_result, tuple):
            homepage = cache_result
        else:
            await asyncio.sleep(1)
            cache_result = await _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback, try_playwright_first=try_pw_first)
            if isinstance(cache_result, tuple):
                homepage = cache_result
            else:
                return f"❌ {homepage}"

    page_url, html = homepage
    # Reject Google error/captcha pages — try archive/Playwright/Common Crawl only (skip Google cache)
    if _is_google_error_or_captcha_page(html):
        cache_result = await _fetch_from_cache_fallbacks(
            session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback,
            skip_google_cache=True,
        )
        if isinstance(cache_result, tuple):
            page_url, html = cache_result
            if _is_google_error_or_captcha_page(html):
                return "❌ Google cache returned an error page; content unavailable. Try again later."
        else:
            await asyncio.sleep(1)
            cache_result = await _fetch_from_cache_fallbacks(
                session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback,
                skip_google_cache=True,
            )
            if isinstance(cache_result, tuple):
                page_url, html = cache_result
                if _is_google_error_or_captcha_page(html):
                    return "❌ Google cache returned an error page; content unavailable. Try again later."
            else:
                return "❌ Content was a Google error page; cache fallback had no better result. Try again later."
    # Reject Cloudflare/challenge pages so we try archive and other fallbacks instead
    if _is_challenge_or_verification_page(html):
        cache_result = await _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback, try_playwright_first=True)
        if isinstance(cache_result, tuple):
            page_url, html = cache_result
            if _is_challenge_or_verification_page(html) or _is_google_error_or_captcha_page(html):
                return "❌ Challenge/verification page; cache fallback had no better result. Try again later."
        else:
            await asyncio.sleep(1)
            cache_result = await _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback, try_playwright_first=True)
            if isinstance(cache_result, tuple):
                page_url, html = cache_result
                if _is_challenge_or_verification_page(html) or _is_google_error_or_captcha_page(html):
                    return "❌ Challenge/verification page; cache fallback had no better result. Try again later."
            else:
                return "❌ Challenge/verification page (e.g. Cloudflare); cache fallback had no better result. Try again later."
    
    # STRICT validation: Verify we're on the correct domain
    from urllib.parse import urlparse
    parsed_page = urlparse(page_url)
    parsed_original = urlparse(normalized_url)
    
    page_domain = parsed_page.netloc.replace('www.', '').lower()
    original_domain = parsed_original.netloc.replace('www.', '').lower()
    
    # Only log redirect when it's meaningful (different domain), not http->https or www variation
    if page_url != normalized_url and page_domain != original_domain:
        errors.append(f"ℹ️ Redirected from {normalized_url} to {page_url}")
    
    # More lenient domain validation - allow common redirect patterns
    # Extract base domains for comparison
    original_base = '.'.join(original_domain.split('.')[-2:]) if '.' in original_domain else original_domain
    page_base = '.'.join(page_domain.split('.')[-2:]) if '.' in page_domain else page_domain
    
    # Only reject if base domains are completely different
    if page_domain != original_domain:
        # Check if it's a subdomain variation
        if not (page_domain.endswith('.' + original_domain) or original_domain.endswith('.' + page_domain)):
            # Check if base domains match (e.g., www.example.com -> example.com)
            if original_base != page_base:
                # Completely different domain - but be lenient, allow it
                # Many sites redirect to different domains (e.g., country-specific domains)
                # Only log for debugging, don't block
                pass
    
    # Validate HTML content before processing
    if not html or len(html.strip()) < 50:
        return f"❌ Empty or too short HTML content from {page_url}"
    
    # Check if HTML looks valid (has some HTML structure)
    html_lower = html.lower()
    
    # More robust validation - check for common content indicators
    has_content = (
        '<html' in html_lower or 
        '<!doctype' in html_lower or 
        '<body' in html_lower or
        '<main' in html_lower or
        '<article' in html_lower or
        '<div' in html_lower[:2000]  # Check first 2000 chars for div tags
    )
    
    if not has_content:
        # Check if it's a redirect page
        if 'location.href' in html_lower or 'window.location' in html_lower or 'meta http-equiv="refresh"' in html_lower:
            return f"❌ JavaScript/Meta redirect detected at {page_url}. Content may not be accessible."
        # Might still have content, continue
    
    # ADDITIONAL SAFETY CHECK: Detect if content seems wrong for the domain
    # Check for common e-commerce patterns that shouldn't be on a business consultant site
    if len(html) > 500:
        html_sample = html_lower[:10000]  # Check first 10k chars
        
        # E-commerce indicators
        ecommerce_patterns = [
            'add to cart', 'shopping cart', 'buy now', 'add to bag',
            'product reviews', 'customer reviews', 'star rating',
            'amazon.com', 'ebay.com', 'etsy.com', 'shopify store',
            'price:', '$', 'usd', 'eur', 'gbp',
            'remote helicopter', 'rc helicopter', 'replacement parts'
        ]
        
        ecommerce_count = sum(1 for pattern in ecommerce_patterns if pattern in html_sample)
        
        # If we find many e-commerce patterns, it's likely wrong content
        if ecommerce_count >= 5:
            # Check if domain name appears in content (if not, it's definitely wrong)
            domain_in_content = original_domain.split('.')[0] in html_sample
            if not domain_in_content:
                return f"❌ Content mismatch detected! Page at {page_url} contains e-commerce content that doesn't match expected domain {original_domain}. This might be a redirect or wrong content."
    
    cleaned = cleanup_html(html)
    # Catch challenge/verification text that only appears after cleanup (e.g. split across tags)
    if cleaned and _cleaned_text_is_challenge(cleaned):
        cache_result = await _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback, try_playwright_first=True)
        if isinstance(cache_result, tuple):
            page_url, html = cache_result
            if _is_challenge_or_verification_page(html) or _is_google_error_or_captcha_page(html):
                return "❌ Challenge/verification page; cache fallback had no better result. Try again later."
            cleaned = cleanup_html(html)
            if cleaned and _cleaned_text_is_challenge(cleaned):
                return "❌ Challenge/verification page; cache fallback had no better result. Try again later."
        else:
            return "❌ Challenge/verification page (e.g. Cloudflare); cache fallback had no better result. Try again later."
    
    # No content or insufficient (under 200 chars) = try fallbacks once, then error
    if not cleaned or len(cleaned.strip()) < _MIN_CONTENT_LEN:
        cache_result = await _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback)
        if isinstance(cache_result, tuple):
            page_url, html = cache_result
            cleaned = cleanup_html(html)
        if not cleaned or len(cleaned.strip()) < _MIN_CONTENT_LEN:
            return "❌ Insufficient content (under 200 characters); likely error or blocked page. Try again later."
    
    # Validate cleaned content (require meaningful length and reject challenge/verification pages)
    if cleaned and len(cleaned.strip()) >= _MIN_CONTENT_LEN and not _cleaned_text_is_challenge(cleaned):
        # Better page header with clear separation
        page_header = f"\n--- PAGE: {page_url} ---\n\n"
        page_content = page_header + cleaned
        # Check if adding this page would exceed max_chars
        if total_chars + len(page_content) + separator_len <= max_chars:
            results.append(page_content)
            total_chars += len(page_content) + separator_len
        elif total_chars == 0:  # At least include homepage even if it's large
            # Truncate homepage to fit (at word boundary)
            available_chars = max_chars - len(page_header) - separator_len
            if available_chars > 0:
                truncated_content = _truncate_at_word(cleaned, available_chars)
                results.append(page_header + truncated_content)
                total_chars = max_chars
        # If total_chars > 0 and adding would exceed, skip
    else:
        # Too short, empty, or challenge page — treat as error
        if not cleaned or len(cleaned.strip()) == 0:
            errors.append(f"❌ No extractable text content on homepage: {page_url}")
        elif _cleaned_text_is_challenge(cleaned):
            errors.append(f"❌ Challenge/verification page (e.g. Cloudflare) on homepage: {page_url}")
        else:
            errors.append(f"❌ Insufficient content from {page_url} (under {_MIN_CONTENT_LEN} characters); likely error or blocked page.")

    # Use the actual fetched URL (after redirects) for link extraction
    links = extract_links(html, page_url)
    # Improved keyword matching - case-insensitive, handles hyphens/underscores
    if keywords:  # Only process if keywords are provided
        for kw in keywords:
            if not kw or not kw.strip():
                continue
            kw_clean = kw.strip()
            match = next((l for l in links if match_keyword_in_url(kw_clean, l)), None)
            if match:
                visited.add(match)
    for l in links[: depth * 2]:
        visited.add(l)

    # Process additional pages while respecting max_chars limit
    for link in visited:
        if total_chars >= max_chars:
            break  # Stop if we've reached the limit
        
        res = await fetch(session, link, timeout, retries, fast_mode)
        if isinstance(res, str):
            errors.append(f"❌ {res}")
        elif isinstance(res, tuple):
            link_url, html2 = res
            
            # Validate HTML content
            if not html2 or len(html2.strip()) < 50:
                errors.append(f"❌ Empty or invalid HTML from: {link_url}")
                continue
            
            # Check for error pages
            html2_lower = html2.lower()
            if 'error' in html2_lower[:500] or '404' in html2_lower[:500] or 'not found' in html2_lower[:500]:
                errors.append(f"❌ Error page detected: {link_url}")
                continue
            
            cleaned2 = cleanup_html(html2)
            
            # Validate cleaned content (meaningful length and not challenge/verification page)
            if cleaned2 and len(cleaned2.strip()) >= _MIN_CONTENT_LEN and not _cleaned_text_is_challenge(cleaned2):
                # Better page header with clear separation
                page_header = f"\n--- PAGE: {link_url} ---\n\n"
                page_content = page_header + cleaned2
                # Check if adding this page would exceed max_chars
                if total_chars + len(page_content) + separator_len <= max_chars:
                    results.append(page_content)
                    total_chars += len(page_content) + separator_len
                else:
                    # Add partial content if there's space (truncate at word boundary)
                    available_chars = max_chars - total_chars - separator_len - len(page_header)
                    if available_chars > 100:  # Only add if meaningful space remains
                        truncated_content = _truncate_at_word(cleaned2, available_chars)
                        results.append(page_header + truncated_content)
                        total_chars = max_chars
                    break  # No more space
            else:
                if not cleaned2 or len(cleaned2.strip()) == 0:
                    errors.append(f"❌ No extractable text content on page: {link_url}")
                elif _cleaned_text_is_challenge(cleaned2):
                    errors.append(f"❌ Challenge/verification page on: {link_url}")
                else:
                    errors.append(f"❌ Insufficient content from {link_url} (under {_MIN_CONTENT_LEN} characters).")

    if not results:
        return errors[0] if errors else f"❌ Unknown error on site: {url}"
    
    # Join results with separator
    final_text = separator.join(results)
    # Final guard: if total content is still under minimum, treat as bad result
    if len(final_text.strip()) < _MIN_CONTENT_LEN:
        return errors[0] if errors else f"❌ Insufficient content (under {_MIN_CONTENT_LEN} characters); likely error or blocked page. Try again later."
    
    # Add errors at the end if there's space (errors are usually short)
    if errors and len(final_text) + len("\n".join(errors)) + 20 <= max_chars:
        final_text += "\n\n" + "\n".join(errors)
    
    # Final safety check - ensure we don't exceed max_chars (improved truncation)
    if len(final_text) > max_chars:
        # Try to cut at a reasonable point (prefer sentence boundaries)
        truncated = final_text[:max_chars]
        
        # Look for sentence endings (period, exclamation, question mark followed by space/newline)
        sentence_endings = []
        for i in range(len(truncated) - 1, max(0, len(truncated) - 200), -1):
            if truncated[i] in '.!?' and i < len(truncated) - 1:
                next_char = truncated[i + 1]
                if next_char in ' \n\t':
                    sentence_endings.append(i + 1)
                    if len(sentence_endings) >= 3:  # Get a few options
                        break
        
        # Also look for paragraph boundaries
        paragraph_endings = []
        for i in range(len(truncated) - 1, max(0, len(truncated) - 200), -1):
            if truncated[i] == '\n' and i < len(truncated) - 1:
                if truncated[i + 1] == '\n' or (i > 0 and truncated[i - 1] == '\n'):
                    paragraph_endings.append(i + 1)
                    if len(paragraph_endings) >= 2:
                        break
        
        # Choose best cut point
        best_cut = len(truncated)
        if sentence_endings:
            # Use the last sentence ending that's not too early
            for end_pos in reversed(sentence_endings):
                if end_pos >= max_chars * 0.85:  # At least 85% of max_chars
                    best_cut = end_pos
                    break
        elif paragraph_endings:
            # Fall back to paragraph boundary
            for end_pos in reversed(paragraph_endings):
                if end_pos >= max_chars * 0.85:
                    best_cut = end_pos
                    break
        # Fallback: truncate at word boundary to avoid mid-word cuts
        if best_cut == len(truncated):
            truncated = _truncate_at_word(final_text, max_chars)
            best_cut = len(truncated)
        
        final_text = truncated[:best_cut].rstrip() + "\n\n[... content truncated to fit character limit ...]"
    
    return final_text

# -------------------------
# Worker + Writer
# -------------------------


def build_lead_data_for_row(df: pd.DataFrame, orig_row: int, lead_cols: dict, has_headers: bool, url: str) -> dict:
    """Build lead_data dict from a single row (for large runs to avoid storing 300k dicts in memory)."""
    lead_data = {'url': url}
    default_company = url.replace('https://', '').replace('http://', '').split('/')[0] if url else ''
    for key, col_name in (lead_cols or {}).items():
        try:
            ci = list(df.columns).index(col_name) if has_headers else int(str(col_name).replace("Column ", "")) - 1
            val = df.iloc[orig_row, ci] if orig_row < len(df) else None
            v = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val).strip()
            lead_data[key] = v
        except (ValueError, TypeError, IndexError):
            lead_data[key] = ""
    if not lead_data.get('company_name'):
        lead_data['company_name'] = default_company
    return lead_data


async def worker_coroutine(name, session, url_queue: asyncio.Queue, result_queue: asyncio.Queue,
                           depth, keywords, max_chars, retries, timeout,
                           ai_enabled=False, ai_api_key=None, ai_provider=None, ai_model=None, ai_prompt=None,
                           email_copy_enabled=False, email_copy_api_key=None, email_copy_provider=None,
                           email_copy_model=None, email_copy_prompt=None,
                           lead_data_map=None, ai_status_callback=None, scrape_status_callback=None, fast_mode: bool = False,
                           use_playwright_fallback: bool = False, use_common_crawl_fallback: bool = False,
                           total_urls: int = 0,
                           original_df=None, url_list_with_idx=None, csv_config: dict = None):
    from urllib.parse import urlparse
    import random
    
    worker_id = int(name.split('-')[-1]) if '-' in name else 0
    initial_delay = worker_id * (random.uniform(0.03, 0.12) if fast_mode else random.uniform(0.5, 1.5))
    await asyncio.sleep(initial_delay)
    
    while True:
        item = await url_queue.get()
        if item is None:
            url_queue.task_done()
            break
        
        # CRITICAL: Wrap entire processing in try-finally to ensure task_done() is always called
        try:
            await asyncio.sleep(random.uniform(0.03, 0.15) if fast_mode else random.uniform(0.5, 2.0))
            # Handle both old format (just URL) and new format (URL, index)
            if isinstance(item, tuple):
                original_url, url_index = item
            else:
                original_url = item
                url_index = None
            
            # Normalize URL (add https:// if missing)
            normalized_url = normalize_url(original_url)
            
            # Get lead data: from map, or on-demand from dataframe for large runs (avoids 300k dicts in memory)
            lead_data = None
            if lead_data_map and url_index is not None and url_index in lead_data_map:
                try:
                    ld = lead_data_map[url_index]
                    lead_data = dict(ld) if ld is not None and hasattr(ld, 'keys') else {}
                except (TypeError, AttributeError, ValueError):
                    lead_data = {}
                lead_data['url'] = original_url  # Ensure URL is set
            elif (original_df is not None and url_list_with_idx and csv_config is not None
                  and url_index is not None and 0 <= url_index < len(url_list_with_idx)):
                try:
                    _, orig_row = url_list_with_idx[url_index]
                    lead_data = build_lead_data_for_row(
                        original_df, orig_row,
                        csv_config.get('lead_cols') or {},
                        csv_config.get('has_headers', True),
                        original_url)
                except Exception:
                    lead_data = {'url': original_url, 'company_name': original_url.replace('https://', '').replace('http://', '').split('/')[0]}
            else:
                # Fallback: extract company name from URL
                lead_data = {
                    'url': original_url,
                    'company_name': original_url.replace('https://', '').replace('http://', '').split('/')[0]
                }
            
            def _scrape_status(status: str, msg: str):
                if scrape_status_callback:
                    try:
                        scrape_status_callback(original_url, status, msg)
                    except Exception:
                        pass

            _scrape_status("scraping", "Fetching...")
            # Be lenient with URL validation - let the actual HTTP request determine validity
            # Only reject obviously invalid URLs (empty or completely malformed)
            if not normalized_url or len(normalized_url.strip()) < 4:
                scraped_text = "❌ Invalid URL: URL is empty or too short"
                ai_summary = "❌ Invalid URL: URL is empty or too short"
                email_copy = ""
                _scrape_status("error", scraped_text)
            else:
                # Global timeout per URL - high tier gets more time so slow sites + cache fallback succeed
                fast_cap = 240 if (total_urls or 0) >= 500 else 120  # was 180/90; 240s gives cache fallback more room
                max_total_time = min((timeout * (retries + 1) * (depth + 1) * 2) + (30 if fast_mode else 60), fast_cap if fast_mode else 300)
                
                try:
                    # Wrap scraping in a timeout to prevent infinite hangs
                    scraped_text = await asyncio.wait_for(
                        scrape_site(session, normalized_url, depth, keywords, max_chars, retries, timeout, fast_mode,
                                    use_playwright_fallback, use_common_crawl_fallback),
                        timeout=max_total_time
                    )
                except asyncio.TimeoutError:
                    url_short = normalized_url[:70] + ("…" if len(normalized_url) > 70 else "")
                    err_msg = (
                        f"❌ Timeout: full scrape for {url_short} exceeded {max_total_time}s. "
                        f"Site works in browser but not scraper — often blocked by anti-bot. "
                        f"Tip: 1) Increase Timeout in Step 2 Settings, 2) Install Playwright (pip install playwright && playwright install chromium) for browser fallback."
                    )
                    _scrape_status("error", err_msg)
                    try:
                        cache_result = await asyncio.wait_for(
                            _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback, quick_mode=True, playwright_full_time=True),
                            timeout=90
                        )
                    except asyncio.TimeoutError:
                        cache_result = ""
                    if isinstance(cache_result, tuple):
                        page_url, html = cache_result
                        scraped_text = _process_cached_html_to_text(page_url, html, max_chars)
                        if scraped_text and not _scraped_result_is_bad(scraped_text):
                            _scrape_status("scraped", f"Recovered from cache ({len(scraped_text):,} chars)")
                        else:
                            scraped_text = err_msg
                            _scrape_status("error", scraped_text)
                    else:
                        scraped_text = err_msg
                        _scrape_status("error", scraped_text)
                except Exception as e:
                    # If scraping fails with an exception, try once more with a more lenient approach
                    try:
                        # Try with http:// if https:// failed, but also wrap in timeout
                        if normalized_url.startswith("https://"):
                            fallback_url = normalized_url.replace("https://", "http://", 1)
                            try:
                                scraped_text = await asyncio.wait_for(
                                    scrape_site(session, fallback_url, depth, keywords, max_chars, retries, timeout, fast_mode,
                                                use_playwright_fallback, use_common_crawl_fallback),
                                    timeout=max_total_time
                                )
                            except asyncio.TimeoutError:
                                url_short = fallback_url[:70] + ("…" if len(fallback_url) > 70 else "")
                                scraped_text = (
                                    f"❌ Timeout: scrape for {url_short} exceeded {max_total_time}s. "
                                    f"Tip: increase Timeout in Step 2 Settings."
                                )
                        else:
                            scraped_text = f"❌ Error scraping {normalized_url}: {str(e)}"
                    except Exception as e2:
                        scraped_text = f"❌ Error scraping {normalized_url}: {str(e2)}"
                    if scraped_text and scraped_text.startswith("❌"):
                        _scrape_status("error", scraped_text)
                        try:
                            cache_result = await asyncio.wait_for(
                                _fetch_from_cache_fallbacks(session, normalized_url, timeout, use_playwright_fallback, use_common_crawl_fallback, quick_mode=True, playwright_full_time=True),
                                timeout=90
                            )
                        except asyncio.TimeoutError:
                            cache_result = ""
                        if isinstance(cache_result, tuple):
                            page_url, html = cache_result
                            cached_text = _process_cached_html_to_text(page_url, html, max_chars)
                            if cached_text and not _scraped_result_is_bad(cached_text):
                                scraped_text = cached_text
                                _scrape_status("scraped", f"Recovered from cache ({len(scraped_text):,} chars)")
                    if scraped_text and scraped_text.startswith("❌"):
                        _scrape_status("error", scraped_text)
                
                # Single gate: never treat bad content as success (short, challenge, or error-like)
                if scraped_text and not scraped_text.startswith("❌") and _scraped_result_is_bad(scraped_text):
                    scraped_text = "❌ Insufficient content (under 200 characters); likely error or blocked page. Try again later."
                    _scrape_status("error", scraped_text)
                
                # RELAXED VALIDATION: Only flag obvious mismatches
                # Many sites legitimately redirect to different domains (same organization)
                # Only check for obvious content mismatches, not domain redirects
                if scraped_text and not scraped_text.startswith("❌"):
                    # Check if scraped content contains PAGE: header
                    page_header_match = re.search(r'PAGE:\s*(https?://[^\s\n]+)', scraped_text[:5000])
                    if page_header_match:
                        actual_scraped_url = page_header_match.group(1)
                        # Normalize both URLs for comparison
                        actual_domain = urlparse(actual_scraped_url).netloc.replace('www.', '').lower()
                        expected_domain = urlparse(normalized_url).netloc.replace('www.', '').lower()
                        
                        # Extract base domains (last two parts)
                        actual_base = '.'.join(actual_domain.split('.')[-2:]) if '.' in actual_domain else actual_domain
                        expected_base = '.'.join(expected_domain.split('.')[-2:]) if '.' in expected_domain else expected_domain
                        
                        # Only flag if:
                        # 1. Base domains are completely different (not related)
                        # 2. AND it's not a subdomain variation
                        # 3. AND content is suspiciously short (might be wrong page)
                        if actual_base != expected_base:
                            if not (actual_domain.endswith('.' + expected_domain) or expected_domain.endswith('.' + actual_domain)):
                                # Check if content is suspiciously short (might be wrong page)
                                if len(scraped_text.strip()) < _SHORT_CONTENT_FOR_MISMATCH:
                                    # Very short content from different domain - likely wrong
                                    scraped_text = f"❌ CONTENT MISMATCH: Scraped content is from {actual_domain} but expected {expected_domain}. Original URL: {original_url}"
                                    _scrape_status("error", scraped_text)
                                # Otherwise, allow it - many sites redirect to related domains
                if scraped_text and not scraped_text.startswith("❌"):
                    _scrape_status("scraped", f"{len(scraped_text):,} chars scraped")
                
                # Generate AI summary if enabled (skip when scraping already failed or insufficient content)
                if ai_enabled and ai_api_key and ai_provider and ai_model:
                    if not scraped_text or scraped_text.startswith("❌") or len(scraped_text.strip()) < _MIN_CONTENT_LEN:
                        ai_summary = "❌ Summary skipped: site unreachable or insufficient content"
                    else:
                        _scrape_status("ai_summarizing", "Generating AI summary...")
                        def url_status_callback(msg):
                            if scrape_status_callback:
                                try:
                                    scrape_status_callback(original_url, "ai_summarizing", msg)
                                except Exception:
                                    pass
                            if ai_status_callback:
                                ai_status_callback(original_url, msg)
                        try:
                            ai_summary = await asyncio.wait_for(
                                generate_ai_summary(
                                    ai_api_key, ai_provider, ai_model, ai_prompt or "",
                                    lead_data, scraped_text, status_callback=url_status_callback
                                ),
                                timeout=120
                            )
                        except asyncio.TimeoutError:
                            ai_summary = "❌ AI Summary timeout: Generation exceeded 2 minutes"
                            _scrape_status("error", ai_summary)
                        except Exception as e:
                            ai_summary = f"❌ AI Summary error: {str(e)}"
                            _scrape_status("error", ai_summary)
                        else:
                            if ai_summary and ai_summary.startswith("❌"):
                                _scrape_status("error", ai_summary)
                else:
                    ai_summary = ""
                # Generate email copy if enabled (skip when scraping already failed or insufficient content)
                if email_copy_enabled and email_copy_api_key and email_copy_provider and email_copy_model:
                    if not scraped_text or scraped_text.startswith("❌") or len(scraped_text.strip()) < _MIN_CONTENT_LEN:
                        email_copy = "❌ Email skipped: site unreachable or insufficient content"
                    else:
                        _scrape_status("email_copy", "Generating email copy...")
                        def email_status_callback(msg):
                            if scrape_status_callback:
                                try:
                                    scrape_status_callback(original_url, "email_copy", msg)
                                except Exception:
                                    pass
                        try:
                            email_copy = await asyncio.wait_for(
                                generate_email_copy(
                                    email_copy_api_key, email_copy_provider, email_copy_model,
                                    email_copy_prompt or "", lead_data, scraped_text,
                                    status_callback=email_status_callback
                                ),
                                timeout=120
                            )
                        except asyncio.TimeoutError:
                            email_copy = "❌ Email copy timeout: Generation exceeded 2 minutes"
                            _scrape_status("error", email_copy)
                        except Exception as e:
                            email_copy = f"❌ Email copy error: {str(e)}"
                            _scrape_status("error", email_copy)
                        else:
                            if email_copy and email_copy.startswith("❌"):
                                _scrape_status("error", email_copy)
                else:
                    email_copy = ""
            
            # Reject bad results (short or challenge pages) so they never get written as success
            if scraped_text and _scraped_result_is_bad(scraped_text):
                scraped_text = f"❌ Insufficient content (under {_MIN_CONTENT_LEN} characters) or blocked/verification page. Try again later."
                _scrape_status("error", scraped_text)

            # PERFECT CSV CLEANING - Zero tolerance for formatting errors
            def clean_csv_field(field_value):
                """
                Perfect CSV field cleaning:
                - Removes all newlines, carriage returns, tabs
                - Removes null bytes and control characters
                - Normalizes whitespace
                - Ensures valid UTF-8 encoding
                - Returns empty string for None/invalid values
                """
                if field_value is None:
                    return ""
                
                # Convert to string
                field_str = str(field_value)
                
                # Remove null bytes and control characters (except space, tab)
                # Keep only printable characters and whitespace
                import re
                # Remove null bytes
                field_str = field_str.replace('\x00', '')
                # Replace all newlines, carriage returns, form feeds, vertical tabs with space
                field_str = re.sub(r'[\n\r\f\v]', ' ', field_str)
                # Replace tabs with space
                field_str = field_str.replace('\t', ' ')
                # Remove other control characters (0x00-0x1F except space)
                field_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', field_str)
                
                # Normalize whitespace: collapse multiple spaces/tabs into single space
                field_str = re.sub(r'\s+', ' ', field_str)
                
                # Strip leading/trailing whitespace
                field_str = field_str.strip()
                
                # Ensure valid UTF-8 encoding (replace invalid sequences)
                try:
                    field_str = field_str.encode('utf-8', errors='replace').decode('utf-8')
                except:
                    field_str = ""
                
                return field_str
            
            # Final gate: never enqueue bad content as success (defense in depth)
            if scraped_text and not scraped_text.startswith("❌") and _scraped_result_is_bad(scraped_text):
                scraped_text = "❌ Insufficient content (under 200 characters); likely error or blocked page. Try again later."
            # Clean all fields perfectly; normalize AI/email errors for clearer CSV display
            ai_for_csv = _normalize_ai_error_for_display(ai_summary, "Summary")
            email_for_csv = _normalize_ai_error_for_display(email_copy, "Email") if email_copy else ""
            cleaned_url = clean_csv_field(original_url)
            cleaned_scraped_text = clean_csv_field(scraped_text)
            cleaned_ai_summary = clean_csv_field(ai_for_csv)
            cleaned_email_copy = clean_csv_field(email_for_csv) if email_copy_enabled else ""
            # Output: 5 columns with row_index when email copy enabled, else 4 (includes url_index)
            # Format: (url, scraped_text, ai_summary, email_copy, url_index)
            row_idx = url_index if url_index is not None else -1
            if email_copy_enabled:
                out = (cleaned_url, cleaned_scraped_text, cleaned_ai_summary, cleaned_email_copy, row_idx)
            else:
                out = (cleaned_url, cleaned_scraped_text, cleaned_ai_summary, row_idx)
            await result_queue.put(out)
        except Exception as e:
            # If anything goes wrong, still output an error row
            try:
                # Helper function for cleaning (defined here since it might not be in scope)
                def clean_csv_field(field_value):
                    if field_value is None:
                        return ""
                    field_str = str(field_value)
                    import re
                    field_str = field_str.replace('\x00', '')
                    field_str = re.sub(r'[\n\r\f\v]', ' ', field_str)
                    field_str = field_str.replace('\t', ' ')
                    field_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', field_str)
                    field_str = re.sub(r'\s+', ' ', field_str)
                    field_str = field_str.strip()
                    try:
                        field_str = field_str.encode('utf-8', errors='replace').decode('utf-8')
                    except:
                        field_str = ""
                    return field_str
                
                original_url_val = original_url if 'original_url' in locals() else "unknown URL"
                row_idx = url_index if 'url_index' in locals() and url_index is not None else -1
                error_msg = f"❌ Worker error processing {original_url_val}: {str(e)}"
                cleaned_url = clean_csv_field(original_url_val)
                cleaned_scraped_text = clean_csv_field(error_msg)
                cleaned_ai_summary = clean_csv_field("")
                cleaned_email_copy = clean_csv_field("")
                if email_copy_enabled:
                    out = (cleaned_url, cleaned_scraped_text, cleaned_ai_summary, cleaned_email_copy, row_idx)
                else:
                    out = (cleaned_url, cleaned_scraped_text, cleaned_ai_summary, row_idx)
                await result_queue.put(out)
            except:
                try:
                    row_idx = url_index if 'url_index' in locals() and url_index is not None else -1
                    if email_copy_enabled:
                        await result_queue.put(("", f"❌ Critical worker error: {str(e)}", "", "", row_idx))
                    else:
                        await result_queue.put(("", f"❌ Critical worker error: {str(e)}", "", row_idx))
                except:
                    pass  # If even this fails, just continue
        finally:
            # CRITICAL: Always mark task as done, even if there was an error
            url_queue.task_done()


async def writer_coroutine(result_queue: asyncio.Queue, rows_per_file: int, output_dir: str,
                           total_urls: int, progress_callback,
                           start_part: int = 0, checkpoint_data: dict | None = None,
                           checkpoint_path: str | None = None,
                           log_callback=None, include_email_copy: bool = False,
                           original_df: pd.DataFrame = None, csv_config: dict = None,
                           use_database: bool = True):
    """
    Write results to disk and/or database. If checkpoint_data and checkpoint_path are provided,
    saves progress after each file write for crash recovery and resume.
    
    If original_df is provided, merges scraped data with original CSV columns.
    
    For large datasets (150k+ leads), use_database=True enables SQLite persistence
    which handles results in memory-efficient chunks.
    """
    def emit(msg: str):
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode("ascii", errors="replace").decode("ascii"))
        if log_callback:
            try:
                log_callback(msg)
            except Exception:
                pass

    # CRITICAL: Create output directory first
    try:
        os.makedirs(output_dir, exist_ok=True)
        emit(f"📁 Writer: Created output directory: {output_dir}")
    except Exception as e:
        emit(f"❌ Writer: Failed to create output directory: {e}")
        import traceback
        traceback.print_exc()
        return  # Exit if we can't create directory
    
    # Initialize database for large dataset support (lower threshold in cloud for memory safety)
    db = None
    db_threshold = 500 if is_cloud_mode() else 1000
    if use_database and total_urls > db_threshold:
        try:
            db = init_scrape_db()
            emit(f"🗄️ Writer: Initialized SQLite database for {total_urls:,} URLs")
        except Exception as e:
            emit(f"⚠️ Writer: Could not initialize database: {e}")
    
    # Prepare original DataFrame for merging if provided
    has_original_df = original_df is not None and not original_df.empty
    url_col = csv_config.get('url_column') if csv_config else None
    has_headers = csv_config.get('has_headers', True) if csv_config else True
    # URL column index for no-headers CSV (Column 1 -> 0)
    if csv_config and not has_headers and url_col is not None:
        try:
            url_col_idx = int(str(url_col).replace("Column ", "").strip()) - 1
            url_col_idx = max(0, min(url_col_idx, len(original_df.columns) - 1)) if has_original_df else 0
        except (ValueError, TypeError):
            url_col_idx = 0
    else:
        url_col_idx = 0
    
    if has_original_df:
        # Make a copy to avoid modifying the original
        enriched_df = original_df.copy()
        # Add new columns for scraped data
        enriched_df['ScrapedText'] = ''
        enriched_df['CompanySummary'] = ''
        if include_email_copy:
            enriched_df['EmailCopy'] = ''
        emit(f"📊 Writer: Will merge results with original CSV ({len(enriched_df)} rows, {len(enriched_df.columns)} columns)")
    else:
        enriched_df = None
    
    loop = asyncio.get_running_loop()
    buffer = []
    part = start_part
    processed = 0
    files_written = []
    last_flush_ts = time.time()
    # In cloud, flush more often to minimize data loss if session drops
    emergency_flush_every_s = 3 if is_cloud_mode() else 5
    emergency_flush_min_rows = max(10, min(30, rows_per_file // 25)) if is_cloud_mode() else max(15, min(50, rows_per_file // 20))

    while True:
        try:
            item = await result_queue.get()
        except BaseException as e:
            emit(f"⚠️ Writer: result_queue.get error (non-fatal): {e}")
            continue
        if item is None:
            emit(f"📝 Writer: Received stop signal. Processed {processed}/{total_urls}. Buffer has {len(buffer)}.")
            result_queue.task_done()
            break
        buffer.append(item)
        processed += 1
        result_queue.task_done()  # CRITICAL: one per item (join() counts these)
        if progress_callback:
            try:
                progress_callback(processed, total_urls)
            except Exception:
                pass

        try:
            # Save to database for large datasets (memory-efficient)
            if db and buffer:
                for row in buffer:
                    if len(row) >= 2:
                        url = str(row[0]).strip() if row[0] else ""
                        scraped_text = str(row[1]) if row[1] else ""
                        ai_summary = str(row[2]) if len(row) >= 3 and row[2] else ""
                        email_copy = str(row[3]) if include_email_copy and len(row) >= 4 and row[3] else ""
                        row_index = row[4] if len(row) >= 5 and row[4] is not None else processed - 1
                        
                        # Determine result status
                        result_status = 'success' if scraped_text and not scraped_text.startswith('❌') else 'failed'
                        if scraped_text and (scraped_text.startswith('⚠️') or scraped_text.startswith('No relevant content')):
                            result_status = 'partial'
                        
                        try:
                            db.insert_result(
                                url=url,
                                row_index=row_index,
                                scraped_text=scraped_text,
                                company_summary=ai_summary,
                                email_copy=email_copy,
                                result_status=result_status
                            )
                        except Exception as e:
                            emit(f"⚠️ Database: Error saving result for {url[:50]}: {e}")

            # Process results and update enriched_df if available
            if has_original_df and buffer:
                for row in buffer:
                    if len(row) >= 3:
                        url = str(row[0]).strip() if row[0] else ""
                        scraped_text = str(row[1]) if row[1] else ""
                        ai_summary = str(row[2]) if row[2] else ""
                        email_copy = str(row[3]) if include_email_copy and len(row) >= 4 else ""
                        
                        if url:
                            # Find matching row in original DataFrame by URL
                            try:
                                if has_headers and url_col and url_col in enriched_df.columns:
                                    url_series = enriched_df[url_col].astype(str).str.strip()
                                else:
                                    url_series = enriched_df.iloc[:, url_col_idx].astype(str).str.strip()
                                match_mask = url_series == url
                                if match_mask.any():
                                    idx = match_mask.idxmax()
                                    enriched_df.at[idx, 'ScrapedText'] = scraped_text
                                    enriched_df.at[idx, 'CompanySummary'] = ai_summary
                                    if include_email_copy:
                                        enriched_df.at[idx, 'EmailCopy'] = email_copy
                            except Exception as e:
                                emit(f"⚠️ Writer: Error merging row for {url[:50]}: {e}")
                
                # Clear buffer after processing (we've merged into enriched_df)
                buffer = []
            
            # Write chunks to disk when:
            # 1) regular threshold is reached, or
            # 2) enough time passed with partial buffer (prevents progress loss on sudden refresh/crash)
            # NOTE: When using original_df, we don't chunk - we save at the end
            while len(buffer) >= rows_per_file or (
                len(buffer) >= emergency_flush_min_rows and (time.time() - last_flush_ts) >= emergency_flush_every_s and not has_original_df
            ):
                part += 1
                chunk_size = rows_per_file if len(buffer) >= rows_per_file else len(buffer)
                chunk_rows = buffer[:chunk_size]
                buffer = buffer[chunk_size:]
                last_flush_ts = time.time()
                if chunk_size < rows_per_file:
                    emit(f"⚠️ Early flush: writing {chunk_size} buffered rows to reduce loss risk")
                
                try:
                    # Filter out ONLY truly empty rows (keep error messages as they are valid data)
                    filtered_rows = []
                    for row in chunk_rows:
                        if len(row) >= 2:
                            url = str(row[0]).strip() if row[0] else ""
                            scraped_text = str(row[1]).strip() if row[1] else ""
                            # Include rows with valid URL - keep error messages (they start with ❌)
                            # Only exclude rows where URL is empty or scraped_text is completely empty
                            if url and url != "":
                                # Include even if scraped_text is empty or is an error message
                                # Error messages are valid data that should be preserved
                                filtered_rows.append(row)
                    
                    if not filtered_rows:
                        # Skip writing if no valid rows
                        continue
                    
                    # Normalize rows: 4 columns when include_email_copy, else 3
                    n_cols = 4 if include_email_copy else 3
                    cols = ["Website", "ScrapedText", "CompanySummary", "EmailCopy"] if include_email_copy else ["Website", "ScrapedText", "CompanySummary"]
                    normalized_rows = []
                    for row in filtered_rows:
                        normalized_row = list(row[:n_cols])
                        while len(normalized_row) < n_cols:
                            normalized_row.append("")
                        normalized_rows.append(tuple(normalized_row[:n_cols]))
                    
                    df = pd.DataFrame(normalized_rows, columns=cols)
                    
                    # Normalize scrape/AI/email error strings so narrow columns show clear messages
                    if "ScrapedText" in df.columns:
                        df["ScrapedText"] = df["ScrapedText"].apply(
                            lambda x: _normalize_scrape_error_for_display(str(x) if pd.notna(x) else "")
                        )
                    if "CompanySummary" in df.columns:
                        df["CompanySummary"] = df["CompanySummary"].apply(
                            lambda x: _normalize_ai_error_for_display(str(x) if pd.notna(x) else "", "Summary")
                        )
                    if include_email_copy and "EmailCopy" in df.columns:
                        df["EmailCopy"] = df["EmailCopy"].apply(
                            lambda x: _normalize_ai_error_for_display(str(x) if pd.notna(x) else "", "Email")
                        )
                    
                    # PERFECT CSV CLEANING - Clean all DataFrame columns before writing
                    def clean_dataframe_for_csv(df):
                        """Clean all string columns in DataFrame for perfect CSV formatting"""
                        import re
                        for col in df.columns:
                            # Convert to string, handle None/NaN
                            df[col] = df[col].astype(str).replace('nan', '').replace('None', '')
                            # Remove null bytes
                            df[col] = df[col].str.replace('\x00', '', regex=False)
                            # Replace newlines, carriage returns, tabs with space
                            df[col] = df[col].str.replace(r'[\n\r\f\v\t]', ' ', regex=True)
                            # Remove other control characters
                            df[col] = df[col].str.replace(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', regex=True)
                            # Normalize whitespace
                            df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
                            # Strip whitespace
                            df[col] = df[col].str.strip()
                            # Replace empty strings with empty string (not NaN)
                            df[col] = df[col].replace('', '')
                        return df
                    
                    # Clean DataFrame
                    df = clean_dataframe_for_csv(df.copy())
                    
                    # Remove ONLY rows with empty URLs (keep error messages)
                    df = df[df["Website"].astype(str).str.strip() != ""]
                    
                    # Enforce column structure
                    for c in cols:
                        if c not in df.columns:
                            df[c] = ""
                    df = df[cols]
                    
                    if len(df) == 0:
                        # Skip writing if DataFrame is empty after filtering
                        continue
                    
                    # CRITICAL: Write Excel file FIRST (more reliable than CSV)
                    excel_path = os.path.join(output_dir, f"output_part_{part}.xlsx")
                    excel_written = False
                    try:
                        df_excel = df[cols].copy()
                        from openpyxl import Workbook
                        wb = Workbook()
                        ws = wb.active
                        _write_excel_sheet(ws, cols, df_excel)
                        
                        # Run heavy I/O in executor to avoid blocking event loop (fixes progress freeze)
                        def _save_excel():
                            wb.save(excel_path)
                        try:
                            await loop.run_in_executor(None, _save_excel)
                        except Exception as save_error:
                            # If save fails, try to save with minimal data
                            print(f"⚠️ Excel save failed, attempting recovery: {save_error}")
                            # Create a fresh workbook with just headers
                            wb_recovery = Workbook()
                            ws_recovery = wb_recovery.active
                            ws_recovery.title = "Scraped Data"
                            ws_recovery.append(cols)
                            wb_recovery.save(excel_path)
                            raise save_error
                        
                        # Verify Excel file was written
                        if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                            files_written.append(excel_path)
                            excel_written = True
                            emit(f"✅ Writer: Wrote Excel file: {os.path.basename(excel_path)} ({len(df_excel)} rows)")
                        else:
                            emit(f"❌ Writer: Excel file missing/empty: {os.path.basename(excel_path)}")
                    except Exception as e:
                        # If Excel fails, log error but continue
                        emit(f"⚠️ Writer: Excel export failed for part {part}: {e}")
                        import traceback
                        traceback.print_exc()
                    
                    # PERFECT CSV WRITING - Zero tolerance for errors
                    csv_path = os.path.join(output_dir, f"output_part_{part}.csv")
                    
                    # Write CSV with perfect formatting:
                    # - UTF-8 with BOM for Excel compatibility
                    # - QUOTE_ALL: All fields quoted to handle commas, quotes, special chars
                    # - lineterminator='\n': Standard Unix line endings
                    # - doublequote=True: Escape quotes by doubling them (CSV standard)
                    try:
                        df = df[cols]
                        def _write_csv():
                            with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                                writer = csv.writer(f, quoting=csv.QUOTE_ALL, doublequote=True, lineterminator='\n', quotechar='"')
                                writer.writerow(cols)
                                for idx in range(len(df)):
                                    row = df.iloc[idx]
                                    def clean_val(v):
                                        if not v: return ""
                                        v = str(v).replace('\x00', '')
                                        v = re.sub(r'[\n\r\f\v\t]', ' ', v)
                                        v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', v)
                                        return re.sub(r'\s+', ' ', v).strip()
                                    row_values = [clean_val(row[c]) for c in cols]
                                    row_values = (row_values + [""] * len(cols))[:len(cols)]
                                    writer.writerow(row_values)
                        await loop.run_in_executor(None, _write_csv)
                        
                        # Verify file was written
                        if os.path.exists(csv_path) and os.path.getsize(csv_path) > 0:
                            files_written.append(csv_path)
                            emit(f"✅ Writer: Wrote CSV file: {os.path.basename(csv_path)} ({len(df)} rows)")
                        else:
                            emit(f"❌ Writer: CSV file missing/empty: {os.path.basename(csv_path)}")
                        
                        # VALIDATION: Verify CSV file is valid by reading it back
                        try:
                            test_df = pd.read_csv(csv_path, encoding='utf-8-sig', quoting=csv.QUOTE_ALL, engine='python')
                            # Check row count matches
                            if len(test_df) != len(df):
                                raise ValueError(f"CSV validation failed: row count mismatch (expected {len(df)}, got {len(test_df)})")
                            if len(test_df.columns) != len(cols):
                                raise ValueError(f"CSV validation failed: column count is {len(test_df.columns)}, expected {len(cols)}. Columns: {list(test_df.columns)}")
                            if list(test_df.columns) != cols:
                                raise ValueError(f"CSV validation failed: column names don't match. Expected {cols}, got {list(test_df.columns)}")
                        except Exception as e:
                            # If validation fails, log but don't rewrite (already used manual writer)
                            import logging
                            logging.warning(f"CSV validation warning: {e}")
                            # Also print to console for debugging
                            emit(f"⚠️ CSV validation warning for {os.path.basename(csv_path)}: {e}")
                        # CRASH RECOVERY: Update checkpoint after each successful write
                        if checkpoint_data is not None and checkpoint_path and len(df) > 0:
                            try:
                                urls_in_chunk = df["Website"].astype(str).str.strip().tolist()
                                checkpoint_data.setdefault("completed_urls", set()).update(u for u in urls_in_chunk if u)
                                checkpoint_data["last_part"] = part
                                save_checkpoint(checkpoint_data, checkpoint_path)
                            except Exception as cp_err:
                                print(f"⚠️ Checkpoint update failed: {cp_err}")
                    except Exception as e:
                        import logging
                        logging.error(f"CSV write failed: {e}. Using manual writer...")
                        df = df[cols]
                        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                            writer = csv.writer(f, quoting=csv.QUOTE_ALL, doublequote=True, lineterminator='\n')
                            writer.writerow(cols)
                            for idx in range(len(df)):
                                row = df.iloc[idx]
                                row_values = ["" if pd.isna(row[c]) else str(row[c]) for c in cols]
                                writer.writerow(row_values)
                    
                    # Excel file already written above, skip duplicate writing
                    pass
                except Exception as e:
                    # If Excel fails for large file, at least save CSV with perfect formatting
                    try:
                        # Clean DataFrame first
                        def clean_dataframe_for_csv(df):
                            import re
                            for col in df.columns:
                                df[col] = df[col].astype(str).replace('nan', '').replace('None', '')
                                df[col] = df[col].str.replace('\x00', '', regex=False)
                                df[col] = df[col].str.replace(r'[\n\r\f\v\t]', ' ', regex=True)
                                df[col] = df[col].str.replace(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', regex=True)
                                df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
                                df[col] = df[col].str.strip()
                                df[col] = df[col].replace('', '')
                            return df
                        df = clean_dataframe_for_csv(df.copy())
                        
                        # CRITICAL: Ensure CompanySummary exists before writing
                        if "CompanySummary" not in df.columns:
                            df["CompanySummary"] = ""
                        df = df[["Website", "ScrapedText", "CompanySummary"]]
                        
                        # Use manual CSV writer for perfect quoting
                        # CRITICAL: Use iloc instead of iterrows() to prevent misalignment
                        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                            writer = csv.writer(f, quoting=csv.QUOTE_ALL, doublequote=True, lineterminator='\n', quotechar='"')
                            # ALWAYS write 3-column header
                            writer.writerow(["Website", "ScrapedText", "CompanySummary"])
                            
                            # CRITICAL: Use iloc to prevent iterrows() misalignment issues
                            for idx in range(len(df)):
                                row = df.iloc[idx]
                                row_values = []
                                
                                # Extract by column name to ensure correct order
                                website = "" if pd.isna(row["Website"]) else str(row["Website"])
                                scraped_text = "" if pd.isna(row["ScrapedText"]) else str(row["ScrapedText"])
                                company_summary = "" if pd.isna(row["CompanySummary"]) else str(row["CompanySummary"])
                                
                                # Clean values - CRITICAL: Remove newlines and normalize whitespace
                                def clean_csv_val(v):
                                    if not v:
                                        return ''
                                    import re
                                    v = str(v)
                                    v = v.replace('\x00', '')  # Remove null bytes
                                    v = v.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')  # Replace newlines/tabs with space
                                    v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', v)  # Remove control chars
                                    v = re.sub(r'\s+', ' ', v).strip()  # Normalize whitespace
                                    return v
                                
                                row_values.append(clean_csv_val(website))
                                row_values.append(clean_csv_val(scraped_text))
                                row_values.append(clean_csv_val(company_summary))
                                
                                # Ensure exactly 3 values
                                while len(row_values) < 3:
                                    row_values.append('')
                                row_values = row_values[:3]
                                
                                # CRITICAL: csv.writer with QUOTE_ALL will automatically quote all fields
                                # and escape internal quotes by doubling them (Excel standard)
                                writer.writerow(row_values)
                    except Exception:
                        # Ultimate fallback: manual writer - CRITICAL: Always write 3 columns
                        if "CompanySummary" not in df.columns:
                            df["CompanySummary"] = ""
                        df = df[["Website", "ScrapedText", "CompanySummary"]]
                        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                            writer = csv.writer(f, quoting=csv.QUOTE_ALL, doublequote=True, lineterminator='\n')
                            # ALWAYS write 3-column header
                            writer.writerow(["Website", "ScrapedText", "CompanySummary"])
                            # CRITICAL: Use iloc instead of iterrows() to prevent misalignment
                            for idx in range(len(df)):
                                row = df.iloc[idx]
                                # Extract by column name to ensure correct order, always 3 values
                                website = "" if pd.isna(row["Website"]) else str(row["Website"])
                                scraped_text = "" if pd.isna(row["ScrapedText"]) else str(row["ScrapedText"])
                                company_summary = "" if pd.isna(row["CompanySummary"]) else str(row["CompanySummary"])
                                writer.writerow([website, scraped_text, company_summary])
                except BaseException as write_err:
                    # CRITICAL: Never crash the writer - log and continue so progress is never lost
                    emit(f"⚠️ Writer: Error writing chunk (skipping to prevent crash): {write_err}")
                    import traceback
                    traceback.print_exc()
        except BaseException as writer_loop_err:
            try:
                emit(f"⚠️ Writer: Loop error (progress kept in buffer): {writer_loop_err}")
            except Exception:
                pass

    # Write remaining buffer (only for non-merged mode)
    if buffer and not has_original_df:
        # Filter out ONLY rows with empty URLs (keep error messages as they are valid data)
        filtered_buffer = []
        for row in buffer:
            if len(row) >= 2:
                url = str(row[0]).strip() if row[0] else ""
                scraped_text = str(row[1]).strip() if row[1] else ""
                # Include rows with valid URL - keep error messages (they start with ❌)
                # Only exclude rows where URL is empty
                if url and url != "":
                    n_c = 4 if include_email_copy else 3
                    norm = list(row[:n_c])
                    while len(norm) < n_c:
                        norm.append("")
                    filtered_buffer.append(tuple(norm[:n_c]))
        
        if filtered_buffer:
            part += 1
            cols_buf = ["Website", "ScrapedText", "CompanySummary", "EmailCopy"] if include_email_copy else ["Website", "ScrapedText", "CompanySummary"]
            try:
                df = pd.DataFrame(filtered_buffer, columns=cols_buf)
                
                # Remove ONLY rows with empty URLs (keep error messages)
                df = df[df["Website"].astype(str).str.strip() != ""]
                
                if len(df) == 0:
                    # Skip writing if DataFrame is empty after filtering
                    pass
                else:
                    csv_path = os.path.join(output_dir, f"output_part_{part}.csv")
                    df = df[cols_buf]
                    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                        writer = csv.writer(f, quoting=csv.QUOTE_ALL, doublequote=True, lineterminator='\n', quotechar='"')
                        writer.writerow(cols_buf)
                        
                        for idx in range(len(df)):
                            row = df.iloc[idx]
                            def clean_val(v):
                                if not v: return ""
                                import re
                                v = str(v).replace('\x00', '')
                                v = v.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                                v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', v)
                                return re.sub(r'\s+', ' ', v).strip()
                            row_values = [clean_val(row[c]) for c in cols_buf]
                            row_values = (row_values + [""] * len(cols_buf))[:len(cols_buf)]
                            writer.writerow(row_values)
                    
                    excel_path = os.path.join(output_dir, f"output_part_{part}.xlsx")
                    try:
                        df_excel_buf = df[cols_buf].copy()
                        from openpyxl import Workbook
                        wb = Workbook()
                        ws = wb.active
                        _write_excel_sheet(ws, cols_buf, df_excel_buf)
                        
                        # CRITICAL: Save with explicit error handling
                        try:
                            wb.save(excel_path)
                        except Exception as save_error:
                            print(f"⚠️ Excel save failed, attempting recovery: {save_error}")
                            wb_recovery = Workbook()
                            ws_recovery = wb_recovery.active
                            ws_recovery.title = "Scraped Data"
                            ws_recovery.append(cols_buf)
                            wb_recovery.save(excel_path)
                            raise save_error
                        
                        # Verify Excel file was written
                        if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                            files_written.append(excel_path)
                            print(f"✅ Writer: Successfully wrote Excel file: {excel_path} ({len(df)} rows)")
                        else:
                            print(f"❌ Writer: Excel file was not created or is empty: {excel_path}")
                    except Exception as e:
                        # If Excel fails, log but continue (CSV is more important)
                        print(f"⚠️ Writer: Excel export failed for part {part}: {e}")
                        import traceback
                        traceback.print_exc()
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                emit(f"❌ Writer: Failed to write output part {part}: {type(e).__name__}: {e}")
                for line in tb.strip().split("\n"):
                    emit(line)
    
    # When using original_df, write the enriched DataFrame at the end
    if has_original_df and enriched_df is not None:
        try:
            # Clean the enriched DataFrame
            def clean_enriched_df(df):
                import re
                for col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].astype(str).replace('nan', '').replace('None', '')
                        df[col] = df[col].str.replace('\x00', '', regex=False)
                        df[col] = df[col].str.replace(r'[\n\r\f\v\t]', ' ', regex=True)
                        df[col] = df[col].str.replace(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', regex=True)
                        df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
                        df[col] = df[col].str.strip()
                return df
            
            enriched_df = clean_enriched_df(enriched_df)
            
            # Write enriched CSV with strict formatting (no newlines in fields, QUOTE_ALL for Excel)
            enriched_csv_path = os.path.join(output_dir, "enriched_output.csv")
            cols = list(enriched_df.columns)
            with open(enriched_csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                w = csv.writer(f, quoting=csv.QUOTE_ALL, doublequote=True, lineterminator='\n', quotechar='"')
                w.writerow(cols)
                for i in range(len(enriched_df)):
                    row = enriched_df.iloc[i]
                    vals = []
                    for c in cols:
                        v = row[c] if c in row.index else ''
                        if pd.isna(v):
                            v = ''
                        v = str(v).replace('\x00', '').replace('\n', ' ').replace('\r', ' ')
                        v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', v)
                        v = re.sub(r'\s+', ' ', v).strip()
                        vals.append(v)
                    w.writerow(vals)
            files_written.append(enriched_csv_path)
            emit(f"✅ Writer: Wrote enriched CSV: {os.path.basename(enriched_csv_path)} ({len(enriched_df)} rows, {len(enriched_df.columns)} columns)")
            
            # Write enriched Excel
            enriched_excel_path = os.path.join(output_dir, "enriched_output.xlsx")
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Enriched Data"
                _write_excel_sheet(ws, list(enriched_df.columns), enriched_df)
                wb.save(enriched_excel_path)
                files_written.append(enriched_excel_path)
                emit(f"✅ Writer: Wrote enriched Excel: {os.path.basename(enriched_excel_path)}")
            except Exception as e:
                emit(f"⚠️ Writer: Excel export failed for enriched data: {e}")
        except Exception as e:
            import traceback
            emit(f"❌ Writer: Failed to write enriched output: {type(e).__name__}: {e}")
            traceback.print_exc()
    
    # Final summary
    emit(f"📊 Writer: Finished. Wrote {len(files_written)} files total.")
    if files_written:
        emit(f"📁 Writer: Files written: {', '.join([os.path.basename(f) for f in files_written])}")

# -------------------------
# Test preview (no file write)
# -------------------------


async def _collector_coroutine(result_queue: asyncio.Queue, results_list: list):
    """Collect results into a list instead of writing to disk."""
    while True:
        item = await result_queue.get()
        result_queue.task_done()
        if item is None:
            break
        results_list.append(item)


async def run_test_preview(urls: list, n: int, retries, timeout, depth, keywords, max_chars,
                           user_agent: str, ai_enabled: bool, ai_api_key, ai_provider, ai_model, ai_prompt,
                           email_copy_enabled: bool, email_copy_api_key, email_copy_provider,
                           email_copy_model, email_copy_prompt,
                           lead_data_map: dict, log_callback=None) -> list:
    """Run scraping + AI on first n URLs, return results for in-browser preview. No file output."""
    def emit(msg: str):
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode("ascii", errors="replace").decode("ascii"))
        if log_callback:
            try:
                log_callback(msg)
            except Exception:
                pass

    urls_subset = urls[:n]
    remaining = [(u, idx) for idx, u in enumerate(urls_subset)]
    results_list = []

    url_queue = asyncio.Queue()
    result_queue = asyncio.Queue(maxsize=n + 5)
    for u, idx in remaining:
        await url_queue.put((u, idx))

    connector = aiohttp.TCPConnector(limit=10, limit_per_host=3, ssl=False)
    session = aiohttp.ClientSession(
        headers={"User-Agent": user_agent},
        connector=connector,
        timeout=aiohttp.ClientTimeout(total=None),
        trust_env=True,
    )

    collector_task = asyncio.create_task(_collector_coroutine(result_queue, results_list))
    workers = [
        asyncio.create_task(worker_coroutine(
            f"test-worker-{i+1}", session, url_queue, result_queue, depth, keywords, max_chars,
            retries, timeout, ai_enabled, ai_api_key, ai_provider, ai_model, ai_prompt,
            email_copy_enabled, email_copy_api_key, email_copy_provider, email_copy_model, email_copy_prompt,
            lead_data_map, None, None, fast_mode=True, use_playwright_fallback=True, use_common_crawl_fallback=True))
        for i in range(min(2, n))
    ]

    emit(f"🧪 Test run: scraping + AI on {n} URL(s)...")
    try:
        await url_queue.join()
        for _ in workers:
            try:
                await url_queue.put(None)
            except Exception:
                pass
        await asyncio.gather(*workers)
        await result_queue.put(None)
        await asyncio.gather(collector_task)
    except Exception as e:
        emit(f"⚠️ Test error: {e}")
    finally:
        await session.close()

    return results_list


# -------------------------
# Runner
# -------------------------


async def run_scraper(urls, concurrency, retries, timeout, depth, keywords, max_chars,
                      user_agent, rows_per_file, output_dir, progress_callback,
                      ai_enabled=False, ai_api_key=None, ai_provider=None, ai_model=None, ai_prompt=None,
                      email_copy_enabled=False, email_copy_api_key=None, email_copy_provider=None,
                      email_copy_model=None, email_copy_prompt=None,
                      lead_data_map=None, ai_status_callback=None, scrape_status_callback=None,
                      run_folder: str = "", fast_mode: bool = False,
                      low_resource: bool = False, log_callback=None,
                      use_playwright_fallback: bool = True, use_common_crawl_fallback: bool = True,
                      original_df: pd.DataFrame = None, csv_config: dict = None, url_list_with_idx: list = None):
    """
    Run the scraper. Supports resume from checkpoint if output_dir contains a checkpoint file.
    Uses backpressure on result_queue to prevent memory exhaustion with large datasets.
    Saves checkpoint on every file write and in finally block for crash recovery.
    
    If original_df is provided, merges scraped results with original CSV data.
    """
    def emit(msg: str):
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode("ascii", errors="replace").decode("ascii"))
        if log_callback:
            try:
                log_callback(msg)
            except Exception:
                pass

    if not urls:
        emit("❌ No URLs to process. Exiting.")
        return

    total_urls_count = len(urls)

    # Cloud: tiered by run size — small/medium runs get normal speed; large runs stay gentle
    if is_cloud_mode():
        if total_urls_count < 5000:
            low_resource = False
            concurrency = min(concurrency, 20)
            emit(f"☁️ Cloud: balanced settings for {total_urls_count:,} URLs (concurrency={concurrency}).")
        else:
            low_resource = True
            concurrency = min(concurrency, _cloud_concurrency_max())
            emit(f"☁️ Cloud: gentle settings for {total_urls_count:,} URLs (concurrency={concurrency}) for stability.")
    else:
        # Local: settings already tuned by get_auto_run_settings from system tier (RAM/CPU)
        tier = get_system_tier()
        ram_gb = _get_system_ram_gb()
        try:
            import psutil
            cpus = psutil.cpu_count(logical=True) or os.cpu_count() or 0
        except Exception:
            cpus = os.cpu_count() or 0
        if total_urls_count >= 200000:
            ram_str = f"{ram_gb:.0f}GB" if ram_gb >= 1 else "?"
            emit(f"🖥️ Local: {total_urls_count:,} URLs — tier={tier} (RAM≈{ram_str}, {cpus} CPUs). Concurrency={concurrency} — full force on capable hardware, stable & <6h on 8GB.")
        else:
            emit(f"🖥️ Local: {total_urls_count:,} URLs (concurrency={concurrency}, tier={tier}).")

    # For 50k–100k only: optional cap (get_auto_run_settings already tiers 200k+)
    if total_urls_count > 50000 and total_urls_count <= 100000:
        # 50k+ URLs: reduce concurrency to prevent memory exhaustion
        original_concurrency = concurrency
        concurrency = min(concurrency, 6 if is_cloud_mode() else 10)
        if concurrency != original_concurrency:
            emit(f"🔧 Large dataset ({total_urls_count:,} URLs). Reduced concurrency to {concurrency} for stability.")
    elif total_urls_count > 20000:
        # 20k+ URLs: moderate concurrency
        original_concurrency = concurrency
        concurrency = min(concurrency, 8 if is_cloud_mode() else 15)
        if concurrency != original_concurrency:
            emit(f"🔧 Large dataset ({total_urls_count:,} URLs). Reduced concurrency to {concurrency} for stability.")

    os.makedirs(output_dir, exist_ok=True)
    checkpoint_path = get_checkpoint_path(output_dir)
    checkpoint_data = load_checkpoint(checkpoint_path)
    
    if checkpoint_data is not None:
        completed_urls = checkpoint_data.get("completed_urls", set())
        start_part = checkpoint_data.get("last_part", 0)
        stored_urls = checkpoint_data.get("urls", [])
        # Only resume if stored URLs match (same run). Compare length first to avoid building huge sets when different.
        if len(stored_urls) == len(urls) and set(stored_urls) == set(urls):
            remaining_with_idx = [(u, idx) for idx, u in enumerate(urls) if u not in completed_urls]
            if not remaining_with_idx:
                emit("✅ All URLs already completed (resume found nothing to do)")
                return
            emit(f"📂 Resuming: {len(completed_urls)} already done, {len(remaining_with_idx)} remaining")
        else:
            completed_urls = set()
            start_part = 0
            remaining_with_idx = [(u, idx) for idx, u in enumerate(urls)]
    else:
        completed_urls = set()
        start_part = 0
        remaining_with_idx = [(u, idx) for idx, u in enumerate(urls)]
    
    total_overall = len(urls)
    total_this_run = len(remaining_with_idx)
    completed_before = len(completed_urls)
    
    checkpoint_data = {
        "urls": urls,
        "completed_urls": completed_urls,
        "last_part": start_part,
        "run_folder": run_folder or os.path.basename(output_dir),
        "started_at": datetime.now().isoformat(),
    }
    save_checkpoint(checkpoint_data, checkpoint_path)
    
    progress_state = {"last_count": completed_before, "last_time": time.time()}
    
    def adapted_progress(done, total_queue):
        progress_state["last_count"] = completed_before + done
        progress_state["last_time"] = time.time()
        if progress_callback:
            progress_callback(completed_before + done, total_overall)
    
    # Backpressure: queue size by environment, run size, and tier (low RAM = smaller queue to avoid OOM)
    if is_cloud_mode():
        queue_max = min(300, total_this_run + 10) if total_urls_count < 5000 else min(100, total_this_run + 10)
    else:
        if total_urls_count >= 200000:
            # 200k–300k: tier-based queue — low=100 (8GB), medium=150, high=200 (full force)
            tier = get_system_tier()
            cap = 100 if tier == "low" else (150 if tier == "medium" else 200)
            queue_max = min(cap, total_this_run + 10)
        else:
            queue_max = min(200 if low_resource else 1000, total_this_run + 10)
    result_queue = asyncio.Queue(maxsize=queue_max)
    url_queue = asyncio.Queue()

    for u, idx in remaining_with_idx:
        await url_queue.put((u, idx))

    timeout_obj = aiohttp.ClientTimeout(total=None)
    # Use cookie jar for session persistence (makes requests look more legitimate)
    cookie_jar = aiohttp.CookieJar(unsafe=True)  # Allow cross-domain cookies
    # Connector - tune for low-resource vs fast mode; local runs get higher limits
    if low_resource:
        conn_limit, conn_per_host = 30, 3
    elif fast_mode and not is_cloud_mode():
        conn_limit, conn_per_host = 400, 20
    elif fast_mode:
        conn_limit, conn_per_host = 200, 8
    else:
        conn_limit, conn_per_host = 150 if is_cloud_mode() else 300, 5 if is_cloud_mode() else 15
    connector = aiohttp.TCPConnector(
        limit=conn_limit,
        limit_per_host=conn_per_host,
        ssl=False,
        enable_cleanup_closed=True,
        keepalive_timeout=30,
        ttl_dns_cache=300,
        use_dns_cache=True,
    )
    # Store user agent in session headers for fetch() to use when generating realistic headers
    session = aiohttp.ClientSession(
        headers={"User-Agent": user_agent}, 
        connector=connector, 
        timeout=timeout_obj, 
        cookie_jar=cookie_jar,
        trust_env=True  # Use system proxy settings if available
    )

    writer_task = asyncio.create_task(writer_coroutine(
        result_queue, rows_per_file, output_dir, total_this_run, adapted_progress,
        start_part=start_part, checkpoint_data=checkpoint_data, checkpoint_path=checkpoint_path,
        log_callback=log_callback, include_email_copy=email_copy_enabled,
        original_df=original_df, csv_config=csv_config))

    workers = [asyncio.create_task(worker_coroutine(
        f"worker-{i+1}", session, url_queue, result_queue, depth, keywords, max_chars, retries, timeout,
        ai_enabled, ai_api_key, ai_provider, ai_model, ai_prompt,
        email_copy_enabled, email_copy_api_key, email_copy_provider, email_copy_model, email_copy_prompt,
        lead_data_map, ai_status_callback, scrape_status_callback, fast_mode,
        use_playwright_fallback, use_common_crawl_fallback, total_urls=total_overall,
        original_df=original_df, url_list_with_idx=url_list_with_idx, csv_config=csv_config)) for i in range(concurrency)]

    # Timeout: cap at 3 hours, plus 10 min headroom for pause/recovery phases
    base_time = (timeout * (retries + 1) * (depth + 1) * 2 * total_this_run) + (30 * total_this_run)
    max_queue_time = min(base_time + 600, 3 * 3600 + 600)  # +10 min for pause-and-wait recovery
    tier = get_system_tier() if not is_cloud_mode() else "cloud"
    emit(f"⚙️ Run config: tier={tier}, workers={concurrency}, timeout={timeout}s, total={total_overall}, remaining={total_this_run}, queue_max={queue_max}, max_queue_time={int(max_queue_time)}s, fast_mode={fast_mode}, low_resource={low_resource}")
    
    recovery_wait_s = 300  # 5 minutes: wait for slow network/system before giving up
    recovery_chunk_s = 30  # Check progress every 30s during recovery
    stall_complete_event = asyncio.Event()  # Set when stall confirmed and we should force shutdown
    near_completion_recovery = 90  # When 99%+ done and stuck, use shorter recovery

    async def force_completion_on_stall():
        """Detect stall and PAUSE for recovery (slow network/disconnect) before draining."""
        stall_threshold = 240 if low_resource else 300  # 4-5 min before considering stall
        while True:
            await asyncio.sleep(15)
            elapsed = time.time() - progress_state["last_time"]
            last_count = progress_state["last_count"]
            remaining = total_overall - last_count
            if remaining <= 0:
                break
            # Near completion (1-5 URLs left): use shorter threshold so we don't wait hours for 1 stuck worker
            pct_done = last_count / max(total_overall, 1)
            use_shorter = remaining <= 5 and pct_done >= 0.99
            effective_threshold = min(stall_threshold, near_completion_recovery) if use_shorter else stall_threshold
            if use_shorter:
                effective_recovery = 60  # 1 min for near-completion stall
            else:
                effective_recovery = recovery_wait_s
            if elapsed > effective_threshold:
                emit(f"⏸️ Pausing: no progress for {int(elapsed)}s ({last_count}/{total_overall} done). Waiting up to {effective_recovery // 60} min for recovery...")
                waited = 0
                recovered = False
                while waited < effective_recovery:
                    await asyncio.sleep(recovery_chunk_s)
                    waited += recovery_chunk_s
                    new_count = progress_state["last_count"]
                    if new_count > last_count:
                        recovered = True
                        emit(f"✅ Recovery: progress resumed ({new_count - last_count} completed). Continuing.")
                        break
                    emit(f"⏳ Still waiting... {waited}s / {effective_recovery}s (no progress yet)")
                if not recovered:
                    emit(f"⚠️ No recovery. Stopping stuck workers (1 URL may be missing; will retry on resume).")
                    drained = 0
                    while drained < remaining + 100:
                        try:
                            item = url_queue.get_nowait()
                        except asyncio.QueueEmpty:
                            break
                        if item is None or item == (None, None):
                            url_queue.task_done()
                            continue
                        url, idx = (item[0], item[1]) if isinstance(item, tuple) else (item, None)
                        row_idx = idx if idx is not None else -1
                        if email_copy_enabled:
                            err_result = (url, "❌ Stall: worker stopped after recovery wait", "", "", row_idx)
                        else:
                            err_result = (url, "❌ Stall: worker stopped after recovery wait", "", row_idx)
                        for _ in range(20):
                            try:
                                result_queue.put_nowait(err_result)
                                break
                            except asyncio.QueueFull:
                                await asyncio.sleep(0.3)
                        else:
                            emit("⚠️ Result queue full during stall drain")
                        url_queue.task_done()
                        drained += 1
                    stall_complete_event.set()
                break
    
    stall_monitor = asyncio.create_task(force_completion_on_stall())
    
    async def heartbeat():
        while True:
            await asyncio.sleep(20)
            done = progress_state["last_count"]
            elapsed = int(time.time() - progress_state["last_time"])
            emit(f"💓 Heartbeat: {done}/{total_overall} done, idle_for={elapsed}s, url_q={url_queue.qsize()}, result_q={result_queue.qsize()}")
    heartbeat_task = asyncio.create_task(heartbeat())
    
    try:
        join_task = asyncio.create_task(url_queue.join())
        stall_wait_task = asyncio.create_task(stall_complete_event.wait())
        done, pending = await asyncio.wait(
            [join_task, stall_wait_task],
            return_when=asyncio.FIRST_COMPLETED,
            timeout=max_queue_time
        )
        stall_wait_task.cancel()
        try:
            await stall_wait_task
        except asyncio.CancelledError:
            pass
        join_completed = join_task.done() and not join_task.cancelled() and join_task.exception() is None
        stall_triggered = stall_complete_event.is_set()
        if not join_completed:
            join_task.cancel()
            try:
                await join_task
            except asyncio.CancelledError:
                pass
            if not stall_triggered:
                emit(f"⏸️ Queue join timed out. Pausing 3 min for recovery...")
                await asyncio.sleep(180)
                emit(f"⚠️ Draining remaining URLs...")
            drained = 0
            while drained < total_this_run + 50:
                try:
                    item = url_queue.get_nowait()
                except asyncio.QueueEmpty:
                    break
                if item is None or item == (None, None):
                    url_queue.task_done()
                    continue
                url = item[0] if isinstance(item, tuple) else item
                idx = item[1] if isinstance(item, tuple) and len(item) > 1 else None
                row_idx = idx if idx is not None else -1
                err_msg = "❌ Stall: worker stopped" if stall_triggered else f"❌ Timeout: exceeded {max_queue_time}s"
                if email_copy_enabled:
                    err_result = (url, err_msg, "", "", row_idx)
                else:
                    err_result = (url, err_msg, "", row_idx)
                for _ in range(30):
                    try:
                        result_queue.put_nowait(err_result)
                        break
                    except asyncio.QueueFull:
                        await asyncio.sleep(0.3)
                else:
                    emit("⚠️ Result queue full during drain")
                url_queue.task_done()
                drained += 1

        stall_monitor.cancel()
        heartbeat_task.cancel()
        try:
            await stall_monitor
        except asyncio.CancelledError:
            pass
        except Exception as e:
            emit(f"⚠️ Stall monitor error: {e}")
        try:
            await heartbeat_task
        except asyncio.CancelledError:
            pass
        except Exception as e:
            emit(f"⚠️ Heartbeat monitor error: {e}")

        for _ in workers:
            try:
                await url_queue.put(None)
            except Exception:
                pass
        
        try:
            await asyncio.wait_for(asyncio.gather(*workers, return_exceptions=True), timeout=30)
        except asyncio.TimeoutError:
            for worker in workers:
                if not worker.done():
                    worker.cancel()

        try:
            await result_queue.put(None)
        except Exception:
            pass
        
        try:
            await asyncio.wait_for(result_queue.join(), timeout=60)
        except asyncio.TimeoutError:
            emit("⚠️ Result queue join timed out. Proceeding anyway...")
        
        try:
            await asyncio.wait_for(writer_task, timeout=120)
        except asyncio.TimeoutError:
            emit("⚠️ Writer task timed out. Proceeding anyway...")
            writer_task.cancel()
        except Exception as e:
            emit(f"⚠️ Error waiting for writer: {e}")

        await session.close()
        await asyncio.sleep(2)
    except BaseException as e:
        import traceback
        tb = traceback.format_exc()
        emit(f"⚠️ Scraper error (progress saved): {type(e).__name__}: {e}")
        emit("--- Full traceback ---")
        for line in tb.strip().split("\n"):
            emit(line)
        emit("--- End traceback ---")
    finally:
        # Always clean up running tasks/session to avoid hidden hangs between runs
        for task in [stall_monitor, heartbeat_task, writer_task, *workers]:
            try:
                if task and not task.done():
                    task.cancel()
            except BaseException:
                pass
        try:
            if not session.closed:
                await session.close()
        except BaseException:
            pass
        try:
            save_checkpoint(checkpoint_data, checkpoint_path)
        except BaseException:
            pass
        try:
            emit("💾 Checkpoint saved (progress safe).")
        except Exception:
            pass

# -------------------------
# Streamlit UI
# -------------------------

st.set_page_config(
    page_title="Web Scraper", 
    layout="wide", 
    page_icon="🌐",
    initial_sidebar_state="collapsed"
)

# Initialize session state defaults
SESSION_DEFAULTS = {
    'keywords': [],
    'keywords_input': 'about,service,product',
    'concurrency': 20,
    'force_max_workers': False,
    'depth': 3,
    'retries': 3,
    'timeout': 35,
    'max_chars': 10000,
    'rows_per_file': 2000,
    'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'ai_enabled': False,
    'ai_enabled_checkbox': False,
    'ai_provider': None,
    'ai_model': None,
    'ai_api_key': None,
    'ai_prompt': None,
    'master_prompt': None,
    'email_copy_enabled': False,
    'email_copy_enabled_checkbox': False,
    'email_copy_provider': None,
    'email_copy_model': None,
    'email_copy_api_key': None,
    'email_copy_prompt': None,
    'email_copy_use_step3': False,
    'csv_config': None,
    '_csv_url_count': 0,
    'step3_preview_row': 0,
    'step4_preview_row': 0,
    'current_step': 1,
    '_test_running': False,
    '_test_results': None,
    '_test_error': None,
}

for key, value in SESSION_DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = value

# Dark theme UI styling (design alignment)
st.markdown("""
<style>
    /* 1. Global theme and layout */
    .stApp {
        background: radial-gradient(ellipse 80% 50% at 50% -20%, rgba(6, 182, 212, 0.15), transparent),
                    radial-gradient(ellipse 60% 40% at 80% 50%, rgba(99, 102, 241, 0.08), transparent),
                    linear-gradient(180deg, #050508 0%, #0c0f1a 30%, #0f172a 70%, #050508 100%);
    }
    .main .block-container {
        padding-top: 2.5rem;
        padding-bottom: 3rem;
        max-width: 1100px;
        margin: 0 auto;
        background: rgba(15, 23, 42, 0.5);
        border-radius: 1.25rem;
        border: 1px solid rgba(51, 65, 85, 0.4);
        backdrop-filter: blur(20px);
        box-shadow: 0 4px 24px rgba(0,0,0,0.25);
    }
    .main, .block-container {
        color: #e2e8f0;
    }
    p, span, label, .stMarkdown {
        color: #cbd5e1 !important;
    }
    ::selection {
        background: rgba(6, 182, 212, 0.3);
        color: #e0f2fe;
    }
    
    /* 2. Hero - handled in HTML below */
    
    /* 3-4. Section headers and typography */
    h1 {
        color: #f1f5f9 !important;
        font-weight: 700;
        margin-bottom: 0.5rem;
        font-size: 2.5rem;
    }
    h3 {
        color: #f1f5f9 !important;
        font-weight: 600;
        margin-top: 0;
        margin-bottom: 1rem;
        font-size: 1.5rem;
    }
    h4 {
        color: #e2e8f0 !important;
        font-weight: 600;
        margin-top: 1.5rem;
        margin-bottom: 0.75rem;
    }
    
    /* Step card class - refined with step number badge */
    .step-card {
        background: linear-gradient(135deg, rgba(15, 23, 42, 0.9) 0%, rgba(30, 41, 59, 0.6) 100%);
        backdrop-filter: blur(16px);
        border: 1px solid rgba(51, 65, 85, 0.5);
        border-radius: 1rem;
        padding: 1.5rem 1.75rem;
        margin-bottom: 1.75rem;
        box-shadow: 0 4px 20px rgba(0,0,0,0.2);
        transition: border-color 0.25s ease, box-shadow 0.25s ease;
    }
    .step-card:hover {
        border-color: rgba(51, 65, 85, 0.8);
        box-shadow: 0 8px 28px rgba(0,0,0,0.3);
    }
    .step-card.active {
        box-shadow: 0 0 30px -5px rgba(6, 182, 212, 0.25);
        border-color: rgba(6, 182, 212, 0.4);
    }
    .step-card .step-badge {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 28px;
        height: 28px;
        background: linear-gradient(135deg, #06b6d4, #0891b2);
        border-radius: 8px;
        font-size: 0.8rem;
        font-weight: 700;
        color: white;
        margin-right: 0.5rem;
        vertical-align: middle;
    }
    
    /* Tip boxes inside step cards - dark with left accent */
    .tip-box {
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin-bottom: 1rem;
        color: #e2e8f0;
        font-size: 0.95rem;
    }
    .tip-box-cyan {
        background: rgba(6, 182, 212, 0.1);
        border-left: 4px solid #06b6d4;
    }
    .tip-box-emerald {
        background: rgba(16, 185, 129, 0.1);
        border-left: 4px solid #10b981;
    }
    .tip-box-fuchsia {
        background: rgba(217, 70, 239, 0.1);
        border-left: 4px solid #d946ef;
    }
    
    /* Info / tip boxes - dark with left accent */
    .stAlert, [data-testid="stAlert"] {
        background: rgba(6, 182, 212, 0.08) !important;
        border-left: 4px solid #06b6d4 !important;
        border-radius: 0 8px 8px 0 !important;
        color: #e0f2fe !important;
    }
    .stSuccess, [data-testid="stSuccess"] {
        background: rgba(16, 185, 129, 0.12) !important;
        border-left: 4px solid #10b981 !important;
        border-radius: 0 8px 8px 0 !important;
        color: #a7f3d0 !important;
    }
    .stWarning, [data-testid="stWarning"] {
        background: rgba(245, 158, 11, 0.12) !important;
        border-left: 4px solid #f59e0b !important;
        border-radius: 0 8px 8px 0 !important;
        color: #fde68a !important;
    }
    .stError, [data-testid="stError"] {
        background: rgba(239, 68, 68, 0.12) !important;
        border-left: 4px solid #ef4444 !important;
        border-radius: 0 8px 8px 0 !important;
        color: #fecaca !important;
    }
    
    /* Primary CTA: Start Scraping button */
    .main [data-testid="stVerticalBlock"]:has(.cta-heading-block) .stButton > button {
        background: linear-gradient(135deg, #06b6d4 0%, #0891b2 40%, #4f46e5 100%) !important;
        color: white !important;
        font-weight: 700 !important;
        font-size: 1.05rem !important;
        padding: 1rem 2.5rem !important;
        border-radius: 12px !important;
        border: none !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 20px rgba(6, 182, 212, 0.35) !important;
    }
    .main [data-testid="stVerticalBlock"]:has(.cta-heading-block) .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 28px rgba(6, 182, 212, 0.45) !important;
    }
    
    /* Secondary buttons (all other buttons) */
    .stButton > button {
        background: rgba(30, 41, 59, 0.8) !important;
        color: #cbd5e1 !important;
        border: 1px solid #475569 !important;
        font-weight: 500 !important;
        padding: 0.5rem 1rem !important;
        border-radius: 8px !important;
        transition: all 0.3s ease !important;
    }
    .stButton > button:hover {
        border-color: #22d3ee !important;
        color: #22d3ee !important;
        background: rgba(6, 182, 212, 0.1) !important;
    }
    
    /* Inputs */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input {
        background: rgba(2, 6, 23, 0.6) !important;
        border: 1px solid #475569 !important;
        border-radius: 8px !important;
        color: #e2e8f0 !important;
        transition: border-color 0.2s, box-shadow 0.2s !important;
    }
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        border-color: #06b6d4 !important;
        box-shadow: 0 0 0 2px rgba(6, 182, 212, 0.3) !important;
    }
    .stTextInput > div > div > input::placeholder {
        color: #64748b !important;
    }
    
    /* Select boxes */
    .stSelectbox > div > div {
        background: rgba(2, 6, 23, 0.6) !important;
        border: 1px solid #475569 !important;
        border-radius: 8px !important;
        color: #e2e8f0 !important;
    }
    .stSelectbox > div > div:focus-within {
        border-color: #06b6d4 !important;
        box-shadow: 0 0 0 2px rgba(6, 182, 212, 0.3) !important;
    }
    
    /* Text area */
    .stTextArea > div > div > textarea {
        background: rgba(2, 6, 23, 0.6) !important;
        border: 1px solid #475569 !important;
        border-radius: 8px !important;
        color: #e2e8f0 !important;
    }
    .stTextArea > div > div > textarea:focus {
        border-color: #06b6d4 !important;
        box-shadow: 0 0 0 2px rgba(6, 182, 212, 0.3) !important;
    }
    
    /* Sliders - cyan accent */
    .stSlider [data-baseweb="slider"] span[role="slider"] {
        background: #06b6d4 !important;
    }
    .stSlider [data-baseweb="slider"] > div > div {
        background: #334155 !important;
        border-radius: 8px !important;
    }
    
    /* File uploader - inviting drop zone with consistent dark theme */
    [data-testid="stFileUploader"] {
        width: 100% !important;
    }
    [data-testid="stFileUploader"] > section,
    [data-testid="stFileUploader"] > div > div {
        background: rgba(15, 23, 42, 0.5) !important;
        border: 2px dashed rgba(71, 85, 105, 0.6) !important;
        border-radius: 1rem !important;
        padding: 2rem !important;
        transition: all 0.3s ease !important;
    }
    [data-testid="stFileUploader"] > section:hover,
    [data-testid="stFileUploader"] > div > div:hover {
        border-color: rgba(6, 182, 212, 0.5) !important;
        background: rgba(6, 182, 212, 0.08) !important;
        box-shadow: 0 0 20px rgba(6, 182, 212, 0.1);
    }
    [data-testid="stFileUploader"] [data-testid="stMarkdownContainer"] p {
        color: #94a3b8 !important;
        font-size: 0.95rem !important;
    }
    [data-testid="stFileUploader"] button {
        background: rgba(6, 182, 212, 0.15) !important;
        border: 1px solid rgba(6, 182, 212, 0.4) !important;
        color: #22d3ee !important;
        border-radius: 8px !important;
        padding: 0.5rem 1.25rem !important;
        font-weight: 500 !important;
        transition: all 0.2s ease !important;
    }
    [data-testid="stFileUploader"] button:hover {
        background: rgba(6, 182, 212, 0.25) !important;
        border-color: rgba(6, 182, 212, 0.6) !important;
        transform: translateY(-1px);
    }
    
    /* Expanders - refined dark styling */
    .streamlit-expanderHeader {
        font-weight: 600;
        color: #e2e8f0 !important;
        background: linear-gradient(135deg, rgba(30, 41, 59, 0.9) 0%, rgba(51, 65, 85, 0.7) 100%) !important;
        border: 1px solid rgba(71, 85, 105, 0.6) !important;
        border-radius: 12px !important;
        padding: 0.75rem 1rem !important;
        transition: all 0.2s ease !important;
    }
    .streamlit-expanderHeader:hover {
        background: linear-gradient(135deg, rgba(51, 65, 85, 0.9) 0%, rgba(71, 85, 105, 0.7) 100%) !important;
        border-color: rgba(6, 182, 212, 0.4) !important;
    }
    [data-testid="stExpander"] {
        border: 1px solid rgba(71, 85, 105, 0.5) !important;
        border-radius: 12px !important;
        margin-bottom: 0.75rem !important;
        background: rgba(15, 23, 42, 0.6) !important;
        overflow: hidden;
    }
    [data-testid="stExpander"]:hover {
        border-color: rgba(6, 182, 212, 0.3) !important;
    }
    .streamlit-expanderContent {
        border-top: 1px solid rgba(71, 85, 105, 0.4) !important;
        padding: 1.25rem !important;
        background: rgba(15, 23, 42, 0.4) !important;
        color: #e2e8f0 !important;
    }
    .streamlit-expanderContent .stMarkdown,
    .streamlit-expanderContent p,
    .streamlit-expanderContent label,
    .streamlit-expanderContent .stCaption {
        color: #cbd5e1 !important;
    }
    /* Expander icon color */
    .streamlit-expanderHeader svg {
        color: #22d3ee !important;
        fill: #22d3ee !important;
    }
    
    /* Number input wrapper - grey, no white */
    .stNumberInput > div,
    [data-testid="stNumberInput"] > div {
        background: rgba(30, 41, 59, 0.6) !important;
        border-radius: 8px !important;
    }
    
    /* Captions */
    .stCaption {
        color: #94a3b8 !important;
        font-size: 0.9rem;
    }
    
    /* Dividers - gradient line */
    hr {
        margin: 2rem 0;
        border: none;
        height: 1px;
        background: linear-gradient(90deg, transparent, rgba(6, 182, 212, 0.25), rgba(99, 102, 241, 0.2), transparent);
    }
    
    /* Subheaders in sections */
    .main h4, .main .stSubheader {
        color: #e2e8f0 !important;
        font-weight: 600;
    }
    
    /* Radio / checkbox labels */
    .stRadio > div > label, .stCheckbox > label {
        color: #cbd5e1 !important;
    }
    .stRadio > div {
        gap: 1rem;
    }
    
    /* Progress bar - gradient fill */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #06b6d4, #3b82f6, #4f46e5) !important;
        border-radius: 9999px !important;
    }
    [data-testid="stProgressBar"] > div > div {
        background: #1e293b !important;
        border-radius: 9999px !important;
    }
    
    /* Metrics - dark cards */
    [data-testid="stMetric"] {
        background: rgba(15, 23, 42, 0.8) !important;
        border: 1px solid rgba(51, 65, 85, 0.5) !important;
        border-radius: 12px !important;
        padding: 1rem !important;
        box-shadow: inset 0 0 15px rgba(0,0,0,0.3);
    }
    [data-testid="stMetric"] label {
        color: #94a3b8 !important;
    }
    [data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #22d3ee !important;
    }
    
    /* Code / terminal log */
    .stCodeBlock, [data-testid="stCodeBlock"] {
        background: #0a0a0f !important;
        border: 1px solid #334155 !important;
        border-radius: 8px !important;
    }
    .stCodeBlock code, [data-testid="stCodeBlock"] code {
        color: #6ee7b7 !important;
        font-family: ui-monospace, monospace !important;
    }
    
    /* Download section - card-style layout */
    .main [data-testid="stVerticalBlock"]:has(.download-section-anchor) {
        background: linear-gradient(135deg, rgba(15, 23, 42, 0.8) 0%, rgba(30, 41, 59, 0.5) 100%) !important;
        border: 1px solid rgba(51, 65, 85, 0.5) !important;
        border-radius: 1rem !important;
        padding: 1.5rem !important;
        margin: 1rem 0 !important;
    }
    .main [data-testid="stVerticalBlock"]:has(.download-section-anchor) .stDownloadButton > button {
        background: rgba(30, 41, 59, 0.9) !important;
        border: 1px solid #475569 !important;
        color: #e2e8f0 !important;
        transition: all 0.2s ease !important;
    }
    .main [data-testid="stVerticalBlock"]:has(.download-section-anchor) .stDownloadButton > button:hover {
        border-color: #06b6d4 !important;
        color: #22d3ee !important;
        background: rgba(6, 182, 212, 0.1) !important;
    }
    .main [data-testid="stVerticalBlock"]:has(.download-section-anchor) .stDownloadButton:nth-of-type(1) > button {
        border-color: rgba(6, 182, 212, 0.5) !important;
        box-shadow: 0 0 20px rgba(6, 182, 212, 0.2) !important;
    }
    .main [data-testid="stVerticalBlock"]:has(.download-section-anchor) .stDownloadButton:nth-of-type(2) > button,
    .main [data-testid="stVerticalBlock"]:has(.download-section-anchor) .stColumns > div:nth-child(2) .stDownloadButton > button {
        border-color: rgba(6, 182, 212, 0.5) !important;
        box-shadow: 0 0 20px rgba(6, 182, 212, 0.2) !important;
    }
    
    /* Custom scrollbar */
    .main ::-webkit-scrollbar { width: 8px; height: 8px; }
    .main ::-webkit-scrollbar-track { background: rgba(15, 23, 42, 0.5); border-radius: 4px; }
    .main ::-webkit-scrollbar-thumb { background: rgba(51, 65, 85, 0.8); border-radius: 4px; }
    .main ::-webkit-scrollbar-thumb:hover { background: rgba(6, 182, 212, 0.5); }
    
    /* Resume / partial results */
    .resume-runs .stMarkdown { color: #e2e8f0 !important; }
    .resume-runs .stCaption { color: #94a3b8 !important; }
    
    /* Dialog / modal - dark overlay and panel */
    [data-testid="stDialog"] {
        background: rgba(0, 0, 0, 0.6) !important;
        backdrop-filter: blur(8px);
    }
    [data-testid="stDialog"] > div {
        background: #0f172a !important;
        border: 1px solid #334155 !important;
        border-radius: 1rem !important;
        box-shadow: 0 0 40px rgba(0,0,0,0.5) !important;
    }
    [data-testid="stDialog"] .stMarkdown { color: #e2e8f0 !important; }
    
    /* Token estimator - terminal-style panel */
    .main [data-testid="stExpander"]:has([class*="token-estimator-anchor"]) {
        border: 1px solid rgba(34, 197, 94, 0.3) !important;
        background: rgba(0, 0, 0, 0.4) !important;
    }
    .main [data-testid="stExpander"]:has([class*="token-estimator-anchor"]) .streamlit-expanderHeader {
        color: #6ee7b7 !important;
    }
    
    /* Test run area - grey card, readable text */
    .test-run-anchor ~ * { color: inherit; }
    .main [data-testid="stVerticalBlock"]:has(.test-run-anchor) {
        background: #334155 !important;
        border: 1px solid #475569 !important;
        border-radius: 1rem !important;
        padding: 1rem !important;
    }
    .main [data-testid="stVerticalBlock"]:has(.test-run-anchor):hover {
        background: #374151 !important;
        border-color: #64748b !important;
    }
    .main [data-testid="stVerticalBlock"]:has(.test-run-anchor) .stCaption,
    .main [data-testid="stVerticalBlock"]:has(.test-run-anchor) label {
        color: #cbd5e1 !important;
    }

    /* Step indicator styles */
    .step-indicator {
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 0.5rem;
        margin: 1.5rem 0 2rem;
        padding: 0.75rem;
        background: rgba(15, 23, 42, 0.6);
        border-radius: 1rem;
        border: 1px solid rgba(51, 65, 85, 0.5);
    }
    .step-item {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        padding: 0.5rem 1rem;
        border-radius: 0.75rem;
        transition: all 0.3s ease;
        cursor: default;
    }
    .step-item.completed {
        background: rgba(16, 185, 129, 0.15);
        border: 1px solid rgba(16, 185, 129, 0.3);
    }
    .step-item.active {
        background: rgba(6, 182, 212, 0.15);
        border: 1px solid rgba(6, 182, 212, 0.4);
        box-shadow: 0 0 20px rgba(6, 182, 212, 0.15);
    }
    .step-item.pending {
        background: rgba(51, 65, 85, 0.3);
        border: 1px solid rgba(71, 85, 105, 0.3);
        opacity: 0.7;
    }
    .step-number {
        width: 24px;
        height: 24px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 700;
        font-size: 0.8rem;
    }
    .step-item.completed .step-number {
        background: #10b981;
        color: white;
    }
    .step-item.active .step-number {
        background: #06b6d4;
        color: white;
    }
    .step-item.pending .step-number {
        background: #475569;
        color: #94a3b8;
    }
    .step-label {
        font-size: 0.85rem;
        font-weight: 500;
        white-space: nowrap;
    }
    .step-item.completed .step-label {
        color: #10b981 !important;
    }
    .step-item.active .step-label {
        color: #22d3ee !important;
    }
    .step-item.pending .step-label {
        color: #64748b !important;
    }
    .step-connector {
        width: 24px;
        height: 2px;
        background: rgba(71, 85, 105, 0.5);
        border-radius: 1px;
    }
    .step-connector.completed {
        background: linear-gradient(90deg, #10b981, #06b6d4);
    }

    /* Required field indicator */
    .required-field::after {
        content: " *";
        color: #ef4444;
    }

    /* Validation message styles */
    .validation-message {
        background: rgba(239, 68, 68, 0.1);
        border-left: 3px solid #ef4444;
        padding: 0.75rem 1rem;
        border-radius: 0 8px 8px 0;
        margin: 0.5rem 0;
        color: #fecaca;
        font-size: 0.9rem;
    }

    /* Remove default Streamlit branding */
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# Hero: refined icon badge + gradient title + subtitle
st.markdown("""
<div class="hero-section" style="text-align: center; padding: 2.5rem 0 2rem; margin-bottom: 1.5rem;">
    <div style="display: inline-flex; align-items: center; justify-content: center; padding: 1rem 1.25rem; 
         background: linear-gradient(135deg, rgba(6, 182, 212, 0.15) 0%, rgba(99, 102, 241, 0.1) 100%); 
         border-radius: 1.25rem; border: 1px solid rgba(6, 182, 212, 0.3); margin-bottom: 1.25rem;
         box-shadow: 0 0 40px rgba(6, 182, 212, 0.1);">
        <span style="font-size: 2.5rem;">🌐</span>
    </div>
    <h1 style="background: linear-gradient(90deg, #22d3ee 0%, #38bdf8 25%, #3b82f6 50%, #6366f1 75%, #8b5cf6 100%);
               -webkit-background-clip: text;
               -webkit-text-fill-color: transparent;
               background-clip: text;
               margin-bottom: 0.5rem;
               font-size: 2.85rem;
               font-weight: 800;
               letter-spacing: -0.03em;
               line-height: 1.2;">
        Website Scraper
    </h1>
    <p style="color: #94a3b8; font-size: 1.05rem; margin-top: 0.5rem; max-width: 36rem; margin-left: auto; margin-right: auto; line-height: 1.6;">
        Scrape websites from your CSV file and get clean, structured text content — with optional AI summaries and email copy
    </p>
</div>
""", unsafe_allow_html=True)

# Desktop app download
RELEASES_URL = "https://github.com/ahmadzabir/WebScrapper/releases"
RUN_BUILD_URL = "https://github.com/ahmadzabir/WebScrapper/actions/workflows/desktop-build.yml"

dl_col1, dl_col2 = st.columns(2)
with dl_col1:
    st.link_button("⬇️ Download for Windows", RELEASES_URL, type="primary", use_container_width=True)
with dl_col2:
    st.link_button("🔧 Create first release", RUN_BUILD_URL, type="secondary", use_container_width=True)
st.caption("**Download:** Opens releases → click WebScrapperDesktop-Setup.exe. **Create first release:** Run the build workflow (takes ~15 min), then the download will work.")

# Step indicator - shows progress through the workflow
def render_step_indicator(current_step):
    steps = [
        (1, "📁 Upload"),
        (2, "⚙️ Settings"),
        (3, "🤖 AI Setup"),
        (4, "🚀 Run")
    ]
    html = '<div class="step-indicator">'
    for i, (num, label) in enumerate(steps):
        if num < current_step:
            status = "completed"
            icon = "✓"
        elif num == current_step:
            status = "active"
            icon = str(num)
        else:
            status = "pending"
            icon = str(num)
        
        html += f'<div class="step-item {status}"><span class="step-number">{icon}</span><span class="step-label">{label}</span></div>'
        if i < len(steps) - 1:
            connector_status = "completed" if num < current_step else ""
            html += f'<div class="step-connector {connector_status}"></div>'
    html += '</div>'
    return html

# Determine current step based on session state
current_step = 1
if st.session_state.get('csv_config'):
    current_step = 2
if st.session_state.get('current_step', 1) >= 3:
    current_step = 3
if st.session_state.get('_test_running') or st.session_state.get('_last_run_complete'):
    current_step = 4

st.markdown(render_step_indicator(current_step), unsafe_allow_html=True)

# -------- RESUME / PARTIAL RESULTS --------
# Scan outputs/ for checkpoints - show runs that can be resumed or have partial data
outputs_dir = "outputs"
if os.path.isdir(outputs_dir):
    incomplete_runs = []
    for run_folder in sorted(os.listdir(outputs_dir), reverse=True)[:10]:
        run_path = os.path.join(outputs_dir, run_folder)
        if not os.path.isdir(run_path):
            continue
        ck_path = get_checkpoint_path(run_path)
        ck = load_checkpoint(ck_path)
        if not ck:
            continue
        total_urls = len(ck.get("urls", []))
        # Use actual row count from files (source of truth) — checkpoint can be wrong after crash
        actual_rows, _ = get_actual_completed_from_files(run_path)
        completed = actual_rows if actual_rows > 0 else len(ck.get("completed_urls", []))
        if completed > 0 and completed < total_urls:
            incomplete_runs.append({"folder": run_folder, "completed": completed, "total": total_urls, "path": run_path})
        elif completed >= total_urls and total_urls > 0:
            incomplete_runs.append({"folder": run_folder, "completed": completed, "total": total_urls, "path": run_path, "done": True})

    if incomplete_runs:
        with st.expander("📂 Resume or Download Partial Results", expanded=True):
            st.markdown("""
            <div style="background: rgba(6, 182, 212, 0.08); border-left: 4px solid #06b6d4; padding: 0.75rem 1rem; border-radius: 0 8px 8px 0; margin-bottom: 1rem; color: #e2e8f0;">
                <strong>If the app crashed or stopped</strong>, you can resume or download what was saved.
            </div>
            """, unsafe_allow_html=True)
            st.markdown('<div class="resume-runs">', unsafe_allow_html=True)
            for run in incomplete_runs:
                is_done = run.get("done", False)
                label = f"✅ {run['folder']}: {run['completed']:,}/{run['total']:,} rows" if is_done else f"⏸️ {run['folder']}: {run['completed']:,}/{run['total']:,} rows in files (partial)"
                with st.container():
                    col_a, col_b = st.columns([2, 1])
                    with col_a:
                        st.markdown(f"**{label}**")
                        if not is_done:
                            st.caption("To resume: Re-upload your CSV and click Start — the app will auto-detect and continue.")
                        st.caption("ZIP includes combined_all_results.csv with all rows in one file.")
                    with col_b:
                        zip_path = os.path.join(run["path"], f"{run['folder']}.zip")
                        if not os.path.exists(zip_path):
                            csv_files = sorted([f for f in os.listdir(run["path"]) if f.startswith("output_part_") and f.endswith(".csv") and "combined" not in f.lower()])
                            if csv_files:
                                try:
                                    import zipfile
                                    # Create combined CSV with ALL rows in one file (so user gets 900 rows, not 100 per file)
                                    combined_path = os.path.join(run["path"], "combined_all_results.csv")
                                    dfs = []
                                    for f in csv_files:
                                        fp = os.path.join(run["path"], f)
                                        try:
                                            df = pd.read_csv(fp, encoding="utf-8-sig")
                                            if len(df) > 0:
                                                dfs.append(df)
                                        except Exception:
                                            pass
                                    if dfs:
                                        combined_df = pd.concat(dfs, ignore_index=True)
                                        combined_df.to_csv(combined_path, index=False, encoding="utf-8-sig")
                                    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                                        for f in csv_files:
                                            zf.write(os.path.join(run["path"], f), arcname=f)
                                        if os.path.exists(combined_path):
                                            zf.write(combined_path, arcname="combined_all_results.csv")
                                        for f in os.listdir(run["path"]):
                                            if (f.endswith(".xlsx") or (f.endswith(".txt") and f.startswith("crash_log_"))) and "combined" not in f.lower():
                                                zf.write(os.path.join(run["path"], f), arcname=f)
                                except Exception:
                                    pass
                        
                        # Store in session state for cloud persistence
                        if os.path.exists(zip_path):
                            try:
                                with open(zip_path, "rb") as f:
                                    zip_data = f.read()
                                    # Store in session state with unique key
                                    session_key = f"resume_zip_{run['folder']}"
                                    st.session_state[session_key] = zip_data
                            except Exception:
                                zip_data = None
                        
                        # Try session state first, then disk
                        session_key = f"resume_zip_{run['folder']}"
                        zip_data = st.session_state.get(session_key)
                        
                        if not zip_data and os.path.exists(zip_path):
                            try:
                                with open(zip_path, "rb") as f:
                                    zip_data = f.read()
                                    st.session_state[session_key] = zip_data
                            except Exception:
                                zip_data = None
                        
                        if zip_data:
                            st.download_button(
                                "⬇️ Download Results",
                                zip_data,
                                file_name=f"{run['folder']}.zip",
                                mime="application/zip",
                                key=f"resume_dl_{run['folder']}"
                            )
                        else:
                            st.caption("Results expired - please re-upload CSV and run again")
            st.markdown('</div>', unsafe_allow_html=True)

# Step 1: Upload CSV
st.markdown("""
<div class="step-card">
    <h3 style="color: #f1f5f9 !important; margin-top: 0; display: flex; align-items: center;">
        <span class="step-badge">1</span>📁 Upload CSV
    </h3>
    <div class="tip-box tip-box-cyan"><strong>💡 Tip:</strong> Your CSV should have a column with website URLs (one per row)</div>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Choose CSV file",
    type=["csv"],
    help="Upload a CSV file with website URLs. If you had a file before, the session may have reset — re-upload here."
)

# If we had config or a completed run but the file is now missing, session was reset (rerun/refresh/long run).
_had_session = bool(
    st.session_state.get("csv_config")
    or st.session_state.get("scraping_complete")
    or st.session_state.get("output_dir")
)
if uploaded_file is None and _had_session:
    st.info(
        "**Session was reset** (e.g. after a long run, reconnection, or page refresh). "
        "Re-upload your CSV above to run again. You’ll need to select the URL column in Step 1. "
        "If you had results, check **Resume or Download Partial Results** above or the download section below."
    )

# CSV Configuration (shown after file upload)
csv_has_headers = None
url_column = None
lead_data_columns = {}
df_preview = None

if uploaded_file is not None:
    # CSV Configuration - Cleaner layout
    st.markdown("---")
    st.markdown("### 📋 CSV Configuration")

    # Preview first few rows
    try:
        # Reset file pointer to beginning
        uploaded_file.seek(0)
        
        # Read a sample to detect structure
        sample_df = pd.read_csv(uploaded_file, nrows=5)
        uploaded_file.seek(0)  # Reset again
        
        col_csv1, col_csv2 = st.columns(2)
        
        with col_csv1:
            csv_has_headers = st.radio(
                "First row has column names?",
                ["Yes", "No"],
                help="Does the first row contain column names like 'URL', 'Website', etc?",
                key="csv_headers_radio"
            )
        
        with col_csv2:
            st.caption("💡 **Tip:** If yes, you'll see column names to choose from")
        
        # Read CSV based on header selection
        uploaded_file.seek(0)  # Reset file pointer
        if csv_has_headers == "Yes":
            df_preview = pd.read_csv(uploaded_file, nrows=10)
            st.success(f"✅ Found {len(df_preview.columns)} columns: {', '.join(df_preview.columns.tolist()[:5])}{'...' if len(df_preview.columns) > 5 else ''}")
        else:
            df_preview = pd.read_csv(uploaded_file, header=None, nrows=10)
            st.info(f"ℹ️ Found {len(df_preview.columns)} columns (no headers)")
        
        # Column selection
        st.markdown("#### Select Columns")
        
        col_sel1, col_sel2 = st.columns(2)
        
        with col_sel1:
            # URL column selection
            if csv_has_headers == "Yes":
                url_column = st.selectbox(
                    "Which column has URLs?",
                    options=list(df_preview.columns),
                    index=0,
                    help="Output Column A (Website) will be this column. Pick the one with website addresses (e.g. https://...), not lead names."
                )
            else:
                url_column = st.selectbox(
                    "Which column has URLs?",
                    options=[f"Column {i+1}" for i in range(len(df_preview.columns))],
                    index=0,
                    help="Output Column A (Website) will be this column. Pick the one with website addresses (e.g. https://...), not lead names."
                )
        
        with col_sel2:
            st.caption("⚠️ **Required** - Must select this")
        
        # All non-URL columns become variables for AI prompts (Step 3 & 4)
        st.markdown("---")
        col_options = list(df_preview.columns) if csv_has_headers == "Yes" else [f"Column {i+1}" for i in range(len(df_preview.columns))]
        lead_data_columns = {}
        for col in col_options:
            if col == url_column:
                continue
            key = _col_to_placeholder(col)
            if key:
                lead_data_columns[key] = col
        with st.expander("📊 Variables from your sheet", expanded=True):
            st.caption("Every column (except the URL column) is available as a variable in Steps 3 & 4. E.g. column \"Website Name\" → {website_name} in your prompt.")
            if lead_data_columns:
                st.code(" • ".join(f"{{{k}}}" for k in sorted(lead_data_columns.keys())[:20]) + (" …" if len(lead_data_columns) > 20 else ""), language=None)
        
        # Store in session state for use during scraping and token estimator
        num_cols_preview = len(df_preview.columns) if df_preview is not None else 0
        if num_cols_preview == 0:
            st.error("❌ No columns detected in CSV. Check the file has headers or at least one column.")
        else:
            try:
                if csv_has_headers == "Yes":
                    if url_column not in (list(df_preview.columns) or []):
                        st.warning("⚠️ Selected URL column not found in preview. Using first column.")
                        url_col_idx = 0
                    else:
                        url_col_idx = list(df_preview.columns).index(url_column)
                else:
                    url_col_idx = int(str(url_column).replace("Column ", "").strip()) - 1
                    if url_col_idx < 0 or url_col_idx >= num_cols_preview:
                        url_col_idx = 0
                url_col_idx = min(url_col_idx, num_cols_preview - 1)
            except (ValueError, TypeError, IndexError):
                url_col_idx = 0
            url_series_preview = df_preview.iloc[:, url_col_idx].fillna("").astype(str)
        url_count = sum(1 for u in url_series_preview if str(u).strip())
        # Warn if selected URL column doesn't look like URLs (e.g. user picked "Lead Name" by mistake)
        def _looks_like_url(s):
            s = (s or "").strip()
            if not s:
                return False
            s_lower = s.lower()
            if s_lower.startswith("http://") or s_lower.startswith("https://"):
                return True
            if "." in s and any(t in s_lower for t in [".com", ".org", ".edu", ".gov", ".net", ".io", ".co"]):
                return True
            return False
        sample_vals = [u for u in url_series_preview if str(u).strip()][:20]
        url_like_count = sum(1 for u in sample_vals if _looks_like_url(u))
        if sample_vals and url_like_count < max(1, len(sample_vals) * 0.3):
            st.warning(
                "⚠️ **Selected column doesn't look like website URLs** (e.g. only names or text). "
                "Output **Column A** will be exactly this column. If you want URLs in Column A, choose the column that contains website addresses in Step 1."
            )
        st.session_state['csv_config'] = {
            'has_headers': csv_has_headers == "Yes",
            'url_column': url_column,
            'lead_data_columns': lead_data_columns,
            'df_preview': df_preview,
            'url_count': url_count
        }
        st.session_state['_csv_url_count'] = url_count
        
    except Exception as e:
        st.error(f"❌ Error reading CSV: {str(e)}")
        st.info("Please check your CSV file format and try again.")

# Step 2: Settings
st.markdown("""
<div class="step-card">
    <h3 style="color: #f1f5f9 !important; margin-top: 0; display: flex; align-items: center;">
        <span class="step-badge">2</span>⚙️ Settings
    </h3>
    <div class="tip-box tip-box-emerald"><strong>⚡ Speed is automatic.</strong> Workers, timeouts, and retries are set from your list size—nothing to tune.</div>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns(2)

with col1:
    st.markdown("#### What to scrape")
    keywords_input = st.text_input(
        "Keywords to find (optional)",
        value=st.session_state.get('keywords_input', "about,service,product"),
        help="Comma-separated. The scraper looks for links with these in the URL (e.g. about, service, product).",
        key="keywords_input"
    )
    keywords = process_keywords(keywords_input)
    st.session_state['keywords'] = keywords
    if keywords:
        st.caption(f"✅ Looking for: {', '.join(keywords)}")
    else:
        st.caption("💡 Leave empty to scrape homepage only")

with col2:
    with st.expander("ℹ️ Unreachable sites", expanded=False):
        st.caption(
            "If a URL is unreachable, the app tries **Google cache** and **Archive.org**. "
            "It only scrapes the URLs in your sheet—no separate web search."
        )
    with st.expander("🔧 Advanced (rarely needed)", expanded=False):
        user_agent = st.text_input(
            "User-Agent",
            value=st.session_state.get('user_agent', "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"),
            help="Browser identifier sent to websites.",
            key="user_agent"
        )

# Step 3: Company Summary Generator (optional)
st.markdown("""
<div class="step-card">
    <h3 style="color: #f1f5f9 !important; margin-top: 0; display: flex; align-items: center;">
        <span class="step-badge">3</span>📋 Company Summary <span style="font-size: 0.75rem; font-weight: 500; color: #64748b; margin-left: 0.5rem;">(optional)</span>
    </h3>
    <div class="tip-box tip-box-fuchsia">Turn your web scrapes into structured company summaries — clear summaries, grounded facts, and commercial hypotheses. Requires an API key.</div>
</div>
""", unsafe_allow_html=True)

ai_enabled = st.checkbox(
    "Enable company summaries",
    value=st.session_state.get('ai_enabled', False),
    help="Generate structured summaries for each scraped website using AI",
    key="ai_enabled_checkbox"
)
st.session_state['ai_enabled'] = ai_enabled

if ai_enabled:
    st.markdown("---")
    st.markdown("#### Provider")
    ai_provider = st.selectbox(
        "AI provider",
        ["OpenAI", "Gemini", "OpenRouter"],
        help="OpenAI (GPT models), Google Gemini, or OpenRouter (one key for 600+ models)",
        key="ai_provider_select"
    )
    st.session_state['ai_provider'] = ai_provider
    
    st.markdown("#### API key")
    api_key_key = f"{ai_provider.lower()}_api_key"
    stored_api_key = st.session_state.get(api_key_key, "")
    
    if ai_provider == "OpenAI":
        st.caption("Create a key at [platform.openai.com/api-keys](https://platform.openai.com/api-keys)")
    elif ai_provider == "Gemini":
        st.caption("Create a key at [makersuite.google.com/app/apikey](https://makersuite.google.com/app/apikey)")
    else:
        st.caption("Create a key at [openrouter.ai/keys](https://openrouter.ai/keys) — one key for GPT, Claude, Gemini, and more")
    
    ai_api_key = st.text_input(
        f"{ai_provider} API key",
        value=stored_api_key,
        type="password",
        help="Paste your API key. It's stored in your session and never saved to disk.",
    )
    
    # Save API key
    if ai_api_key:
        if ai_api_key != stored_api_key:
            st.session_state[api_key_key] = ai_api_key
            st.session_state['ai_api_key'] = ai_api_key
            st.session_state[f"{ai_provider.lower()}_models"] = None
        st.success("API key saved")
    elif stored_api_key and not ai_api_key:
        st.session_state[api_key_key] = ""
        st.session_state['ai_api_key'] = ""
        st.session_state[f"{ai_provider.lower()}_models"] = None
    
    # Always sync generic key
    st.session_state['ai_api_key'] = st.session_state.get(api_key_key, "")
    
    st.markdown("#### Model")
    models_cache_key = f"{ai_provider.lower()}_models"
    cached_models = st.session_state.get(models_cache_key)
    
    if ai_api_key and ai_api_key.strip():
        if cached_models is None:
            with st.spinner(f"Loading available models..."):
                try:
                    if os.name == "nt":
                        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
                    
                    if ai_provider == "OpenAI":
                        models = asyncio.run(fetch_openai_models(ai_api_key))
                    elif ai_provider == "Gemini":
                        models = asyncio.run(fetch_gemini_models(ai_api_key))
                    else:
                        models = asyncio.run(fetch_openrouter_models(ai_api_key))
                    
                    if models:
                        st.session_state[models_cache_key] = models
                        cached_models = models
                    else:
                        cached_models = (
                            ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"] if ai_provider == "OpenAI"
                            else ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-pro"] if ai_provider == "Gemini"
                            else ["openai/gpt-4o-mini", "openai/gpt-4o", "google/gemini-2.0-flash-exp:free"]
                        )
                        st.session_state[models_cache_key] = cached_models
                except Exception as e:
                    st.warning("⚠️ Could not load models. Using defaults.")
                    cached_models = (
                        ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"] if ai_provider == "OpenAI"
                        else ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-pro"] if ai_provider == "Gemini"
                        else ["openai/gpt-4o-mini", "openai/gpt-4o", "google/gemini-2.0-flash-exp:free"]
                    )
                    st.session_state[models_cache_key] = cached_models
        
        if cached_models:
            default_index = 0
            try:
                if ai_provider == "OpenAI":
                    if "gpt-4o-mini" in cached_models:
                        default_index = cached_models.index("gpt-4o-mini")
                    elif "gpt-4o" in cached_models:
                        default_index = cached_models.index("gpt-4o")
                elif ai_provider == "Gemini":
                    if "gemini-1.5-flash" in cached_models:
                        default_index = cached_models.index("gemini-1.5-flash")
                    elif "gemini-1.5-pro" in cached_models:
                        default_index = cached_models.index("gemini-1.5-pro")
                else:
                    for preferred in ["openai/gpt-4o-mini", "openai/gpt-4o", "google/gemini-2.0-flash-exp:free"]:
                        if preferred in cached_models:
                            default_index = cached_models.index(preferred)
                            break
            except ValueError:
                default_index = 0
            
            ai_model = st.selectbox(
                "Model",
                options=cached_models,
                index=default_index,
                help="Recommended: gpt-4o-mini (fast) or gpt-4o (best).",
                key=f"{ai_provider}_model_select"
            )
            st.session_state['ai_model'] = ai_model
            st.session_state['ai_provider'] = ai_provider
            
            st.caption(f"{len(cached_models)} models available")
        else:
            ai_model = None
    else:
        st.info("Enter your API key above to see available models.")
        ai_model = None
    
    st.markdown("#### Prompt")
    prompt_mode = st.radio(
        "Prompt",
        ["Use default prompt", "Customize prompt"],
        horizontal=True,
        help="The default prompt is tuned for company summaries, facts, and hypotheses."
    )
    
    default_prompt_template = """You are Hypothesis Bot… an advanced commercial analysis agent.

INPUT
You will receive ONLY one input: raw website copy scraped from a company's website (may include multiple pages).

CORE JOB
Turn messy webcopy into:
1) a sharp company intelligence SUMMARY
2) a clean set of FACTS grounded in the text
3) commercially relevant HYPOTHESES inferred from signals, gaps, tone, and structure

You are not writing outreach. You are building the intelligence layer that enables personalization later.

NON NEGOTIABLE RULES
1) Do not invent facts. If it is not explicitly supported by the webcopy, it is not a fact.
2) Facts must be short and must include Evidence Quote… an exact snippet from the webcopy.
3) Hypotheses must be explicitly labeled as hypotheses and must include:
   • Signal… the specific wording or structural clue that triggered the inference
   • Commercial implication… why it matters in a sales or growth context
   • Confidence: High, Medium, or Low
4) Never mention any external tools or data sources (Apollo, LinkedIn, Crunchbase, funding, headcount, etc.). You only have webcopy.
5) If the company name is unclear, write "Company: Not explicitly stated" and proceed.
6) Avoid cringe adjectives like great, amazing, innovative. Be surgical.
7) Avoid criticism. Frame gaps neutrally as "signals" or "absence suggests".
8) Keep it concise. No fluff. No extra sections.

ANALYSIS GUIDELINES
When extracting facts, prioritize:
• What they do (offer categories, deliverables)
• Who they serve (industries, segments, buyer language)
• How they sell (engagement models, pricing mentions, process)
• Proof (case studies, client names, testimonials, quantified claims)
• Differentiators (positioning phrases, guarantees, compliance, security)
• Operational signals (hiring, support hours, global language, locations)
• Technology signals (stacks, platforms, integrations) only if stated

When generating hypotheses, prioritize:
• Likely buying triggers (growth, hiring, new initiatives, modernization)
• Likely pain points (capacity, speed, differentiation, trust, compliance, delivery)
• Likely maturity level (specialist vs generalist, product vs services)
• Likely stakeholder priorities (risk reduction, outcomes, speed, cost certainty)
• Contradictions or gaps between claims and proof
Each hypothesis must connect to a commercial implication.

LEAD INFORMATION:
- Website URL: {url}
- Company Name: {company_name}

WEBSITE CONTENT (ONLY SOURCE OF INFORMATION):
{scraped_content}

OUTPUT FORMAT (STRICT - NO MARKDOWN, NO TRUNCATION)
Return exactly these 3 sections with EXACT headers:

===SUMMARY===
Write 3 to 5 sentences in one paragraph describing:
• what the company appears to do
• who it appears to serve
• how it positions itself
• one notable proof element (only if present)
Mark inferences with (obs).
Do NOT use markdown (**, __, #). Use plain text only.

===FACTS===
Provide 10 to 18 facts.
Format each EXACTLY as:
Fact: [statement]
Evidence Quote: "[COMPLETE quote from webcopy - do NOT truncate with ...]"
Source: [Page name if present, otherwise "Webcopy"]

CRITICAL: Evidence quotes must be COMPLETE sentences from the webcopy. Do NOT truncate mid-word or mid-sentence.

===HYPOTHESES===
Provide 6 to 12 hypotheses.
Format each EXACTLY as:
Hypothesis: [inference statement] (obs)
Signal: "[specific wording or structural clue from webcopy]"
Commercial implication: [why it matters in sales/growth context]
Confidence: High or Medium or Low

CRITICAL: 
- Do NOT use markdown formatting (**, __, #)
- Do NOT truncate words mid-word
- Do NOT mix facts into hypotheses section
- Keep facts grounded (from webcopy), hypotheses inferred (from signals)"""
    
    if 'master_prompt' not in st.session_state:
        st.session_state['master_prompt'] = default_prompt_template
    
    if prompt_mode == "Customize prompt":
        st.markdown("#### Prompt template (edit below)")
        vars_def = _get_variable_definitions()
        placeholders_line = ", ".join(p for (_, p, _) in vars_def)
        st.caption(f"**Available variables:** {placeholders_line} — type these in the template below.")
        ai_prompt = st.text_area(
            "Prompt template",
            value=st.session_state.get("master_prompt", default_prompt_template),
            height=280,
            help="This is the prompt you edit. Variables like {url}, {scraped_content} are replaced when the AI runs. Preview below shows the final version.",
            key="ai_prompt_edit",
            label_visibility="collapsed"
        )
        st.session_state["master_prompt"] = ai_prompt
        prompt_cur = ai_prompt or ""
        ends_with_brace = prompt_cur.rstrip().endswith("{{") or prompt_cur.rstrip().endswith("{")
        if ends_with_brace:
            placeholders_comp = [p for (_, p, _) in vars_def]
            st.caption("Complete your variable — select one to insert:")
            complete_choice = st.selectbox(
                "Variable to insert",
                options=placeholders_comp,
                key="step3_complete_var",
                label_visibility="collapsed"
            )
            if st.button("Insert variable", key="step3_insert_var_btn"):
                base = prompt_cur.rstrip()
                if base.endswith("{{"):
                    base = base[:-2]
                    key = complete_choice.strip("{}")
                    insert = "{{" + key + "}}"
                else:
                    base = base[:-1]
                    insert = complete_choice if complete_choice.startswith("{") else "{" + complete_choice + "}"
                st.session_state["master_prompt"] = base + insert
                st.rerun()
        st.markdown("#### Final prompt (preview with sample data)")
        st.caption("This is the exact prompt sent to the AI for each lead (variables filled with sample data from your sheet).")
        sample_s3 = _get_sample_lead_data_for_preview()
        sample_s3["scraped_content"] = sample_s3.get("scraped_content") or EXAMPLE_SCRAPED_CONTENT
        final_preview_s3 = build_company_summary_prompt(ai_prompt, sample_s3, sample_s3["scraped_content"])
        st.text_area("", value=final_preview_s3[:14000] + ("…" if len(final_preview_s3) > 14000 else ""), height=240, disabled=True, key="step3_final_preview", label_visibility="collapsed")
        csv_cfg_s3 = st.session_state.get("csv_config") or {}
        df_s3 = csv_cfg_s3.get("df_preview")
        n_rows_s3 = len(df_s3) if df_s3 is not None and not df_s3.empty else 1
        row_options_s3 = list(range(n_rows_s3))
        prev_row_s3 = st.session_state.get("step3_preview_row", 0)
        if prev_row_s3 not in row_options_s3:
            prev_row_s3 = 0
        c1, c2 = st.columns([1, 3])
        with c1:
            step3_preview_row = st.selectbox(
                "Preview row",
                options=row_options_s3,
                index=row_options_s3.index(prev_row_s3) if row_options_s3 else 0,
                format_func=lambda i: f"Row {i + 1}",
                key="step3_preview_row_select"
            )
        with c2:
            st.session_state["step3_preview_row"] = step3_preview_row
            if st.button("Preview with sample lead", key="step3_sample_btn"):
                _sample_prompt_dialog("master_prompt", step3_preview_row)
    else:
        ai_prompt = st.session_state.get("master_prompt", default_prompt_template)
    
    st.session_state["ai_prompt"] = ai_prompt

else:
    ai_api_key = None
    ai_provider = None
    ai_model = None
    ai_prompt = None

# Step 4: Email Copy Writer (optional, works independently)
st.markdown("---")
st.markdown("""
<div class="step-card">
    <h3 style="color: #f1f5f9 !important; margin-top: 0; display: flex; align-items: center;">
        <span class="step-badge">4</span>✉️ Email Copy Writer <span style="font-size: 0.75rem; font-weight: 500; color: #64748b; margin-left: 0.5rem;">(optional)</span>
    </h3>
    <div class="tip-box tip-box-emerald">Generate personalized email copy for each lead based on scraped content. Runs independently from company summaries. Same AI providers supported.</div>
</div>
""", unsafe_allow_html=True)

email_copy_enabled = st.checkbox(
    "Enable email copy generation",
    value=st.session_state.get('email_copy_enabled', False),
    help="Generate email copy for each scraped lead using AI",
    key="email_copy_enabled_checkbox"
)
st.session_state['email_copy_enabled'] = email_copy_enabled

if email_copy_enabled:
    st.markdown("---")
    use_step3_credentials = st.checkbox(
        "Use same API key and model as Step 3",
        value=st.session_state.get('email_copy_use_step3', False),
        help="Reuse the provider, API key, and model from Company Summary (Step 3). Requires Step 3 to be enabled.",
        key="email_copy_use_step3_checkbox"
    )
    st.session_state['email_copy_use_step3'] = use_step3_credentials
    if use_step3_credentials and ai_enabled:
        email_copy_provider = ai_provider
        email_copy_api_key = ai_api_key
        email_copy_model = ai_model
        st.session_state['email_copy_provider'] = email_copy_provider
        st.session_state['email_copy_model'] = email_copy_model
        st.caption(f"Using Step 3: {ai_provider} / {ai_model}")
    else:
        if use_step3_credentials and not ai_enabled:
            st.warning("Enable Step 3 (Company Summary) to use its API key and model.")
        st.markdown("#### Provider")
        email_copy_provider = st.selectbox(
            "AI provider",
            ["OpenAI", "Gemini", "OpenRouter"],
            key="email_copy_provider_select"
        )
        st.session_state['email_copy_provider'] = email_copy_provider
        st.markdown("#### API key")
        ec_api_key_name = f"{email_copy_provider.lower()}_api_key"
        email_copy_api_key = st.text_input(
            "API key",
            value=st.session_state.get(ec_api_key_name, ""),
            type="password",
            key="email_copy_api_key_input"
        )
        st.session_state[ec_api_key_name] = email_copy_api_key
        st.markdown("#### Model")
        if email_copy_provider == "OpenAI":
            email_copy_model = st.selectbox("Model", ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"], key="email_copy_model_select")
        elif email_copy_provider == "Gemini":
            email_copy_model = st.selectbox("Model", ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-pro"], key="email_copy_model_select")
        else:
            email_copy_model = st.text_input("OpenRouter model (e.g. openai/gpt-4o-mini)", value="openai/gpt-4o-mini", key="email_copy_model_input")
        st.session_state['email_copy_model'] = email_copy_model
    st.markdown("#### Prompt template (edit below)")
    default_email_prompt = """You are an expert B2B sales copywriter.

Write a personalized outreach email for this company based on their website content.

LEAD INFO:
- URL: {url}
- Company: {company_name}

WEBSITE CONTENT:
{scraped_content}

Requirements:
- Keep the email under 150 words
- Be specific to their business (reference facts from their site)
- Professional, conversational tone
- Clear call-to-action
- No generic fluff"""
    if 'email_copy_prompt' not in st.session_state:
        st.session_state['email_copy_prompt'] = default_email_prompt
    vars_def_ec = _get_variable_definitions()
    placeholders_line_ec = ", ".join(p for (_, p, _) in vars_def_ec)
    st.caption(f"**Available variables:** {placeholders_line_ec} — type these in the template below.")
    email_copy_prompt = st.text_area(
        "Prompt template",
        value=st.session_state.get('email_copy_prompt', default_email_prompt),
        height=160,
        help="This is the prompt you edit. Variables are replaced when the AI runs. Preview below shows the final version.",
        key="email_copy_prompt_edit",
        label_visibility="collapsed"
    )
    st.session_state['email_copy_prompt'] = email_copy_prompt
    prompt_cur_ec = email_copy_prompt or ""
    ends_with_brace_ec = prompt_cur_ec.rstrip().endswith("{{") or prompt_cur_ec.rstrip().endswith("{")
    if ends_with_brace_ec:
        placeholders_comp_ec = [p for (_, p, _) in vars_def_ec]
        st.caption("Complete your variable — select one to insert:")
        complete_choice_ec = st.selectbox(
            "Variable to insert",
            options=placeholders_comp_ec,
            key="step4_complete_var",
            label_visibility="collapsed"
        )
        if st.button("Insert variable", key="step4_insert_var_btn"):
            base_ec = prompt_cur_ec.rstrip()
            if base_ec.endswith("{{"):
                base_ec = base_ec[:-2]
                key_ec = complete_choice_ec.strip("{}")
                insert_ec = "{{" + key_ec + "}}"
            else:
                base_ec = base_ec[:-1]
                insert_ec = complete_choice_ec if complete_choice_ec.startswith("{") else "{" + complete_choice_ec + "}"
            st.session_state["email_copy_prompt"] = base_ec + insert_ec
            st.rerun()
    st.markdown("#### Final prompt (preview with sample data)")
    st.caption("This is the exact prompt sent to the AI for each lead (variables filled with sample data).")
    sample_ec = _get_sample_lead_data_for_preview()
    sample_ec["scraped_content"] = sample_ec.get("scraped_content") or EXAMPLE_SCRAPED_CONTENT
    final_preview_ec = build_email_copy_prompt(email_copy_prompt, sample_ec, sample_ec.get("scraped_content")) or ""
    preview_text = final_preview_ec[:10000] + ("…" if len(final_preview_ec) > 10000 else "")
    st.text_area("Prompt preview", value=preview_text, height=200, disabled=True, key="step4_final_preview", label_visibility="collapsed")
    csv_cfg_ec = st.session_state.get("csv_config") or {}
    df_ec = csv_cfg_ec.get("df_preview")
    n_rows_ec = len(df_ec) if df_ec is not None and not df_ec.empty else 1
    row_options_ec = list(range(n_rows_ec))
    prev_row_ec = st.session_state.get("step4_preview_row", 0)
    if prev_row_ec not in row_options_ec:
        prev_row_ec = 0
    c1_ec, c2_ec = st.columns([1, 3])
    with c1_ec:
        step4_preview_row = st.selectbox(
            "Preview row",
            options=row_options_ec,
            index=row_options_ec.index(prev_row_ec) if row_options_ec else 0,
            format_func=lambda i: f"Row {i + 1}",
            key="step4_preview_row_select"
        )
    with c2_ec:
        st.session_state["step4_preview_row"] = step4_preview_row
        if st.button("Preview with sample lead", key="step4_sample_btn"):
            _sample_prompt_dialog("email_copy_prompt", step4_preview_row)
else:
    email_copy_api_key = None
    email_copy_provider = None
    email_copy_model = None
    email_copy_prompt = None

# Token usage estimator - Always show current settings
st.markdown("---")
with st.expander("💰 Cost Estimator", expanded=False):
    MODEL_PRICING = {
        "gpt-4o-mini": (0.15, 0.60), "gpt-4o": (2.50, 10.00), "gpt-4-turbo": (10.00, 30.00),
        "gpt-3.5-turbo": (0.50, 1.50), "gemini-1.5-flash": (0.075, 0.30),
        "gemini-1.5-pro": (1.25, 5.00), "gemini-pro": (0.50, 1.50),
    }
    def get_model_pricing(model):
        for k, v in MODEL_PRICING.items():
            if k in (model or "").lower():
                return v
        return (1.0, 3.0)
    def chars_to_tokens(chars):
        return max(0, int(chars / 4))
    
    # Get current settings from session state (with defaults)
    csv_cfg = st.session_state.get("csv_config") or {}
    num_urls_csv = st.session_state.get("_csv_url_count", 0) or csv_cfg.get("url_count", 0) or 0
    default_urls = max(1, num_urls_csv) if num_urls_csv > 0 else 100
    
    # Settings
    col1, col2 = st.columns(2)
    with col1:
        est_urls = st.number_input("URLs to process", min_value=1, max_value=300000, value=default_urls, key="est_urls")
    with col2:
        est_success_rate = st.slider("Expected success rate %", 80, 99, 90, key="est_success_rate") / 100
    
    # Current configuration
    max_chars_est = max(100, min(st.session_state.get("max_chars", 10000), 50000))
    depth_est = max(1, st.session_state.get("depth", 3))
    
    # Calculate content size estimate
    base_chars = 3000 + (depth_est - 1) * 2500
    avg_chars_per_site = min(max_chars_est, base_chars)
    
    effective_leads = max(1, int(est_urls * est_success_rate))
    base_overhead = 150
    
    # Check what's enabled
    ai_enabled_est = st.session_state.get("ai_enabled_checkbox", False)
    email_enabled_est = st.session_state.get("email_copy_enabled_checkbox", False)
    ai_model_est = st.session_state.get("ai_model", "gpt-4o-mini") or "gpt-4o-mini"
    email_model_est = st.session_state.get("email_copy_model", "gpt-4o-mini") or "gpt-4o-mini"
    ai_prompt_est = st.session_state.get("ai_prompt", "") or ""
    email_prompt_est = st.session_state.get("email_copy_prompt", "") or ""
    
    # Calculate costs
    total_cost = 0
    cost_details = []
    
    if ai_enabled_est:
        prompt_tokens = chars_to_tokens(len(ai_prompt_est) + 300) + base_overhead
        content_tokens = chars_to_tokens(avg_chars_per_site)
        inp_per = prompt_tokens + content_tokens
        out_per = chars_to_tokens(3000)
        in_p, out_p = get_model_pricing(ai_model_est)
        cost = ((inp_per * effective_leads / 1e6) * in_p) + ((out_per * effective_leads / 1e6) * out_p)
        total_cost += cost
        cost_details.append(("Company Summary", ai_model_est, cost, effective_leads))
    
    if email_enabled_est:
        prompt_tokens = chars_to_tokens(len(email_prompt_est) + 200) + base_overhead
        content_tokens = chars_to_tokens(avg_chars_per_site)
        inp_per = prompt_tokens + content_tokens
        out_per = chars_to_tokens(400)
        in_p, out_p = get_model_pricing(email_model_est)
        cost = ((inp_per * effective_leads / 1e6) * in_p) + ((out_per * effective_leads / 1e6) * out_p)
        total_cost += cost
        cost_details.append(("Email Copy", email_model_est, cost, effective_leads))
    
    # Display results
    st.markdown("#### 💵 Estimated Cost")
    
    if cost_details:
        for name, model, cost, leads in cost_details:
            st.markdown(f"**{name}** ({model}): ~${cost:.2f} for ~{leads} leads")
        st.markdown(f"### **Total: ~${total_cost:.2f}**")
        st.caption(f"Based on ~{avg_chars_per_site:,} chars/site, {depth_est} page(s) per site. Prices in USD.")
    else:
        st.info("Enable Step 3 (Company Summary) or Step 4 (Email Copy) to see cost estimates.")
        st.caption("Scraping alone doesn't use AI tokens — only the AI features do.")

# Test run section
st.markdown("---")
with st.container():
    st.markdown("### 🧪 Test Run")
    st.caption("Preview scraping and AI on a few URLs before running the full batch")
    
    test_col1, test_col2, test_col3 = st.columns([1, 1, 2])
    with test_col1:
        test_rows = st.number_input("URLs to test", min_value=1, max_value=20, value=3, help="Number of URLs to test", key="test_rows_input")
    with test_col2:
        st.write("")
        st.write("")
        can_test = uploaded_file is not None and st.session_state.get('csv_config') is not None
        test_clicked = st.button(
            "🧪 Run Test" if can_test else "⚠️ Upload CSV First",
            key="test_run_btn",
            disabled=not can_test,
            type="secondary"
        )

# Main action area - get variables (speed is set automatically when run starts, from URL count)
keywords = st.session_state.get('keywords', [])
user_agent = st.session_state.get('user_agent', "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
run_name = st.session_state.get('run_name', "")
# Defaults for test run; main run overwrites from get_auto_run_settings(total)
_auto = get_auto_run_settings(500)
concurrency, timeout, retries, depth, rows_per_file, max_chars = _auto["concurrency"], _auto["timeout"], _auto["retries"], _auto["depth"], _auto["rows_per_file"], _auto["max_chars"]
low_resource = is_cloud_mode() or _is_low_resource_default()
# Use checkbox key as source of truth (widget state); fallback to our sync key
ai_enabled = st.session_state.get("ai_enabled_checkbox", st.session_state.get('ai_enabled', False))
if not ai_enabled:
    st.session_state['ai_enabled'] = False  # Ensure sync when disabled
ai_provider = st.session_state.get('ai_provider', None) if ai_enabled else None
# Get API key - check provider-specific key first, then generic
if ai_enabled and ai_provider:
    api_key_key = f"{ai_provider.lower()}_api_key"
    ai_api_key = st.session_state.get(api_key_key) or st.session_state.get('ai_api_key', None)
else:
    ai_api_key = None
ai_model = st.session_state.get('ai_model', None) if ai_enabled else None
ai_prompt = st.session_state.get('ai_prompt', None) if ai_enabled else None

email_copy_enabled_ui = st.session_state.get("email_copy_enabled_checkbox", st.session_state.get('email_copy_enabled', False))
if not email_copy_enabled_ui:
    st.session_state['email_copy_enabled'] = False
email_copy_provider_ui = st.session_state.get('email_copy_provider', None) if email_copy_enabled_ui else None
if email_copy_enabled_ui and email_copy_provider_ui:
    ec_key_name = f"{email_copy_provider_ui.lower()}_api_key"
    email_copy_api_key = st.session_state.get(ec_key_name) or st.session_state.get('email_copy_api_key', None)
else:
    email_copy_api_key = None
email_copy_model_ui = st.session_state.get('email_copy_model', None) if email_copy_enabled_ui else None
email_copy_prompt_ui = st.session_state.get('email_copy_prompt', None) if email_copy_enabled_ui else None

# -------- TEST RUN --------
def _run_test_in_thread():
    """Run test in a background thread to avoid blocking the UI."""
    import traceback
    params = st.session_state.get("_test_params", {})
    try:
        if os.name == "nt":
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        results = asyncio.run(run_test_preview(
            params["urls"], params["test_rows"], params["retries"], params["timeout"],
            params["depth"], params["keywords"], params["max_chars"], params["user_agent"],
            params["ai_enabled"], params["ai_api_key"], params["ai_provider"],
            params["ai_model"], params["ai_prompt"],
            params["email_copy_enabled"], params["email_copy_api_key"], params["email_copy_provider"],
            params["email_copy_model"], params["email_copy_prompt"],
            params["lead_data_map"], log_callback=params.get("log_callback")))
        st.session_state["_test_results"] = results
        st.session_state["_test_logs"] = params.get("_logs", [])
    except Exception as e:
        st.session_state["_test_results"] = []
        st.session_state["_test_error"] = str(e)
        st.session_state["_test_tb"] = traceback.format_exc()
        st.session_state["_test_logs"] = params.get("_logs", []) if params else []
    finally:
        st.session_state["_test_running"] = False

if uploaded_file and test_clicked:
    csv_config = st.session_state.get('csv_config', {}) or {}
    if not csv_config:
        st.error("❌ Please configure your CSV above (select URL column) first.")
    else:
        uploaded_file.seek(0)
        has_headers = csv_config.get('has_headers', True)
        url_col = csv_config.get('url_column')
        lead_cols = csv_config.get('lead_data_columns', {}) or {}
        if not url_col:
            st.error("❌ URL column not configured. Select URL column in Step 1.")
        else:
            try:
                if has_headers:
                    df_in = pd.read_csv(uploaded_file)
                else:
                    df_in = pd.read_csv(uploaded_file, header=None)
            except Exception as e:
                st.error(f"❌ Failed to read CSV: {e}")
                df_in = pd.DataFrame()
            if df_in.empty:
                st.error("❌ CSV is empty or unreadable.")
            else:
                try:
                    url_col_idx = list(df_in.columns).index(url_col) if has_headers else int(str(url_col).replace("Column ", "")) - 1
                except (ValueError, TypeError):
                    st.error("❌ Invalid URL column. Reconfigure CSV in Step 1.")
                    url_col_idx = 0
                url_series = df_in.iloc[:, url_col_idx].fillna("").astype(str)
                url_list_with_idx = [(u.strip(), i) for i, u in enumerate(url_series) if u and str(u).strip()]
                urls = [u for u, _ in url_list_with_idx]
                lead_data_map = {}
                for idx, (url, orig_row) in enumerate(url_list_with_idx):
                    lead_data = {'url': url}
                    default_company = url.replace('https://', '').replace('http://', '').split('/')[0] if url else ''
                    for key, col_name in (lead_cols or {}).items():
                        try:
                            ci = list(df_in.columns).index(col_name) if has_headers else int(str(col_name).replace("Column ", "")) - 1
                            val = df_in.iloc[orig_row, ci] if orig_row < len(df_in) else None
                            v = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val).strip()
                            lead_data[key] = v
                        except (ValueError, TypeError, IndexError):
                            lead_data[key] = ""
                    if not lead_data.get('company_name'):
                        lead_data['company_name'] = default_company
                    lead_data_map[idx] = lead_data

                if not urls:
                    st.error("❌ No valid URLs found in CSV. Check your URL column.")
                else:
                    ai_enabled_test = st.session_state.get("ai_enabled_checkbox", False)
                    ai_provider_test = st.session_state.get('ai_provider', None) if ai_enabled_test else None
                    ai_api_key_test = st.session_state.get(f"{ai_provider_test.lower()}_api_key") or st.session_state.get('ai_api_key') if (ai_enabled_test and ai_provider_test) else None
                    ai_model_test = st.session_state.get('ai_model') if ai_enabled_test else None
                    ai_prompt_test = st.session_state.get('ai_prompt') or ""
                    email_copy_enabled_test = st.session_state.get("email_copy_enabled_checkbox", False)
                    email_copy_provider_test = st.session_state.get('email_copy_provider', None) if email_copy_enabled_test else None
                    email_copy_api_key_test = st.session_state.get(f"{email_copy_provider_test.lower()}_api_key") or st.session_state.get('email_copy_api_key') if (email_copy_enabled_test and email_copy_provider_test) else None
                    email_copy_model_test = st.session_state.get('email_copy_model') if email_copy_enabled_test else None
                    email_copy_prompt_test = st.session_state.get('email_copy_prompt') or ""

                    test_logs = []
                    def test_log(msg):
                        test_logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

                    _test_timeout = timeout
                    _test_retries = min(retries + 2, 8)
                    st.session_state["_test_params"] = {
                        "urls": urls, "test_rows": test_rows, "retries": _test_retries, "timeout": _test_timeout,
                        "depth": depth, "keywords": keywords, "max_chars": max_chars, "user_agent": user_agent,
                        "ai_enabled": ai_enabled_test, "ai_api_key": ai_api_key_test, "ai_provider": ai_provider_test,
                        "ai_model": ai_model_test, "ai_prompt": ai_prompt_test,
                        "email_copy_enabled": email_copy_enabled_test, "email_copy_api_key": email_copy_api_key_test,
                        "email_copy_provider": email_copy_provider_test, "email_copy_model": email_copy_model_test,
                        "email_copy_prompt": email_copy_prompt_test,
                        "lead_data_map": lead_data_map, "log_callback": test_log, "_logs": test_logs
                    }
                    st.session_state["_test_running"] = True
                    st.session_state["_test_error"] = None
                    st.session_state["_test_results"] = None  # Clear previous
                    t = threading.Thread(target=_run_test_in_thread, daemon=True)
                    t.start()
                    st.rerun()

# Test run status display
if st.session_state.get("_test_running"):
    st.markdown("---")
    with st.container():
        col1, col2 = st.columns([3, 1])
        with col1:
            st.info("🧪 **Test is running...** (scraping + AI may take 1-3 min). You can keep editing settings. The results will appear below when ready.")
        with col2:
            if st.button("🔄 Check Status", key="test_check_results_btn", use_container_width=True):
                st.rerun()

# Test results display
if st.session_state.get("_test_results") is not None and not st.session_state.get("_test_running"):
    st.markdown("---")
    results = st.session_state["_test_results"]
    test_logs = st.session_state.get("_test_logs", [])
    test_error = st.session_state.pop("_test_error", None)
    test_tb = st.session_state.pop("_test_tb", None)
    
    if test_error:
        st.error(f"❌ Test failed: {test_error}")
        if test_tb:
            with st.expander("Debug Info"):
                st.code(test_tb, language=None)
    elif results:
        st.success(f"✅ Test completed — {len(results)} URL(s) processed")
    
    if test_logs and not results:
        with st.expander("View Logs"):
            st.code("\n".join(test_logs), language=None)
    
    email_copy_enabled_test = st.session_state.get("email_copy_enabled_checkbox", False)
    
    for i, row in enumerate(results):
        url = row[0]
        scraped_text = row[1] if len(row) >= 2 else ""
        ai_summary = row[2] if len(row) >= 3 else ""
        email_copy_val = row[3] if len(row) >= 4 else ""
        
        # Determine status icon
        has_error = scraped_text.startswith("❌") if scraped_text else False
        status_icon = "✅" if not has_error else "❌"
        
        with st.expander(f"{status_icon} {i+1}. {url[:55]}{'…' if len(url) > 55 else ''}", expanded=(i == 0)):
            tabs = st.tabs(["📄 Scraped", "🤖 Summary", "✉️ Email"])
            
            with tabs[0]:
                if scraped_text and not scraped_text.startswith("❌"):
                    st.text_area("Content", value=scraped_text[:5000] + ("…" if len(scraped_text) > 5000 else ""), 
                                height=250, key=f"test_scraped_{i}", disabled=True, label_visibility="collapsed")
                    st.caption(f"{len(scraped_text):,} characters scraped")
                else:
                    st.error(scraped_text if scraped_text else "No content scraped")
            
            with tabs[1]:
                if ai_summary and not ai_summary.startswith("❌"):
                    formatted = ai_summary.replace("===SUMMARY===", "**Summary**").replace("===FACTS===", "**Facts**").replace("===HYPOTHESES===", "**Hypotheses**")
                    st.markdown(formatted)
                elif st.session_state.get('ai_enabled_checkbox', False):
                    st.info("No AI summary generated — check scraped content or API settings")
                else:
                    st.caption("AI summary disabled (enable in Step 3)")
            
            with tabs[2]:
                if email_copy_val and not email_copy_val.startswith("❌"):
                    st.markdown(email_copy_val)
                elif email_copy_enabled_test:
                    st.info("No email copy generated — check settings")
                else:
                    st.caption("Email copy disabled (enable in Step 4)")
    
    if not results and not test_error:
        st.warning("No results returned. Check your CSV and settings.")

# Phase 0: Measurement Panel - Show failure buckets and stats
if st.session_state.get('_test_results') or st.session_state.get('_last_run_stats'):
    st.markdown("---")
    with st.expander("📊 Scrape Analysis & Failure Buckets", expanded=False):
        stats = st.session_state.get('_last_run_stats', {})
        
        if stats:
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total URLs", stats.get('total', 0))
            with col2:
                st.metric("Success", stats.get('successes', 0), f"{stats.get('success_rate', 0)}%")
            with col3:
                st.metric("Partial", stats.get('partials', 0))
            with col4:
                st.metric("Failed", stats.get('failures', 0))
            
            # Top failure buckets
            top_failures = stats.get('top_failures', [])
            if top_failures:
                st.markdown("#### Top Failure Reasons")
                for bucket, count in top_failures:
                    st.markdown(f"- **{bucket}**: {count} URLs")
            
            # Download failures CSV
            failure_csv = st.session_state.get('_failure_csv', '')
            if failure_csv:
                st.download_button(
                    "📥 Download Failures CSV",
                    failure_csv,
                    file_name=f"scrape_failures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )

# Validation helper
def get_validation_status():
    """Check if all required fields are properly configured."""
    errors = []
    warnings = []
    
    if uploaded_file is None:
        errors.append("Upload a CSV file in Step 1 (if you had one before, the session may have reset — re-upload above)")
    else:
        csv_config = st.session_state.get('csv_config', {})
        if not csv_config:
            errors.append("Configure your CSV (select URL column) in Step 1")
        elif not csv_config.get('url_column'):
            errors.append("Select a URL column in Step 1")
    
    # Check AI configuration if enabled
    if st.session_state.get('ai_enabled_checkbox', False):
        ai_provider = st.session_state.get('ai_provider')
        if not ai_provider:
            warnings.append("AI Provider not selected (Step 3)")
        else:
            api_key = st.session_state.get(f"{ai_provider.lower()}_api_key") or st.session_state.get('ai_api_key')
            if not api_key:
                warnings.append(f"{ai_provider} API key not provided (Step 3)")
        
        if not st.session_state.get('ai_model'):
            warnings.append("AI Model not selected (Step 3)")
    
    # Check Email configuration if enabled
    if st.session_state.get('email_copy_enabled_checkbox', False):
        use_step3 = st.session_state.get('email_copy_use_step3', False)
        if not use_step3 or not st.session_state.get('ai_enabled_checkbox', False):
            email_provider = st.session_state.get('email_copy_provider')
            if not email_provider:
                warnings.append("Email Copy Provider not selected (Step 4)")
            else:
                api_key = st.session_state.get(f"{email_provider.lower()}_api_key")
                if not api_key:
                    warnings.append(f"{email_provider} API key not provided for Email Copy (Step 4)")
    
    return errors, warnings

# -------- SCRAPE BUTTON (CTA section for primary button styling) --------
validation_errors, validation_warnings = get_validation_status()
can_start = uploaded_file is not None and not validation_errors

with st.container():
    st.markdown("---")
    st.markdown("""
    <div class="cta-heading-block" style="text-align: center; padding: 2.5rem 0;">
        <h2 style="color: #f1f5f9; margin-bottom: 0.75rem; font-size: 1.6rem;">🚀 Ready to Start</h2>
        <p style="color: #94a3b8; font-size: 1rem; margin-bottom: 1.75rem; line-height: 1.6;">
            Upload your CSV, configure settings, and click below to scrape.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Show validation messages
    if validation_errors:
        st.markdown('<div class="validation-message"><strong>⚠️ Please fix these issues before starting:</strong><ul>' + 
                    ''.join(f'<li>{e}</li>' for e in validation_errors) + '</ul></div>', 
                    unsafe_allow_html=True)
    elif validation_warnings:
        with st.expander("⚠️ Optional settings not configured (click to see)", expanded=False):
            st.markdown("You can still run the scraper, but these optional features won't work:")
            for w in validation_warnings:
                st.markdown(f"- {w}")
    
    col_btn, _1, _2 = st.columns([1, 1, 1])
    with col_btn:
        start_clicked = st.button(
            "🚀 Start Scraping" if can_start else "⚠️ Check Errors Above",
            use_container_width=True,
            key="start_scraping_btn",
            disabled=not can_start,
            type="primary" if can_start else "secondary",
            help="Upload a CSV in Step 1 and configure the URL column to enable." if not can_start else "Start the scraper run."
        )

if uploaded_file and start_clicked and can_start:
    # Get CSV configuration from session state
    csv_config = st.session_state.get('csv_config', {})
    
    if not csv_config:
        st.error("❌ Please configure your CSV file above (select columns)")
        st.stop()
    
    # Read CSV based on configuration
    has_headers = csv_config.get('has_headers', True)
    url_col = csv_config.get('url_column')
    lead_cols = csv_config.get('lead_data_columns', {}) or {}
    if url_col is None or (isinstance(url_col, str) and not url_col.strip()):
        st.error("❌ URL column not configured. Please select a URL column in Step 1.")
        st.stop()
    # Reset file pointer
    uploaded_file.seek(0)
    try:
        if has_headers:
            df_in = pd.read_csv(uploaded_file)
        else:
            df_in = pd.read_csv(uploaded_file, header=None)
    except Exception as e:
        st.error(f"❌ Failed to read CSV: {e}")
        st.stop()
    if df_in.empty or len(df_in) == 0:
        st.error("❌ CSV file is empty.")
        st.stop()
    # Get URL column index
    try:
        if has_headers:
            if url_col not in df_in.columns:
                st.error(f"❌ URL column '{url_col}' not found in CSV. Columns: {list(df_in.columns)}")
                st.stop()
            url_col_idx = list(df_in.columns).index(url_col)
        else:
            url_col_idx = int(str(url_col).replace("Column ", "").strip()) - 1
            if url_col_idx < 0 or url_col_idx >= len(df_in.columns):
                st.error(f"❌ Invalid column index. Select Column 1 to {len(df_in.columns)}.")
                st.stop()
    except (ValueError, TypeError) as e:
        st.error(f"❌ Invalid URL column selection: {e}")
        st.stop()
    
    # Extract URLs with original row indices (for lead columns)
    url_series = df_in.iloc[:, url_col_idx].fillna("").astype(str)
    url_list_with_idx = [(u.strip(), i) for i, u in enumerate(url_series) if u and str(u).strip()]
    urls = [u for u, _ in url_list_with_idx]
    total = len(urls)
    if total == 0:
        st.error("❌ No valid URLs found in the selected column. Check your CSV and URL column.")
        st.stop()
    
    # Phase 7: Support up to 300k leads (cloud cap slightly lower for stability; local full 300k)
    MAX_URLS_LOCAL = 300000
    CLOUD_MODE_MAX_URLS = 300000  # Allow 300k in cloud with gentle settings
    if total > MAX_URLS_LOCAL:
        st.warning(f"⚠️ **Limit:** Your CSV has {total:,} URLs. Maximum is {MAX_URLS_LOCAL:,} per run.")
        st.info("💡 Split into multiple CSVs and run sequentially for more than 300k.")
        urls = urls[:MAX_URLS_LOCAL]
        url_list_with_idx = url_list_with_idx[:MAX_URLS_LOCAL]
        total = len(urls)
    elif is_cloud_mode() and total > CLOUD_MODE_MAX_URLS:
        st.warning(f"⚠️ **Cloud limit:** Your CSV has {total:,} URLs. Maximum is {CLOUD_MODE_MAX_URLS:,} per run.")
        urls = urls[:CLOUD_MODE_MAX_URLS]
        url_list_with_idx = url_list_with_idx[:CLOUD_MODE_MAX_URLS]
        total = len(urls)
    if total > 100000:
        st.info(f"ℹ️ **Large run ({total:,} URLs):** Reliability-optimized settings. Progress saved every few seconds; you can resume if the session drops.")
        st.caption("Results are saved incrementally; you can resume if the session drops.")
    
    # Prepare lead data mapping (filtered index -> lead data dict). For 100k+ URLs, build on demand in workers to avoid memory.
    LARGE_RUN_THRESHOLD = 100000
    lead_data_map = {}
    if total <= LARGE_RUN_THRESHOLD:
        for idx, (url, orig_row) in enumerate(url_list_with_idx):
            lead_data = build_lead_data_for_row(df_in, orig_row, lead_cols or {}, has_headers, url)
            lead_data_map[idx] = lead_data
    # Store lead data map in session state (empty for large runs; workers use original_df + url_list_with_idx + csv_config)
    st.session_state['lead_data_map'] = lead_data_map
    
    # Show tips for large files (single block, no duplicate messages)
    if total > 10000:
        with st.expander("💡 Tips for Large Datasets (up to 300k leads)", expanded=True):
            st.markdown(f"""
            **Your dataset:** {total:,} URLs

            **Reliability-first (supports up to 300k per run):**
            - ✅ **Concurrency:** Auto-tuned by size (e.g. 4–6 for 100k+, lower for 200k+). Don't override for huge runs.
            - ✅ **Max chars per site:** 10,000–20,000 keeps memory and files manageable.
            - ✅ **Rows per file:** Smaller chunks for 200k+ for frequent saves and resume.
            - ✅ **Be patient:** 100k–300k URLs can take many hours; progress is saved so you can resume.

            **Crash recovery:** Progress is saved every few seconds. If the session drops, re-upload your CSV and click **Start** to resume from the last checkpoint. Use **Resume or Download Partial Results** to get what’s done so far.
            """)
    elif total > 5000:
        st.info(f"ℹ️ Processing {total:,} URLs. Results are saved in chunks; you can resume if needed.")

    # AUTO-RESUME: Find existing run with same URLs before creating new folder
    output_dir = None
    run_folder = None
    url_set = set(urls)
    outputs_dir = "outputs"
    if os.path.isdir(outputs_dir):
        best_match = None  # (output_dir, completed_count)
        for fn in os.listdir(outputs_dir):
            run_path = os.path.join(outputs_dir, fn)
            if not os.path.isdir(run_path):
                continue
            ck = load_checkpoint(get_checkpoint_path(run_path))
            if not ck:
                continue
            stored = set(ck.get("urls", []))
            if stored != url_set:
                continue
            completed = len(ck.get("completed_urls", []))
            total_stored = len(ck.get("urls", []))
            if 0 < completed < total_stored:
                if best_match is None or completed > best_match[1]:
                    best_match = (run_path, completed)
        if best_match:
            output_dir = best_match[0]
            run_folder = os.path.basename(output_dir)
            # Reconcile checkpoint with actual file contents (checkpoint can overstate after crash)
            reconcile_checkpoint_with_files(output_dir)
            actual_rows, _ = get_actual_completed_from_files(output_dir)
            display_count = actual_rows if actual_rows > 0 else best_match[1]
            st.success(f"🔄 **Resuming:** Found existing run with {display_count:,}/{total:,} rows in files. Continuing from where you left off.")

    if output_dir is None:
        run_folder = datetime.now().strftime("run_%Y%m%d_%H%M%S")
        output_dir = os.path.join("outputs", run_folder)
        os.makedirs(output_dir, exist_ok=True)

    # All speed settings from list size (no user tuning)
    auto_settings = get_auto_run_settings(total)
    concurrency = auto_settings["concurrency"]
    timeout = auto_settings["timeout"]
    retries = auto_settings["retries"]
    depth = auto_settings["depth"]
    rows_per_file = auto_settings["rows_per_file"]
    max_chars = auto_settings["max_chars"]

    progress_bar = st.progress(0)
    status_text = st.empty()
    eta_text = st.empty()
    
    # Real-time activity dashboard placeholders
    dashboard_placeholder = st.empty()
    
    # AI status callback (thread-safe)
    ai_status_messages = {}

    fun_messages = [
        "🔍 Scanning the web...",
        "🛠️ Sharpening the scrapers...",
        "🚀 Launching data rockets...",
        "📡 Tuning into websites...",
        "🧩 Piecing text together..."
    ]
    start_time = time.time()
    progress_state_ui = {"done": 0, "total": max(total, 1), "running": True}
    progress_samples = deque(maxlen=200)
    runtime_logs = deque(maxlen=2000)  # Keep enough for full tracebacks + run history
    
    import threading
    runtime_lock = threading.Lock()
    progress_lock = threading.Lock()
    scrape_status_lock = threading.Lock()
    ai_status_lock = threading.Lock()
    
    scrape_in_progress = {}  # url -> {status, message}
    scrape_recent = deque(maxlen=30)  # [{url, status, message}, ...]
    scrape_errors = deque(maxlen=100)  # [{url, message}, ...] - show more in live dashboard
    all_failed_urls = []  # Unbounded list for full report and download

    def progress_cb(done_count, total_count):
        with progress_lock:
            progress_state_ui["done"] = done_count
            progress_state_ui["total"] = max(total_count, 1)
            progress_samples.append((time.time(), done_count))

    def scrape_status_callback(url: str, status: str, message: str):
        with scrape_status_lock:
            if status in ("scraping", "ai_summarizing", "email_copy"):
                scrape_in_progress[url] = {"status": status, "message": message}
            else:
                scrape_in_progress.pop(url, None)
                entry = {"url": url, "status": status, "message": message}
                scrape_recent.append(entry)
                if status == "error":
                    full_msg = (message or "").strip()
                    err_entry = {"url": url, "message": full_msg}
                    scrape_errors.append(err_entry)
                    # One entry per URL: update existing or append so final CSV has exact error (no truncation)
                    existing = next((e for e in all_failed_urls if e.get("url") == url), None)
                    if existing:
                        existing["message"] = full_msg
                    else:
                        all_failed_urls.append(err_entry)

    def ai_status_callback(url, message):
        with ai_status_lock:
            ai_status_messages[url] = message

    scrape_error = [None]  # Mutable to capture exception from thread
    
    def log_cb(msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        with runtime_lock:
            runtime_logs.append(line)
    
    ai_enabled_for_run = st.session_state.get("ai_enabled_checkbox", False)
    if not ai_enabled_for_run:
        ai_api_key_run, ai_provider_run, ai_model_run, ai_prompt_run = None, None, None, None
    else:
        ai_api_key_run, ai_provider_run, ai_model_run, ai_prompt_run = ai_api_key, ai_provider, ai_model, ai_prompt

    # Use checkbox value at run start so email copy setting is reliable
    email_copy_enabled_for_run = bool(st.session_state.get("email_copy_enabled_checkbox", False))
    if not email_copy_enabled_for_run:
        email_copy_api_key_run = email_copy_provider_run = email_copy_model_run = email_copy_prompt_run = None
    else:
        email_copy_api_key_run = email_copy_api_key
        email_copy_provider_run = email_copy_provider_ui
        email_copy_model_run = email_copy_model_ui
        email_copy_prompt_run = email_copy_prompt_ui
        if not (email_copy_api_key_run and email_copy_provider_run and email_copy_model_run):
            st.warning("⚠️ **Email copy** was checked but API key / provider / model is missing. This run will have **3 columns** (no Email Copy). Set Step 4 API key (or use same as Step 3) and run again for email copy.")
            log_cb("⚠️ Email copy disabled for this run: API key/provider/model missing.")
            email_copy_enabled_for_run = False
            email_copy_api_key_run = email_copy_provider_run = email_copy_model_run = email_copy_prompt_run = None

    # Auto-resilience: extra retries for unreliable sites (no checkbox)
    effective_timeout = timeout
    effective_retries = min(retries + 2, 8)
    fast_mode = _should_use_fast_mode(total)

    def run_async_scraper():
        try:
            if os.name == "nt":
                asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
            progress_state_ui["running"] = True
            log_cb(f"🚀 Run started: urls={total:,}, workers={concurrency}, depth={depth}, retries={effective_retries}, timeout={effective_timeout}s")
            log_cb(f"📧 Email copy: {'enabled (output will have 4 columns)' if email_copy_enabled_for_run else 'disabled (output will have 3 columns)'}")
            asyncio.run(
                run_scraper(urls, concurrency, effective_retries, effective_timeout, depth, keywords, max_chars,
                            user_agent, rows_per_file, output_dir, progress_cb,
                            ai_enabled_for_run, ai_api_key_run, ai_provider_run, ai_model_run, ai_prompt_run,
                            email_copy_enabled_for_run, email_copy_api_key_run, email_copy_provider_run,
                            email_copy_model_run, email_copy_prompt_run,
                            lead_data_map,
                            ai_status_callback if ai_enabled_for_run else None, scrape_status_callback,
                            run_folder=run_folder, fast_mode=fast_mode,
                            low_resource=low_resource, log_callback=log_cb,
                            use_playwright_fallback=True, use_common_crawl_fallback=True,
                            original_df=df_in, csv_config=csv_config, url_list_with_idx=url_list_with_idx))
            log_cb("✅ Run completed")
            if all_failed_urls:
                log_cb(f"⚠️ {len(all_failed_urls):,} URLs failed. See 'Failed URLs' section below for details and download.")
        except BaseException as e:
            scrape_error[0] = e
            import traceback
            tb = traceback.format_exc()
            log_cb(f"❌ Run crashed: {type(e).__name__}: {e}")
            log_cb("--- Full traceback (send this for bug reports) ---")
            for line in tb.strip().split("\n"):
                log_cb(line)
            log_cb("--- End traceback ---")
            traceback.print_exc()
            # Persist crash logs to output dir so they survive page refresh
            try:
                with runtime_lock:
                    crash_log_content = "\n".join(runtime_logs)
                crash_log_path = os.path.join(output_dir, f"crash_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
                with open(crash_log_path, "w", encoding="utf-8") as f:
                    f.write(crash_log_content)
                log_cb(f"📄 Crash log saved to: {crash_log_path}")
            except Exception:
                pass
        finally:
            progress_state_ui["running"] = False

    with st.spinner("Scraping — runs in-page. Progress is saved; if you refresh or lose connection, re-upload the CSV and click Start to resume."):
        lead_data_map = st.session_state.get('lead_data_map', None)
        with st.expander("📋 Detailed logs", expanded=False):
            logs_placeholder = st.empty()
        
        thread = threading.Thread(target=run_async_scraper, daemon=False)
        thread.start()
        while thread.is_alive():
            try:
                with progress_lock:
                    d, t = progress_state_ui["done"], progress_state_ui["total"]
                pct = d / max(t, 1)
                now_ts = time.time()
                elapsed = now_ts - start_time
                with progress_lock:
                    if not progress_samples or progress_samples[-1][1] != d:
                        progress_samples.append((now_ts, d))
                    window = [p for p in progress_samples if now_ts - p[0] <= 120]
                # Throughput from trailing window (more stable ETA than global average)
                if len(window) >= 2:
                    dt = max(window[-1][0] - window[0][0], 1e-6)
                    dd = max(window[-1][1] - window[0][1], 0)
                    rate = dd / dt
                else:
                    rate = 0.0
                remaining = ((t - d) / rate) if rate > 1e-6 else float("inf")
                idle_for = int(now_ts - (progress_samples[-1][0] if progress_samples else start_time))
                progress_bar.progress(min(pct, 1.0))
                idx = (d - 1) % len(fun_messages) if d > 0 else 0
                if remaining == float("inf"):
                    status_text.text(f"{fun_messages[idx]} ({d}/{t}) — ETA: calculating...")
                else:
                    status_text.text(f"{fun_messages[idx]} ({d}/{t}) — ETA: {int(remaining // 60)}m {int(remaining % 60)}s")
                if idle_for >= 30:
                    with scrape_status_lock:
                        n_active = len(scrape_in_progress)
                    if n_active > 0:
                        eta_text.warning(f"No completed URLs for {idle_for}s — {n_active} still in progress. Do not refresh.")
                    else:
                        eta_text.warning(f"No progress for {idle_for}s. App will wait for recovery (slow network/system); do not refresh.")
                else:
                    eta_text.text(f"⏱️ Elapsed: {int(elapsed // 60)}m {int(elapsed % 60)}s")
                with runtime_lock:
                    logs_text = "\n".join(runtime_logs)
                logs_placeholder.code(logs_text or "(no logs yet)", language=None)
                
                # Real-time activity dashboard - Compact view
                with dashboard_placeholder.container():
                    with scrape_status_lock:
                        in_progress = list(scrape_in_progress.items())
                        recent = list(scrape_recent)
                    failed_list = list(all_failed_urls)
                    n_scraping = sum(1 for _, d in in_progress if d["status"] == "scraping")
                    n_ai = sum(1 for _, d in in_progress if d["status"] == "ai_summarizing")
                    n_email = sum(1 for _, d in in_progress if d["status"] == "email_copy")
                    rate_per_min = rate * 60 if rate > 0 else 0
                    pct_done = (d / max(t, 1)) * 100
                    
                    st.markdown("#### 📊 Progress")
                    m1, m2, m3, m4 = st.columns(4)
                    with m1:
                        st.metric("Done", f"{d:,}/{t:,}", f"{pct_done:.0f}%")
                    with m2:
                        st.metric("Rate", f"{rate_per_min:.0f}/min")
                    with m3:
                        active_text = f"🔍{n_scraping} 🤖{n_ai} ✉️{n_email}" if in_progress else "Idle"
                        st.metric("Active", active_text)
                    with m4:
                        error_count = len(failed_list)
                        st.metric("Issues (failed)", error_count, delta_color="inverse" if error_count > 0 else "normal")
                    
                    if in_progress:
                        with st.expander(f"⏳ Currently processing ({len(in_progress)})", expanded=False):
                            for url, data in in_progress[-5:]:
                                short_url = escape((url[:50] + "…") if len(url) > 50 else url)
                                status = data.get("status", "scraping")
                                msg = escape(str(data.get("message", ""))[:60])
                                icon = "🔍" if status == "scraping" else ("🤖" if status == "ai_summarizing" else "✉️")
                                st.caption(f"{icon} {short_url} — {msg}")
                    
                    if failed_list:
                        with st.expander(f"⚠️ Issues ({len(failed_list)}) — not scraped or &lt;200 chars", expanded=False):
                            for e in failed_list[-10:]:
                                u = e.get("url", "")
                                m = (e.get("message", "") or "").strip()
                                short_url = escape((u[:40] + "…") if len(u) > 40 else u)
                                msg = escape(m[:500] + ("…" if len(m) > 500 else ""))
                                st.markdown(f"<span style='color:#f87171;'>✗</span> {short_url}<br><span style='color:#94a3b8; font-size:0.8rem; white-space:pre-wrap;'>{msg}</span>", unsafe_allow_html=True)
                            if len(failed_list) > 10:
                                st.caption(f"_Showing last 10 of {len(failed_list)}. Download CSV below for full list with exact errors._")
            except Exception:
                pass
            # Yield so Streamlit can send UI updates and the scraper thread can run (avoid tight loop)
            time.sleep(1.0)
        thread.join()
        # Run finished: progress bar 100%, and explicit "Complete" messaging (no ETA/remaining)
        progress_bar.progress(1.0)
        status_text.markdown(f"**✅ Run complete.** All **{total:,}** URL(s) have been processed. You can download your results below.")
        elapsed_final = int(time.time() - start_time)
        eta_text.caption(f"Total time: {int(elapsed_final // 60)}m {int(elapsed_final % 60)}s")
        with runtime_lock:
            logs_text = "\n".join(runtime_logs) or "(no logs yet)"
        logs_placeholder.code(logs_text, language=None)
        
        # Always show download logs button so users can share for bug reports
        st.download_button(
            "📥 Download run logs",
            logs_text,
            file_name=f"scraper_logs_{run_folder}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            help="Download logs to share when reporting bugs or crashes",
            key="download_run_logs"
        )
        
        if scrape_error[0]:
            err_msg = str(scrape_error[0]).encode("ascii", errors="replace").decode("ascii")
            st.error(f"❌ Scraper error: {err_msg}")
            st.warning("⚠️ **Partial results may have been saved.** Scroll up to **'Resume or Download Partial Results'** at the top to download what you have or re-upload your CSV and click Start to resume automatically.")
            st.info("📋 **Found a bug?** Click **Download run logs** above and send the file when reporting the issue.")

    # Zip all parts at the end with custom name (CSV and Excel files)
    zip_name = f"{run_folder}.zip"
    zip_path = os.path.join(output_dir, zip_name)
    csv_files = []
    excel_files = []
    
    # CRITICAL: Wait a moment for files to be fully written to disk
    import time
    time.sleep(1)
    
    # List files in output directory
    try:
        output_files = os.listdir(output_dir)
        st.info(f"📁 Found {len(output_files)} files in output directory: {', '.join(output_files[:5])}{'...' if len(output_files) > 5 else ''}")
    except Exception as e:
        st.error(f"❌ Error listing output directory: {e}")
        output_files = []
    
    # Create ZIP file: only broken-down part files (output_part_*.csv, output_part_*.xlsx)
    csv_files = [f for f in output_files if f.startswith("output_part_") and f.endswith(".csv")]
    excel_files = [f for f in output_files if f.startswith("output_part_") and f.endswith(".xlsx")]
    # Enriched mode: writer only produces enriched_output.csv / enriched_output.xlsx (no part files)
    enriched_csv_path = os.path.join(output_dir, "enriched_output.csv")
    enriched_xlsx_path = os.path.join(output_dir, "enriched_output.xlsx")
    has_enriched_csv = "enriched_output.csv" in output_files and os.path.isfile(enriched_csv_path) and os.path.getsize(enriched_csv_path) > 0
    has_enriched_xlsx = "enriched_output.xlsx" in output_files and os.path.isfile(enriched_xlsx_path) and os.path.getsize(enriched_xlsx_path) > 0
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in csv_files:
                file_path = os.path.join(output_dir, f)
                if os.path.isfile(file_path):
                    try:
                        zf.write(file_path, arcname=f)
                    except Exception as e:
                        st.warning(f"⚠️ Could not add {f} to ZIP: {e}")
            for f in excel_files:
                file_path = os.path.join(output_dir, f)
                if os.path.isfile(file_path):
                    try:
                        zf.write(file_path, arcname=f)
                    except Exception as e:
                        st.warning(f"⚠️ Could not add {f} to ZIP: {e}")
    except Exception as e:
        st.error(f"❌ Error creating ZIP file: {e}")
        import traceback
        st.error(f"Traceback: {traceback.format_exc()}")
    
    # When run used enriched mode (original CSV + new columns), writer only writes enriched_output.* — no part files. Populate session so download buttons work.
    if (has_enriched_csv or has_enriched_xlsx) and (not csv_files and not excel_files):
        try:
            if has_enriched_csv:
                with open(enriched_csv_path, 'rb') as f:
                    st.session_state['combined_csv_data'] = f.read()
                st.session_state['combined_csv_filename'] = f"{run_folder}_enriched.csv"
                st.session_state['combined_df'] = pd.read_csv(enriched_csv_path, encoding='utf-8-sig')
            if has_enriched_xlsx:
                try:
                    with open(enriched_xlsx_path, 'rb') as f:
                        st.session_state['combined_excel_data'] = f.read()
                    st.session_state['combined_excel_filename'] = f"{run_folder}_combined.xlsx"
                except Exception:
                    pass
        except Exception as e:
            st.warning(f"Could not load enriched output for download: {e}")
    
    # CRITICAL: Consider enriched output as valid result files (enriched mode writes only those, no part files)
    has_any_results = csv_files or excel_files or has_enriched_csv or has_enriched_xlsx
    if not has_any_results:
        st.error("⚠️ **WARNING: No CSV or Excel files were generated!**")
        st.error("This might indicate that the writer coroutine failed or no data was processed.")
        st.info("Check the output directory for any error messages or partial files.")
        try:
            actual_files = os.listdir(output_dir)
            if actual_files:
                st.info(f"Files found in output directory: {', '.join(actual_files)}")
            else:
                st.warning("Output directory is empty!")
        except Exception as e:
            st.error(f"Could not list output directory: {e}")
    else:
        if has_enriched_csv or has_enriched_xlsx:
            st.success(f"✅ Scraping finished! Processed {total:,} website(s). Results saved as combined enriched CSV and Excel.")
        else:
            st.success(f"✅ Scraping finished! Processed {total:,} website(s). Generated {len(csv_files)} CSV file(s) and {len(excel_files)} Excel file(s).")
    
    # Failed URLs report - full list with error details and download
    if all_failed_urls:
        n_failed = len(all_failed_urls)
        # Group by error type for summary
        from collections import Counter
        def _error_type(msg):
            m = (msg or "").lower()
            if "timeout" in m or "exceeded" in m: return "Timeout"
            if "challenge" in m or "cloudflare" in m or "verification" in m: return "Challenge/verification page"
            if "rate limit" in m: return "Rate limit"
            if "redirect" in m: return "Redirect"
            if "insufficient" in m or "too short" in m: return "Insufficient content"
            if "blocked" in m or "403" in m or "429" in m: return "Blocked/HTTP error"
            if "ai" in m or "summary" in m or "email" in m: return "AI/email generation"
            return "Other"
        type_counts = Counter(_error_type(e.get("message", "")) for e in all_failed_urls)
        with st.expander(f"⚠️ Failed URLs ({n_failed:,}) — details and download", expanded=True):
            st.markdown(f"**Summary:** {n_failed:,} of {total:,} URLs failed.")
            st.markdown("**By error type:** " + ", ".join(f"{t}: {c}" for t, c in type_counts.most_common()))
            buf = StringIO()
            w = csv.writer(buf, quoting=csv.QUOTE_ALL)
            w.writerow(["Website", "Error"])
            for e in all_failed_urls:
                w.writerow([e.get("url", ""), (e.get("message", "") or "").strip()])
            failed_csv = buf.getvalue()
            st.download_button(
                "📥 Download failed URLs (CSV)",
                failed_csv,
                file_name=f"failed_urls_{run_folder}.csv",
                mime="text/csv",
                key="download_failed_urls"
            )
            timeout_count = type_counts.get("Timeout", 0)
            if timeout_count > n_failed * 0.5:
                st.warning(f"**Most failures are timeouts ({timeout_count}).** Try increasing **Wait time per site** in Step 2 Settings (e.g. 60–90s), or reduce **Parallel workers** to ease load.")
            st.caption("Use this list to retry failed URLs separately or increase timeout in Settings.")
            st.markdown("**Full list:**")
            display_limit = 50
            for i, e in enumerate(all_failed_urls[:display_limit]):
                u, m = e.get("url", ""), e.get("message", "")
                st.code(f"{u}\n  → {m}", language=None)
            if n_failed > display_limit:
                st.caption(f"_Showing first {display_limit} of {n_failed}. Download CSV above for full list._")
    
    # Store file paths and data in session_state for persistence across reruns
    st.session_state['scraping_complete'] = True
    st.session_state['output_dir'] = output_dir
    st.session_state['run_folder'] = run_folder
    st.session_state['zip_path'] = zip_path
    st.session_state['zip_name'] = zip_name
    st.session_state['csv_files'] = csv_files
    st.session_state['excel_files'] = excel_files
    st.session_state['total_processed'] = total
    
    # Read ZIP file data and store in session_state
    try:
        if os.path.exists(zip_path):
            with open(zip_path, 'rb') as f:
                st.session_state['zip_data'] = f.read()
        else:
            st.warning(f"⚠️ ZIP file not found at {zip_path}")
            st.session_state['zip_data'] = None
    except Exception as e:
        st.error(f"❌ Error reading ZIP file: {e}")
        st.session_state['zip_data'] = None
    
    # Show accuracy info for max_chars
    if total > 0:
        st.info(f"💡 **Accuracy:** Max characters limit ({max_chars:,} chars) was accurately enforced per website. Each website's content was limited to this exact amount.")
    
    # Store for the single download section below (no duplicate section)
    st.session_state['max_chars_info'] = max_chars
    

# Show download section if scraping was completed OR if we have download data
has_download_data = (
    st.session_state.get('zip_data') or 
    st.session_state.get('combined_excel_data') or 
    st.session_state.get('combined_csv_data')
)

# Single download section: show when scraping is complete or we have download data (no duplicate)
if st.session_state.get('scraping_complete', False) or has_download_data:
    # Retrieve data from session_state
    output_dir = st.session_state.get('output_dir')
    run_folder = st.session_state.get('run_folder')
    zip_path = st.session_state.get('zip_path')
    zip_name = st.session_state.get('zip_name')
    csv_files = st.session_state.get('csv_files', [])
    excel_files = st.session_state.get('excel_files', [])
    total = st.session_state.get('total_processed', 0)
    max_chars = st.session_state.get('max_chars_info', 10000)
    
    st.success(f"✅ Scraping completed! Processed {total:,} website(s).")
    
    if total > 0:
        st.info(f"💡 **Accuracy:** Max characters limit ({max_chars:,} chars) was accurately enforced per website.")
    
    # Large Dataset: Offer database-backed download option
    db = get_scrape_db()
    if db and total > 5000:
        st.info(f"📊 **Large Dataset Mode:** Results are stored in database. You can download partial results at any time.")
        
        # Get stats from database
        try:
            stats = db.get_stats()
            col_db1, col_db2, col_db3, col_db4 = st.columns(4)
            with col_db1:
                st.metric("Total", f"{stats['total']:,}")
            with col_db2:
                st.metric("Success", f"{stats['successes']:,}", f"{stats['successes']/stats['total']*100:.1f}%" if stats['total'] > 0 else "0%")
            with col_db3:
                st.metric("Partial", f"{stats['partials']:,}")
            with col_db4:
                st.metric("Failed", f"{stats['failures']:,}")
        except Exception:
            pass
    
    # Cloud mode warning - downloads may expire
    if is_cloud_mode():
        st.warning("⚠️ **Important:** Downloads are stored temporarily. If you refresh or leave this page, files may be lost. Download NOW to save your results.")
    
    # Download section (persistent, same anchor for styling)
    st.markdown("""<h2 class="download-section-anchor" style="color: #f1f5f9; font-size: 1.5rem; margin-bottom: 0.5rem;">📥 Download Results</h2>""", unsafe_allow_html=True)
    st.caption("**ZIP:** All broken-down part files. **Combined Excel:** One file with only Website URL, scraped content, company summary, and email copy (4 columns). **Combined CSV:** Your original uploaded file with the new data appended as columns — clean, no formatting errors.")
    col_dl1, col_dl2, col_dl3 = st.columns(3)
    
    with col_dl1:
        st.subheader("📦 All Files (ZIP)")
        zip_data = st.session_state.get('zip_data')
        if zip_data:
            st.download_button(
                label="⬇️ Download ZIP Archive",
                data=zip_data,
                file_name=zip_name or "results.zip",
                mime="application/zip",
                help="All broken-down CSV and Excel part files (output_part_*.csv, output_part_*.xlsx)",
                key="download_zip_main"
            )
        else:
            # Fallback: read from file
            try:
                if zip_path and os.path.exists(zip_path):
                    with open(zip_path, 'rb') as f:
                        zip_data = f.read()
                        st.session_state['zip_data'] = zip_data
                        st.download_button(
                            label="⬇️ Download ZIP Archive",
                            data=zip_data,
                            file_name=zip_name or "results.zip",
                            mime="application/zip",
                            help="All broken-down CSV and Excel part files (output_part_*.csv, output_part_*.xlsx)",
                            key="download_zip_fallback"
                        )
                else:
                    if is_cloud_mode():
                        st.error("❌ Files expired. In cloud mode, files are temporary. Please re-upload your CSV and run the scraper again.")
                    else:
                        st.info("ZIP file not available on disk")
            except Exception as e:
                st.error(f"Could not read ZIP file: {e}")
        st.caption(f"Contains {len(csv_files)} CSV + {len(excel_files)} Excel files")
    
    with col_dl2:
        st.subheader("📊 Combined Excel")
        excel_data = st.session_state.get('combined_excel_data')
        excel_filename = st.session_state.get('combined_excel_filename')
        combined_df = st.session_state.get('combined_df')
        
        if excel_data and excel_filename:
            st.download_button(
                label="⬇️ Download Combined Excel",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Website URL, scraped content, company summary, and email copy only (4 columns)",
                key="download_excel_main"
            )
            if combined_df is not None:
                st.caption(f"{len(combined_df)} total rows")
        elif output_dir and os.path.isfile(os.path.join(output_dir, "enriched_output.xlsx")) and os.path.getsize(os.path.join(output_dir, "enriched_output.xlsx")) > 0:
            # Enriched-only run (no part files): load from disk so download works after refresh
            try:
                _excel_path = os.path.join(output_dir, "enriched_output.xlsx")
                with open(_excel_path, 'rb') as f:
                    _data = f.read()
                st.session_state['combined_excel_data'] = _data
                st.session_state['combined_excel_filename'] = f"{run_folder}_combined.xlsx"
                st.download_button(
                    label="⬇️ Download Combined Excel",
                    data=_data,
                    file_name=f"{run_folder}_combined.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Website URL, scraped content, company summary, and email copy only (4 columns)",
                    key="download_excel_enriched"
                )
            except Exception as e:
                st.warning(f"Could not load Excel: {e}")
        elif excel_files:
            # Regenerate combined Excel file from existing Excel files
            if not output_dir:
                # Try to get output_dir from session_state
                output_dir = st.session_state.get('output_dir')
                if not output_dir:
                    st.warning("⚠️ Output directory not found. Please restart scraping to regenerate Excel file.")
                    output_dir = None
            
            if output_dir:
                # Prefer enriched output (original sheet + new columns) when present
                enriched_csv_path = os.path.join(output_dir, "enriched_output.csv")
                enriched_xlsx_path = os.path.join(output_dir, "enriched_output.xlsx")
                used_enriched = False
                if os.path.exists(enriched_csv_path) and os.path.getsize(enriched_csv_path) > 0:
                    try:
                        combined_df = pd.read_csv(enriched_csv_path, encoding='utf-8-sig')
                        if not combined_df.empty:
                            with open(enriched_csv_path, 'rb') as f:
                                st.session_state['combined_csv_data'] = f.read()
                            st.session_state['combined_csv_filename'] = f"{run_folder}_enriched.csv"
                            st.session_state['combined_df'] = combined_df
                            # Combined Excel: only Website, ScrapedText, CompanySummary, EmailCopy (4 columns)
                            excel_cols = [c for c in EXCEL_COLS_4 if c in combined_df.columns] or [c for c in EXCEL_COLS_3 if c in combined_df.columns]
                            combined_excel_df = combined_df[excel_cols].copy()
                            def clean_dataframe_for_excel(df):
                                import re
                                for col in df.columns:
                                    df[col] = df[col].astype(str).replace('nan', '').replace('None', '')
                                    df[col] = df[col].str.replace('\x00', '', regex=False)
                                    df[col] = df[col].str.replace(r'[\n\r\f\v\t]', ' ', regex=True)
                                    df[col] = df[col].str.replace(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', regex=True)
                                    df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
                                    df[col] = df[col].str.strip()
                                return df
                            combined_excel_df = clean_dataframe_for_excel(combined_excel_df)
                            from openpyxl import Workbook
                            excel_buffer = BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            _write_excel_sheet(ws, excel_cols, combined_excel_df)
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            st.session_state['combined_excel_data'] = excel_buffer.read()
                            st.session_state['combined_excel_filename'] = f"{run_folder}_combined.xlsx"
                            used_enriched = True
                    except Exception as e:
                        st.warning(f"Could not use enriched output: {e}")
                if not used_enriched:
                    st.info("🔄 Regenerating combined Excel file from existing files...")
                try:
                    if used_enriched:
                        st.download_button(
                            label="⬇️ Download Combined Excel",
                            data=st.session_state.get('combined_excel_data'),
                            file_name=st.session_state.get('combined_excel_filename', f"{run_folder}_combined.xlsx"),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Website URL, scraped content, company summary, and email copy only (4 columns)",
                            key="download_excel_regen"
                        )
                        st.caption(f"{len(combined_df)} total rows — Website, Scraped Text, Company Summary, Email Copy")
                    else:
                        # CRITICAL: Read from CSV files first (source of truth), NOT Excel files
                        csv_all_data = []
                        if csv_files:
                            for csv_file in sorted(csv_files):
                                if "combined" in csv_file.lower() or csv_file == "enriched_output.csv":
                                    continue
                                csv_path = os.path.join(output_dir, csv_file)
                                try:
                                    if os.path.exists(csv_path) and os.path.getsize(csv_path) > 0:
                                        rows_data = []
                                        with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
                                            csv_reader = csv.reader(f, quoting=csv.QUOTE_ALL, doublequote=True)
                                            header = next(csv_reader, None)
                                            if not header:
                                                continue
                                            header_normalized = [str(h).strip().strip('"').strip("'") for h in header]
                                            col_indices = {}
                                            for idx, h in enumerate(header_normalized):
                                                h_lower = h.lower().strip()
                                                if h_lower == 'website' or h_lower == 'url':
                                                    col_indices['Website'] = idx
                                                elif h_lower == 'scrapedtext' or (h_lower.startswith('scraped') and 'text' in h_lower):
                                                    col_indices['ScrapedText'] = idx
                                                elif h_lower == 'companysummary' or (h_lower.startswith('company') and 'summary' in h_lower):
                                                    col_indices['CompanySummary'] = idx
                                                elif 'email' in h_lower and 'copy' in h_lower:
                                                    col_indices['EmailCopy'] = idx
                                            if 'Website' not in col_indices and len(header_normalized) >= 1:
                                                col_indices['Website'] = 0
                                            if 'ScrapedText' not in col_indices and len(header_normalized) >= 2:
                                                col_indices['ScrapedText'] = 1
                                            if 'CompanySummary' not in col_indices and len(header_normalized) >= 3:
                                                col_indices['CompanySummary'] = 2
                                            if 'EmailCopy' not in col_indices and len(header_normalized) >= 4:
                                                col_indices['EmailCopy'] = 3
                                            if len(col_indices) < 3:
                                                continue
                                            use_cols = EXCEL_COLS_4 if 'EmailCopy' in col_indices else EXCEL_COLS_3
                                            num_cols = len(use_cols)
                                            for row in csv_reader:
                                                if len(row) == 0:
                                                    continue
                                                while len(row) < num_cols:
                                                    row.append("")
                                                if len(row) > num_cols:
                                                    row = row[:num_cols]
                                                website_idx = col_indices.get('Website', 0)
                                                scraped_idx = col_indices.get('ScrapedText', 1)
                                                summary_idx = col_indices.get('CompanySummary', 2)
                                                email_idx = col_indices.get('EmailCopy')
                                                if website_idx >= len(row) or scraped_idx >= len(row) or summary_idx >= len(row):
                                                    continue
                                                website = str(row[website_idx]).strip()
                                                scraped_text = str(row[scraped_idx]).strip()
                                                company_summary = str(row[summary_idx]).strip()
                                                email_copy = str(row[email_idx]).strip() if email_idx is not None and email_idx < len(row) else ""
                                                if website:
                                                    row_dict = {'Website': website, 'ScrapedText': scraped_text, 'CompanySummary': company_summary, 'EmailCopy': email_copy if use_cols == EXCEL_COLS_4 else ''}
                                                    rows_data.append(row_dict)
                                        if rows_data:
                                            df_csv = pd.DataFrame(rows_data, columns=EXCEL_COLS_4)
                                            csv_all_data.append(df_csv)
                                except Exception as e:
                                    st.warning(f"⚠️ Error reading CSV {csv_file}: {e}")
                                    continue
                    
                    # Use CSV data if available, otherwise fall back to Excel files
                    if csv_all_data:
                        combined_df = pd.concat(csv_all_data, ignore_index=True)
                        for c in EXCEL_COLS_4:
                            if c not in combined_df.columns:
                                combined_df[c] = ""
                        combined_df = combined_df[EXCEL_COLS_4]
                    elif excel_files:
                        # Fallback to Excel files
                        all_data = []
                        for excel_file in sorted(excel_files):
                            excel_path = os.path.join(output_dir, excel_file)
                            try:
                                if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                                    df_part = pd.read_excel(excel_path, engine='openpyxl', header=0)
                                    if df_part.empty:
                                        continue
                                    df_part.columns = [str(col).strip() for col in df_part.columns]
                                    column_mapping = {}
                                    for col in df_part.columns:
                                        col_lower = col.lower().strip()
                                        if 'website' in col_lower or col_lower == 'url':
                                            column_mapping[col] = 'Website'
                                        elif 'scraped' in col_lower and 'text' in col_lower:
                                            column_mapping[col] = 'ScrapedText'
                                        elif 'company' in col_lower and 'summary' in col_lower:
                                            column_mapping[col] = 'CompanySummary'
                                    df_part = df_part.rename(columns=column_mapping)
                                    if 'Website' not in df_part.columns and len(df_part.columns) >= 1:
                                        df_part.columns.values[0] = 'Website'
                                    if 'ScrapedText' not in df_part.columns and len(df_part.columns) >= 2:
                                        df_part.columns.values[1] = 'ScrapedText'
                                    if 'CompanySummary' not in df_part.columns and len(df_part.columns) >= 3:
                                        df_part.columns.values[2] = 'CompanySummary'
                                    if "Website" not in df_part.columns:
                                        continue
                                    if "ScrapedText" not in df_part.columns:
                                        if len(df_part.columns) >= 2:
                                            second_col = df_part.columns[1]
                                            df_part = df_part.rename(columns={second_col: 'ScrapedText'})
                                        else:
                                            df_part['ScrapedText'] = ""
                                    if "CompanySummary" not in df_part.columns:
                                        df_part["CompanySummary"] = ""
                                    if "EmailCopy" not in df_part.columns:
                                        df_part["EmailCopy"] = ""
                                    std_cols = [c for c in EXCEL_COLS_4 if c in df_part.columns] or EXCEL_COLS_3
                                    df_part = df_part[std_cols]
                                    for col in df_part.columns:
                                        df_part[col] = df_part[col].astype(str).replace('nan', '').replace('None', '')
                                    if len(df_part) > 0:
                                        all_data.append(df_part)
                            except Exception as e:
                                st.warning(f"⚠️ Error reading Excel {excel_file}: {e}")
                                continue
                        if all_data:
                            combined_df = pd.concat(all_data, ignore_index=True)
                            for c in EXCEL_COLS_3:
                                if c not in combined_df.columns:
                                    combined_df[c] = ""
                            if "EmailCopy" not in combined_df.columns:
                                combined_df["EmailCopy"] = ""
                            combined_cols = EXCEL_COLS_4 if "EmailCopy" in combined_df.columns else EXCEL_COLS_3
                            combined_df = combined_df[combined_cols]
                        else:
                            st.error("❌ No data available to combine")
                            combined_df = None
                    else:
                        st.error("❌ No data available to combine")
                        combined_df = None
                    
                    if combined_df is not None and len(combined_df) > 0:
                        # Clean DataFrame
                        def clean_dataframe_for_excel(df):
                            import re
                            for col in df.columns:
                                df[col] = df[col].astype(str).replace('nan', '').replace('None', '')
                                df[col] = df[col].str.replace('\x00', '', regex=False)
                                df[col] = df[col].str.replace(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', regex=True)
                                df[col] = df[col].str.strip()
                            return df
                        combined_df = clean_dataframe_for_excel(combined_df.copy())
                        
                        # Generate combined Excel file
                        from openpyxl import Workbook
                        from io import BytesIO
                        excel_buffer = BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        cols_out = [c for c in EXCEL_COLS_4 if c in combined_df.columns] or EXCEL_COLS_3
                        _write_excel_sheet(ws, cols_out, combined_df)
                        
                        # CRITICAL: Save with explicit error handling
                        try:
                            wb.save(excel_buffer)
                        except Exception as save_error:
                            st.error(f"❌ Excel save failed: {save_error}")
                            import traceback
                            st.error(f"Traceback: {traceback.format_exc()}")
                            raise save_error
                        
                        excel_buffer.seek(0)
                        excel_data = excel_buffer.read()
                        
                        # Store in session_state
                        st.session_state['combined_excel_data'] = excel_data
                        st.session_state['combined_excel_filename'] = f"{run_folder}_combined.xlsx"
                        st.session_state['combined_df'] = combined_df
                        
                        st.download_button(
                            label="⬇️ Download Combined Excel",
                            data=excel_data,
                            file_name=f"{run_folder}_combined.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Download all data in a single Excel file",
                            key="download_excel_regen"
                        )
                        st.caption(f"{len(combined_df)} total rows")
                    else:
                        st.warning("⚠️ No data found in Excel files to combine")
                except Exception as e:
                    st.error(f"❌ Error generating combined Excel file: {e}")
                    import traceback
                    st.error(f"Traceback: {traceback.format_exc()}")
                    st.info("Individual Excel files are available in the ZIP archive")
                    # Still show that Excel files exist
                    if excel_files:
                        st.info(f"Found {len(excel_files)} Excel file(s): {', '.join(excel_files[:3])}{'...' if len(excel_files) > 3 else ''}")
        elif excel_files:
            # Excel files exist but output_dir is missing - this shouldn't happen, but handle it
            st.warning(f"⚠️ Found {len(excel_files)} Excel file(s) but output directory is not set. Please restart scraping.")
        else:
            st.info("No Excel files generated")
    
    with col_dl3:
        st.subheader("📄 Combined CSV (original + new data)")
        csv_data = st.session_state.get('combined_csv_data')
        csv_filename = st.session_state.get('combined_csv_filename')
        combined_df = st.session_state.get('combined_df')
        
        # Check for database-backed large dataset download
        db = get_scrape_db()
        if db and st.session_state.get('total_processed', 0) > 5000:
            # Offer database export for large datasets
            if st.button("📊 Export from Database (Large Dataset)", key="export_db_csv", type="secondary"):
                with st.spinner("Exporting from database..."):
                    try:
                        import io
                        csv_buffer = io.StringIO()
                        db.export_to_csv(csv_buffer, chunk_size=5000)
                        csv_data = csv_buffer.getvalue().encode('utf-8-sig')
                        st.download_button(
                            label="⬇️ Download Database Export",
                            data=csv_data,
                            file_name=f"{run_folder or 'results'}_database_export.csv",
                            mime="text/csv",
                            help="Direct export from database for large datasets",
                            key="download_db_export"
                        )
                    except Exception as e:
                        st.error(f"Database export failed: {e}")
        
        if csv_data and csv_filename:
            st.download_button(
                label="⬇️ Download Combined CSV",
                data=csv_data,
                file_name=csv_filename,
                mime="text/csv",
                help="Your original uploaded file with scraped content, company summary, and email copy appended as new columns. Clean formatting.",
                key="download_csv_main"
            )
            if combined_df is not None:
                st.caption(f"{len(combined_df)} total rows")
        elif output_dir and os.path.isdir(output_dir):
            # Restore combined CSV from disk (enriched_output.csv or part files) so download works after refresh
            try:
                enriched_path = os.path.join(output_dir, "enriched_output.csv")
                if os.path.isfile(enriched_path) and os.path.getsize(enriched_path) > 0:
                    with open(enriched_path, 'rb') as f:
                        csv_data = f.read()
                    st.session_state['combined_csv_data'] = csv_data
                    st.session_state['combined_csv_filename'] = f"{run_folder}_enriched.csv"
                    combined_df = pd.read_csv(enriched_path, encoding='utf-8-sig')
                    st.session_state['combined_df'] = combined_df
                else:
                    dfs = []
                    for f in sorted(csv_files):
                        if "combined" in f.lower() or f == "enriched_output.csv":
                            continue
                        fp = os.path.join(output_dir, f)
                        if os.path.isfile(fp) and f.endswith(".csv"):
                            try:
                                df = pd.read_csv(fp, encoding='utf-8-sig')
                                if len(df) > 0:
                                    dfs.append(df)
                            except Exception:
                                pass
                    if dfs:
                        combined_df = pd.concat(dfs, ignore_index=True)
                        csv_buffer = StringIO()
                        combined_df.to_csv(csv_buffer, index=False, encoding='utf-8-sig', quoting=csv.QUOTE_ALL)
                        csv_data = csv_buffer.getvalue().encode('utf-8-sig')
                        st.session_state['combined_csv_data'] = csv_data
                        st.session_state['combined_csv_filename'] = f"{run_folder}_combined.csv"
                        st.session_state['combined_df'] = combined_df
                csv_data = st.session_state.get('combined_csv_data')
                csv_filename = st.session_state.get('combined_csv_filename')
                combined_df = st.session_state.get('combined_df')
                if csv_data and csv_filename:
                    st.download_button(
                        label="⬇️ Download Combined CSV",
                        data=csv_data,
                        file_name=csv_filename,
                        mime="text/csv",
                        help="Download all data in a single CSV file (Excel/Google Sheets compatible)",
                        key="download_csv_regen"
                    )
                    if combined_df is not None:
                        st.caption(f"{len(combined_df)} total rows")
                else:
                    st.info("Could not regenerate CSV from files. Use the ZIP to get individual files.")
            except Exception as e:
                st.warning(f"Could not load CSV from disk: {e}. Use the ZIP to get individual files.")
        elif csv_files:
            st.info("Output folder not found. Use **Resume or Download Partial Results** above, or re-run to get CSV.")
        else:
            st.info("No CSV files generated")
    
    
    # File list (include enriched_output.* when present, since enriched-only runs have no part files)
    with st.expander("📋 View Generated Files", expanded=False):
        csv_to_show = list(csv_files)
        excel_to_show = list(excel_files)
        if output_dir:
            if os.path.isfile(os.path.join(output_dir, "enriched_output.csv")):
                if "enriched_output.csv" not in csv_to_show:
                    csv_to_show.append("enriched_output.csv")
            if os.path.isfile(os.path.join(output_dir, "enriched_output.xlsx")):
                if "enriched_output.xlsx" not in excel_to_show:
                    excel_to_show.append("enriched_output.xlsx")
        st.write("**CSV Files:**")
        for f in sorted(csv_to_show):
            st.code(f, language=None)
        st.write("**Excel Files:**")
        for f in sorted(excel_to_show):
            st.code(f, language=None)
        if output_dir:
            st.write(f"**Location:** `{output_dir}`")
    
    st.info("💡 **Tip:** CSV files use UTF-8 encoding with BOM for Excel compatibility. Excel files (.xlsx) are ready to open directly!")
    
    # Option to clear results
    if st.button("🗑️ Clear Results", help="Clear download buttons and start fresh", key="clear_results"):
        st.session_state['scraping_complete'] = False
        st.session_state.pop('output_dir', None)
        st.session_state.pop('zip_data', None)
        st.session_state.pop('combined_csv_data', None)
        st.session_state.pop('combined_excel_data', None)
        st.session_state.pop('combined_df', None)
        st.rerun()
